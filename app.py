"""
Robo UpSeller — app.py
Servidor Flask local. Rode com: python3 app.py
Acesse em:  http://127.0.0.1:5001
"""

import asyncio
import json
import os
import platform
import subprocess
import sys
import threading
import queue
import time
import urllib.request
import webbrowser
from datetime import date, timedelta, datetime
from pathlib import Path

from flask import Flask, render_template, request, jsonify, Response, make_response
from supabase import create_client

# ── Configurações ─────────────────────────────────────────────────────────────
SUPABASE_URL = "https://qaqlaqlxxeilouvbwgiv.supabase.co"
SUPABASE_KEY         = "sb_publishable_D0C2IC4Cxmtmu2crnFYXxw_JggkMuNS"
SUPABASE_SERVICE_KEY = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6InFhcWxhcWx4eGVpbG91dmJ3Z2l2Iiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NzAyNTQ1MSwiZXhwIjoyMDkyNjAxNDUxfQ.X8wGOgOsxxoUlCAh5XYdpPtzhMyUnKykiyFYxlfu2mo"

def _supa() -> tuple[str, str]:
    """Retorna (url, key) — usa Supabase privado do cliente se configurado."""
    try:
        cfg = json.loads(CONFIG_FILE.read_text(encoding="utf-8")) if CONFIG_FILE.exists() else {}
        url = (cfg.get("supabase_url") or "").strip()
        key = (cfg.get("supabase_key") or "").strip()
        if url and key:
            return url, key
    except Exception:
        pass
    return SUPABASE_URL, SUPABASE_KEY

# Quando rodado via launcher (importlib), __file__ aponta para robo_app/app.py
# e o diretório de dados/config fica ao lado (BASE_DIR = pasta do exe)
PASTA_RAIZ      = Path(__file__).parent
CONFIG_FILE     = PASTA_RAIZ / "config.json"
PASTA_DADOS     = PASTA_RAIZ / "dados"
PASTA_SESSAO    = PASTA_RAIZ / "sessao"
PASTA_ETIQUETAS = PASTA_RAIZ / "etiquetas"
PASTA_PICKLISTS = PASTA_RAIZ / "picklists"
EXECUCOES_FILE  = PASTA_RAIZ / "execucoes.json"

# ── Flask app ─────────────────────────────────────────────────────────────────
_template_dir = PASTA_RAIZ / "templates"
app = Flask(__name__, template_folder=str(_template_dir))
app.config["TEMPLATES_AUTO_RELOAD"] = True

@app.after_request
def _add_cors_global(response):
    response.headers["Access-Control-Allow-Origin"]  = "*"
    response.headers["Access-Control-Allow-Headers"] = "Content-Type,Authorization"
    response.headers["Access-Control-Allow-Methods"] = "GET,POST,OPTIONS"
    return response

_ignorar_paths = {"/", "/status", "/log", "/log-stream", "/favicon.ico"}

@app.before_request
def _log_req():
    from flask import request
    p = request.path
    if p in _ignorar_paths or p.startswith("/static"):
        return
    log(f"[req] {request.method} {p} | origem={request.remote_addr}")

log_queue        = queue.Queue()
rodando          = False
_pos_import_ativo = False             # captura/NF-e/envio rodando após import (não bloqueia PCP)
_etapa_atual     = ""                 # etapa visível no PCP
_parar           = threading.Event()  # sinaliza parada solicitada pelo usuário
_captura_lock    = threading.Lock()   # evita duas capturas simultâneas
_capturar_skip   = set()              # pedidos com URL inválida nesta sessão (background ignora)
_ultima_rodada: dict = {}             # stats da última rodada — expostos via /status
_API_STATE_LABEL = {
    "allocate":        "Para Reservar",
    "pending_review":  "Para Reservar",
    "to_allocate":     "Para Reservar",
    "to_reserve":      "Para Reservar",
    "in_process":      "Para Imprimir",
    "invoice_pending": "Para Emitir",
    "to_ship":         "Para Enviar",
    "to_pickup":       "Para Retirada",
    "pickup":          "Para Retirada",
}


def _carregar_execucoes() -> set:
    """Carrega execuções do dia de um arquivo, evitando re-execução após reiniciar."""
    hoje = str(date.today())
    try:
        if EXECUCOES_FILE.exists():
            dados = json.loads(EXECUCOES_FILE.read_text(encoding="utf-8"))
            return {k for k in dados.get("chaves", []) if k.startswith(hoje)}
    except Exception:
        pass
    return set()


def _salvar_execucoes(chaves: set):
    try:
        EXECUCOES_FILE.write_text(json.dumps({"chaves": list(chaves)}, ensure_ascii=False), encoding="utf-8")
    except Exception:
        pass


_execucoes_hoje = _carregar_execucoes()


# ── Rota principal ────────────────────────────────────────────────────────────
@app.route("/")
def index():
    config = {}
    if CONFIG_FILE.exists():
        config = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
    return render_template("index.html", config=config)


# ── Salvar configuração ───────────────────────────────────────────────────────
@app.route("/salvar", methods=["POST", "OPTIONS"], provide_automatic_options=False)
def salvar():
    if request.method == "OPTIONS":
        return _cors_preflight()
    dados = request.json
    erros = []

    if not dados.get("token"):          erros.append("Token obrigatório")
    if not dados.get("upseller_email"): erros.append("E-mail obrigatório")
    if not dados.get("upseller_senha"): erros.append("Senha obrigatória")

    if erros:
        return jsonify({"ok": False, "erro": " | ".join(erros)})

    try:
        nome = validar_token(dados["token"])
    except SystemExit:
        return jsonify({"ok": False, "erro": "Token inválido ou não encontrado"})

    # Preserva campos não exibidos na UI (ncm_padrao, ncm_produtos, etc.)
    config_existente = {}
    if CONFIG_FILE.exists():
        try:
            config_existente = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass

    config = {
        "ncm_padrao": "6109.10.00",
        **config_existente,
        "token":           dados["token"].strip(),
        "upseller_email":  dados["upseller_email"].strip(),
        "upseller_senha":  dados["upseller_senha"].strip(),
        "horarios":           dados.get("horarios", []),
        "horarios_semanais":  dados.get("horarios_semanais", None),
        "etapas":          dados.get("etapas", {"importar": True, "picklist": True, "nfe": True, "envio": True}),
        "nome_impressora":         dados.get("nome_impressora", "").strip(),
        "auto_imprimir_etiquetas": bool(dados.get("auto_imprimir_etiquetas", False)),
    }
    CONFIG_FILE.write_text(json.dumps(config, indent=2, ensure_ascii=False), encoding="utf-8")
    return jsonify({"ok": True, "cliente": nome})


@app.route("/salvar-config-parcial", methods=["POST", "OPTIONS"], provide_automatic_options=False)
def salvar_config_parcial():
    """Salva campos específicos no config sem exigir token/email/senha."""
    if request.method == "OPTIONS":
        return _cors_preflight()
    dados = request.json or {}
    config = {}
    if CONFIG_FILE.exists():
        try:
            config = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        except Exception:
            pass
    for k, v in dados.items():
        config[k] = v
    # Se salvou horarios_semanais, mantém backward-compat gerando lista flat de todos os horários únicos
    if "horarios_semanais" in dados:
        todos: set = set()
        for v in (dados["horarios_semanais"].values()):
            if v:
                todos.update(v)
        config["horarios"] = sorted(todos)
    CONFIG_FILE.write_text(json.dumps(config, indent=2, ensure_ascii=False), encoding="utf-8")
    return _cors(jsonify({"ok": True}))


# ── Importação retroativa ─────────────────────────────────────────────────────
@app.route("/importar-retroativo", methods=["POST"])
def importar_retroativo():
    """Importa pedidos de N dias atrás para preencher lacunas no Supabase."""
    dados = request.json or {}
    dias  = int(dados.get("dias", 30))
    if not CONFIG_FILE.exists():
        return jsonify({"ok": False, "erro": "Configure o robô primeiro"})

    def _rodar():
        config = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
        asyncio.run(_backfill_historico(config, dias))

    import threading as _t
    _t.Thread(target=_rodar, daemon=True).start()
    return jsonify({"ok": True, "msg": f"Backfill iniciado — últimos {dias} dias — acompanhe no log"})


async def _backfill_historico(config: dict, dias: int):
    """Busca todos os pedidos dos últimos N dias no UpSeller e insere os faltantes no Supabase."""
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        log("[backfill] ❌ Playwright não instalado")
        return

    token = config.get("token", "")
    log(f"[backfill] 🔄 Importando pedidos dos últimos {dias} dias...")

    from datetime import date, timedelta
    data_ini = (date.today() - timedelta(days=dias)).strftime("%Y-%m-%d")
    data_fim = date.today().strftime("%Y-%m-%d")

    # Estados a varrer (inclui concluídos/enviados)
    ESTADOS = [
        "in_process",
        "invoice_pending",
        "to_ship",
        "shipped",
        "to_pickup",
    ]

    PLAT = {"shopee": "Shopee", "shein": "Shein", "mercado": "Mercado Livre",
            "tiktok": "TikTok", "kwai": "Kwai"}

    todos: list = []

    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            user_data_dir=str(PASTA_SESSAO),
            headless=True, args=["--window-size=1280,900"],
        )
        try:
            page = context.pages[0] if context.pages else await context.new_page()
            await page.goto("https://app.upseller.com/pt/order/in-process",
                            wait_until="domcontentloaded", timeout=60_000)
            await page.wait_for_timeout(1500)

            if "/login" in page.url:
                log("[backfill] ⚠️ Sessão expirada — faça login primeiro")
                return

            for estado in ESTADOS:
                page_num = 1
                total_estado = None
                log(f"[backfill] 🔍 {estado}...")
                while True:
                    body = (
                        f"timeType=1&startTime={data_ini}&endTime={data_fim}"
                        f"&searchType=0&sortName=1&sortValue=1"
                        f"&orderState={estado}&pageNum={page_num}&pageSize=50"
                    )
                    r = await page.evaluate(f"""async () => {{
                        try {{
                            const resp = await fetch('/api/order/index', {{
                                method:'POST',
                                headers:{{'Content-Type':'application/x-www-form-urlencoded'}},
                                body:'{body}'
                            }});
                            return await resp.json();
                        }} catch(e) {{ return null; }}
                    }}""")
                    data = ((r or {}).get("data") or {})
                    lista = data.get("list") or []
                    total = data.get("total") or 0
                    if total_estado is None:
                        total_estado = total
                        log(f"[backfill]   → {total} pedido(s) em {estado}")
                    for o in lista:
                        o["_state"] = estado
                    todos.extend(lista)
                    if page_num * 50 >= total_estado or not lista:
                        break
                    page_num += 1
        except Exception as e:
            log(f"[backfill] ❌ Erro ao consultar API: {e}")
        finally:
            await context.close()

    if not todos:
        log("[backfill] ℹ️ Nenhum pedido encontrado")
        return

    log(f"[backfill] 📊 {len(todos)} pedido(s) encontrados — verificando Supabase...")

    # Monta dicts
    pedidos_dict: dict = {}
    itens_list:   list = []
    for order in todos:
        on = (order.get("orderNumber") or "").strip()
        if not on:
            continue
        plataforma = PLAT.get((order.get("platform") or "").lower(), order.get("platform") or "")
        try:
            valor = float(order.get("orderAmount") or 0) or None
        except Exception:
            valor = None
        api_label_url = None
        for campo in ("labelUrl", "printUrl", "waybillUrl", "expressLabel", "packageLabel"):
            v = (order.get(campo) or "").strip()
            if v and v.startswith("http"):
                api_label_url = v
                break
        items = order.get("orderItemList") or []
        for item in items:
            itens_list.append({
                "order_number": on,
                "sku":          (item.get("variationSku") or item.get("productSku") or "").strip(),
                "product_name": (item.get("productName") or "").strip(),
                "image_url":    (item.get("productImg")  or "").strip(),
                "quantidade":   int(item.get("productCount") or 1),
                "cliente":      token,
                "data":         data_fim,
            })
        first     = items[0] if items else {}
        qtd_total = sum(int(i.get("productCount") or 1) for i in items) or 1
        # Usa payTime (data do pagamento) → createTime → fallback
        data_pedido = data_fim
        for campo_data in ("payTime", "createTime", "orderTime", "invoiceTime"):
            v = (order.get(campo_data) or "")[:10]  # "YYYY-MM-DD"
            if len(v) == 10 and v[4] == "-":
                data_pedido = v
                break

        if on not in pedidos_dict:
            nome_cliente = (
                order.get("receiverName") or order.get("buyerName") or
                order.get("recipientName") or order.get("receiveName") or
                order.get("receiver", {}).get("name") or ""
            ).strip() or None
            if nome_cliente is None and len(pedidos_dict) == 0:
                log(f"[backfill] 🔍 campos do 1º pedido: {sorted(order.keys())}")
            pedidos_dict[on] = {
                "order_number":      on,
                "numero_plataforma": (order.get("orderId") or order.get("extendedId") or "").strip() or None,
                "sku":               (first.get("variationSku") or first.get("productSku") or "").strip(),
                "product_name":      (first.get("productName") or "").strip(),
                "image_url":         (first.get("productImg")  or "").strip(),
                "data":              data_pedido,
                "quantidade":        qtd_total,
                "plataforma":        plataforma,
                "valor":             valor,
                "cliente":           token,
                "label_url":         api_label_url,
                "nome_cliente":      nome_cliente,
            }

    # Filtra só os que não estão no Supabase
    supa = create_client(*_supa())
    nums = list(pedidos_dict.keys())
    existentes: set = set()
    for i in range(0, len(nums), 100):
        lote = nums[i:i + 100]
        r = supa.table("pedidos").select("order_number").in_("order_number", lote).execute()
        existentes.update(row["order_number"] for row in (r.data or []))

    faltando = [pedidos_dict[n] for n in nums if n not in existentes]
    log(f"[backfill] 📥 {len(faltando)} pedido(s) faltando no Supabase")

    if not faltando:
        log("[backfill] ✅ Supabase já está atualizado")
        return

    # Insere em lotes com tratamento de erro por coluna
    inseridos = 0
    erros = 0
    COLUNAS_OPCIONAIS = {"plataforma", "valor", "label_url", "numero_plataforma", "nome_cliente"}
    for i in range(0, len(faltando), 50):
        lote_orig = faltando[i:i + 50]
        excluir: set = set()
        lote = lote_orig
        while True:
            try:
                supa.table("pedidos").upsert(lote, on_conflict="order_number").execute()
                inseridos += len(lote)
                break
            except Exception as e:
                msg = str(e)
                col = next((c for c in COLUNAS_OPCIONAIS - excluir if c in msg), None)
                if col:
                    excluir.add(col)
                    lote = [{k: v for k, v in row.items() if k not in excluir} for row in lote_orig]
                else:
                    log(f"[backfill] ⚠️ Erro insert lote {i//50+1}: {msg[:120]}")
                    erros += len(lote_orig)
                    break

    # Insere itens dos faltando
    ons_faltando = {p["order_number"] for p in faltando}
    itens_inserir = [it for it in itens_list if it["order_number"] in ons_faltando]
    for i in range(0, len(itens_inserir), 50):
        try:
            supa.table("pedido_itens").insert(itens_inserir[i:i + 50]).execute()
        except Exception:
            pass

    log(f"[backfill] ✅ {inseridos} pedido(s) importados! ({len(itens_inserir)} itens){f' | {erros} erros' if erros else ''}")


# ── Supabase privado: testar conexão ─────────────────────────────────────────
@app.route("/testar-supabase", methods=["POST"])
def testar_supabase():
    dados = request.json or {}
    url = (dados.get("supabase_url") or "").strip()
    key = (dados.get("supabase_key") or "").strip()
    if not url or not key:
        return jsonify({"ok": False, "erro": "URL e chave são obrigatórios"})
    try:
        sb = create_client(url, key)
        sb.table("pedidos").select("order_number").limit(1).execute()
        return jsonify({"ok": True, "msg": "✅ Conexão bem-sucedida!"})
    except Exception as e:
        msg = str(e)
        if "does not exist" in msg or "relation" in msg:
            return jsonify({"ok": True, "msg": "✅ Conectado — tabelas ainda não criadas (use Migrar)"})
        return jsonify({"ok": False, "erro": f"❌ {msg}"})


# ── Supabase privado: migrar tabelas e dados ──────────────────────────────────
@app.route("/migrar-supabase", methods=["POST"])
def migrar_supabase():
    dados = request.json or {}
    url_destino = (dados.get("supabase_url") or "").strip()
    key_destino = (dados.get("supabase_key") or "").strip()
    if not url_destino or not key_destino:
        return jsonify({"ok": False, "erro": "URL e chave do Supabase privado são obrigatórios"})

    def _rodar():
        try:
            sb_orig = create_client(SUPABASE_URL, SUPABASE_KEY)
            sb_dest = create_client(url_destino, key_destino)

            # 1. Cria tabelas via RPC (SQL direto) —————————————————————————
            SQL_SETUP = """
CREATE TABLE IF NOT EXISTS pedidos (
    id               BIGSERIAL PRIMARY KEY,
    order_number     TEXT UNIQUE NOT NULL,
    numero_plataforma TEXT,
    sku              TEXT,
    product_name     TEXT,
    image_url        TEXT,
    data             DATE,
    criado_em        TIMESTAMPTZ DEFAULT NOW(),
    cliente          TEXT,
    quantidade       INTEGER DEFAULT 1,
    plataforma       TEXT,
    valor            NUMERIC,
    label_url        TEXT
);
ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS numero_plataforma TEXT;
ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS nome_cliente TEXT;

CREATE TABLE IF NOT EXISTS pedido_itens (
    id           BIGSERIAL PRIMARY KEY,
    order_number TEXT NOT NULL,
    sku          TEXT,
    product_name TEXT,
    image_url    TEXT,
    quantidade   INTEGER DEFAULT 1,
    cliente      TEXT,
    data         DATE,
    criado_em    TIMESTAMPTZ DEFAULT NOW()
);
"""
            try:
                sb_dest.rpc("exec_sql", {"sql": SQL_SETUP}).execute()
                log("[migrar] ✅ Tabelas criadas/verificadas")
            except Exception:
                log("[migrar] ℹ️ RPC exec_sql indisponível — crie as tabelas manualmente pelo painel Supabase")

            # 2. Copia pedidos —————————————————————————————————————————————
            log("[migrar] 📋 Copiando pedidos...")
            offset, total = 0, 0
            while True:
                r = sb_orig.table("pedidos").select("*").range(offset, offset + 199).execute()
                if not r.data:
                    break
                lote = [{k: v for k, v in row.items() if k != "id"} for row in r.data]
                sb_dest.table("pedidos").upsert(lote, on_conflict="order_number").execute()
                total += len(lote)
                offset += 200
                if len(r.data) < 200:
                    break
            log(f"[migrar] ✅ {total} pedido(s) copiados")

            # 3. Copia pedido_itens ————————————————————————————————————————
            log("[migrar] 📋 Copiando itens...")
            offset, total_i = 0, 0
            while True:
                r = sb_orig.table("pedido_itens").select("*").range(offset, offset + 199).execute()
                if not r.data:
                    break
                lote = [{k: v for k, v in row.items() if k != "id"} for row in r.data]
                sb_dest.table("pedido_itens").insert(lote).execute()
                total_i += len(lote)
                offset += 200
                if len(r.data) < 200:
                    break
            log(f"[migrar] ✅ {total_i} item(s) copiados")

            # 4. Salva no config como Supabase ativo ——————————————————————
            cfg = json.loads(CONFIG_FILE.read_text(encoding="utf-8")) if CONFIG_FILE.exists() else {}
            cfg["supabase_url"] = url_destino
            cfg["supabase_key"] = key_destino
            CONFIG_FILE.write_text(json.dumps(cfg, indent=2, ensure_ascii=False), encoding="utf-8")
            log("[migrar] ✅ Supabase privado ativado no config")

        except Exception as e:
            log(f"[migrar] ❌ Erro: {e}")

    import threading as _t
    _t.Thread(target=_rodar, daemon=True).start()
    return jsonify({"ok": True, "msg": "Migração iniciada — acompanhe no log"})


# ── Executar agora ────────────────────────────────────────────────────────────
def _cors(r):
    r.headers["Access-Control-Allow-Origin"]  = "*"
    r.headers["Access-Control-Allow-Methods"] = "GET, POST, OPTIONS"
    r.headers["Access-Control-Allow-Headers"] = "Content-Type"
    return r

def _cors_preflight():
    return _cors(make_response("", 204))

@app.route("/executar", methods=["POST", "OPTIONS"], provide_automatic_options=False)
def executar():
    if request.method == "OPTIONS":
        return _cors_preflight()
    global rodando
    if rodando or _pos_import_ativo:
        return _cors(jsonify({"ok": False, "erro": "Já está rodando, aguarde..."}))

    if not CONFIG_FILE.exists():
        return _cors(jsonify({"ok": False, "erro": "Salve a configuração primeiro"}))

    config = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
    modo   = request.json.get("modo", "hoje")
    agora  = datetime.now()
    ontem  = date.today() - timedelta(days=1)

    if modo == "ambos":
        ontem_inicio = datetime.combine(ontem, datetime.min.time())
        ontem_fim    = datetime.combine(ontem, datetime.max.time()).replace(microsecond=0)
        hoje_inicio  = agora.replace(hour=0, minute=0, second=0, microsecond=0)
        hoje_fim     = agora.replace(microsecond=0)
        threading.Thread(
            target=_rodar_em_thread_duplo,
            args=(config, ontem_inicio, ontem_fim, hoje_inicio, hoje_fim),
            daemon=True
        ).start()
    elif modo == "ontem":
        data_alvo_str = request.json.get("data_alvo", "")
        try:
            alvo = date.fromisoformat(data_alvo_str) if data_alvo_str else ontem
        except ValueError:
            alvo = ontem
        data_inicio = datetime.combine(alvo, datetime.min.time())
        data_fim    = datetime.combine(alvo, datetime.max.time()).replace(microsecond=0)
        threading.Thread(
            target=_rodar_em_thread,
            args=(config, data_inicio, data_fim),
            daemon=True
        ).start()
    elif modo == "datas":
        datas = request.json.get("datas", [])
        if not datas:
            return _cors(jsonify({"ok": False, "erro": "Nenhuma data informada"}))
        def _rodar_multiplas(cfg, lista_datas):
            for ds in lista_datas:
                try:
                    alvo = date.fromisoformat(ds)
                except ValueError:
                    continue
                ini = datetime.combine(alvo, datetime.min.time())
                fim = datetime.combine(alvo, datetime.max.time()).replace(microsecond=0)
                _rodar_em_thread(cfg, ini, fim)
        threading.Thread(target=_rodar_multiplas, args=(config, datas), daemon=True).start()
    else:  # hoje
        data_inicio = agora.replace(hour=0, minute=0, second=0, microsecond=0)
        data_fim    = agora.replace(microsecond=0)
        threading.Thread(
            target=_rodar_em_thread,
            args=(config, data_inicio, data_fim),
            daemon=True
        ).start()

    return _cors(jsonify({"ok": True}))


# ── Login UpSeller (abre browser para autenticação manual) ────────────────────
@app.route("/login_upseller", methods=["POST", "OPTIONS"], provide_automatic_options=False)
def login_upseller():
    if request.method == "OPTIONS":
        return _cors_preflight()
    global rodando
    if rodando:
        return jsonify({"ok": False, "erro": "Robô em execução, aguarde"})
    if not CONFIG_FILE.exists():
        return jsonify({"ok": False, "erro": "Salve a configuração primeiro"})

    config = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
    threading.Thread(
        target=lambda: asyncio.run(_login_upseller_playwright(config)),
        daemon=True
    ).start()
    return jsonify({"ok": True})


# ── Emitir etiqueta no UpSeller ───────────────────────────────────────────────
@app.route("/emitir-etiqueta", methods=["POST", "OPTIONS"])
def emitir_etiqueta():
    # CORS preflight (chamado do dashboard siteadnsys.vercel.app)
    if request.method == "OPTIONS":
        resp = jsonify({})
        resp.headers["Access-Control-Allow-Origin"]  = "*"
        resp.headers["Access-Control-Allow-Methods"] = "POST"
        resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
        return resp

    dados        = request.json or {}
    order_number = dados.get("order_number", "").strip()

    if not order_number:
        r = jsonify({"ok": False, "erro": "order_number obrigatório"})
        r.headers["Access-Control-Allow-Origin"] = "*"
        return r

    if not CONFIG_FILE.exists():
        r = jsonify({"ok": False, "erro": "Configure o robô primeiro"})
        r.headers["Access-Control-Allow-Origin"] = "*"
        return r

    config = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))

    # Roda de forma síncrona para retornar o resultado ao dashboard
    resultado = [{"ok": False, "erro": "Timeout"}]
    concluido = threading.Event()

    def run():
        resultado[0] = asyncio.run(_emitir_etiqueta_playwright(config, order_number))
        concluido.set()

    threading.Thread(target=run, daemon=True).start()
    concluido.wait(timeout=90)

    r = jsonify(resultado[0])
    r.headers["Access-Control-Allow-Origin"] = "*"
    return r


# ── Reimprimir etiqueta/NF-e no UpSeller ─────────────────────────────────────
@app.route("/reimprimir-etiqueta", methods=["POST", "OPTIONS"])
def reimprimir_etiqueta():
    if request.method == "OPTIONS":
        resp = jsonify({})
        resp.headers["Access-Control-Allow-Origin"]  = "*"
        resp.headers["Access-Control-Allow-Methods"] = "POST"
        resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
        return resp

    dados        = request.json or {}
    order_number = dados.get("order_number", "").strip()

    if not order_number:
        r = jsonify({"ok": False, "erro": "order_number obrigatório"})
        r.headers["Access-Control-Allow-Origin"] = "*"
        return r

    if not CONFIG_FILE.exists():
        r = jsonify({"ok": False, "erro": "Configure o robô primeiro"})
        r.headers["Access-Control-Allow-Origin"] = "*"
        return r

    config = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))

    resultado = [{"ok": False, "erro": "Timeout"}]
    concluido = threading.Event()

    def run():
        resultado[0] = asyncio.run(_reimprimir_etiqueta_playwright(config, order_number))
        concluido.set()

    threading.Thread(target=run, daemon=True).start()
    concluido.wait(timeout=90)

    r = jsonify(resultado[0])
    r.headers["Access-Control-Allow-Origin"] = "*"
    return r


# ── Capturar etiquetas (pré-download manual) ─────────────────────────────────
@app.route("/capturar-etiquetas", methods=["POST", "OPTIONS"])
def capturar_etiquetas():
    if request.method == "OPTIONS":
        resp = jsonify({})
        resp.headers["Access-Control-Allow-Origin"]  = "*"
        resp.headers["Access-Control-Allow-Methods"] = "POST"
        resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
        return resp

    if not CONFIG_FILE.exists():
        r = jsonify({"ok": False, "erro": "Configure o robô primeiro"})
        r.headers["Access-Control-Allow-Origin"] = "*"
        return r

    config = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
    threading.Thread(
        target=lambda: asyncio.run(_capturar_etiquetas_playwright(config)),
        daemon=True
    ).start()

    r = jsonify({"ok": True, "msg": "Captura iniciada — acompanhe nos logs"})
    r.headers["Access-Control-Allow-Origin"] = "*"
    return r


# ── Configurar inicialização automática ──────────────────────────────────────
@app.route("/configurar_inicializacao", methods=["POST", "OPTIONS"], provide_automatic_options=False)
def configurar_inicializacao():
    if request.method == "OPTIONS":
        return _cors_preflight()
    if not CONFIG_FILE.exists():
        return jsonify({"ok": False, "erro": "Salve a configuração primeiro"})

    config   = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
    horarios = sorted(config.get("horarios", ["07:00"]))
    sistema  = platform.system()

    if sistema == "Darwin":
        return jsonify(_setup_mac(horarios))
    elif sistema == "Windows":
        return jsonify(_setup_windows(horarios))
    else:
        return jsonify({"ok": False, "erro": f"Sistema não suportado: {sistema}"})


def _wake_time(horario: str) -> str:
    """Retorna HH:MM:SS com 1 minuto de antecedência."""
    h, m = map(int, horario.split(":"))
    m -= 1
    if m < 0:
        m, h = 59, h - 1
    if h < 0:
        h = 23
    return f"{h:02d}:{m:02d}:00"


def _setup_mac(horarios: list) -> dict:
    python_path = sys.executable
    app_path    = str(PASTA_RAIZ / "app.py")
    plist_dir   = Path.home() / "Library" / "LaunchAgents"
    plist_path  = plist_dir / "com.adnsys.roboUpseller.plist"

    plist_dir.mkdir(parents=True, exist_ok=True)
    plist_path.write_text(f"""<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE plist PUBLIC "-//Apple//DTD PLIST 1.0//EN" "http://www.apple.com/DTDs/PropertyList-1.0.dtd">
<plist version="1.0">
<dict>
    <key>Label</key><string>com.adnsys.roboUpseller</string>
    <key>ProgramArguments</key>
    <array>
        <string>{python_path}</string>
        <string>{app_path}</string>
    </array>
    <key>RunAtLoad</key><true/>
    <key>KeepAlive</key><true/>
    <key>WorkingDirectory</key><string>{str(PASTA_RAIZ)}</string>
    <key>StandardOutPath</key><string>{str(PASTA_RAIZ / "robo.log")}</string>
    <key>StandardErrorPath</key><string>{str(PASTA_RAIZ / "robo.log")}</string>
</dict>
</plist>""")

    # Carrega o LaunchAgent (não precisa de sudo)
    subprocess.run(["launchctl", "unload", str(plist_path)], capture_output=True)
    r = subprocess.run(["launchctl", "load",   str(plist_path)], capture_output=True)
    launchagent_ok = r.returncode == 0

    # pmset — precisa de sudo, então gera o comando para o usuário copiar
    wake = _wake_time(horarios[0])
    pmset_cmd = f"sudo pmset repeat wake MTWRFSU {wake}"

    # Tenta rodar sem sudo (funciona quando já tem permissão prévia)
    r2 = subprocess.run(
        ["pmset", "repeat", "wake", "MTWRFSU", wake],
        capture_output=True, text=True
    )
    pmset_ok = r2.returncode == 0

    return {
        "ok":           True,
        "sistema":      "mac",
        "launchagent":  launchagent_ok,
        "pmset_ok":     pmset_ok,
        "pmset_cmd":    pmset_cmd,
        "wake_time":    wake,
    }


def _setup_windows(horarios: list) -> dict:
    pythonw = Path(sys.executable).parent / "pythonw.exe"
    if not pythonw.exists():
        pythonw = Path(sys.executable)
    app_path = str(PASTA_RAIZ / "app.py")

    # Tarefa de inicialização no login
    task_xml = f"""<?xml version="1.0" encoding="UTF-16"?>
<Task version="1.2" xmlns="http://schemas.microsoft.com/windows/2004/02/mit/task">
  <RegistrationInfo><Description>AdnSys Robo UpSeller</Description></RegistrationInfo>
  <Triggers>
    <LogonTrigger><Enabled>true</Enabled></LogonTrigger>
  </Triggers>
  <Settings>
    <MultipleInstancesPolicy>IgnoreNew</MultipleInstancesPolicy>
    <ExecutionTimeLimit>PT0S</ExecutionTimeLimit>
    <WakeToRun>false</WakeToRun>
  </Settings>
  <Actions Context="Author">
    <Exec>
      <Command>{pythonw}</Command>
      <Arguments>"{app_path}"</Arguments>
      <WorkingDirectory>{str(PASTA_RAIZ)}</WorkingDirectory>
    </Exec>
  </Actions>
</Task>"""

    xml_path = PASTA_RAIZ / "robo_inicializacao.xml"
    xml_path.write_text(task_xml, encoding="utf-16")
    r = subprocess.run(
        ["schtasks", "/Create", "/TN", "AdnSysRoboUpSeller", "/XML", str(xml_path), "/F"],
        capture_output=True, text=True
    )
    login_ok = r.returncode == 0

    # Tarefa de wake para cada horário agendado
    wake_tasks = []
    for h in horarios:
        wake   = _wake_time(h)
        hh, mm = wake[:2], wake[3:5]
        name   = f"AdnSysRoboWake_{h.replace(':','')}"
        wake_xml = f"""<?xml version="1.0" encoding="UTF-16"?>
<Task version="1.2" xmlns="http://schemas.microsoft.com/windows/2004/02/mit/task">
  <RegistrationInfo><Description>AdnSys Wake {h}</Description></RegistrationInfo>
  <Triggers>
    <CalendarTrigger>
      <StartBoundary>2000-01-01T{hh}:{mm}:00</StartBoundary>
      <ScheduleByDay><DaysInterval>1</DaysInterval></ScheduleByDay>
      <Enabled>true</Enabled>
    </CalendarTrigger>
  </Triggers>
  <Settings>
    <WakeToRun>true</WakeToRun>
    <ExecutionTimeLimit>PT1M</ExecutionTimeLimit>
  </Settings>
  <Actions Context="Author">
    <Exec><Command>cmd.exe</Command><Arguments>/c echo wake</Arguments></Exec>
  </Actions>
</Task>"""
        wx_path = PASTA_RAIZ / f"robo_wake_{h.replace(':','')}.xml"
        wx_path.write_text(wake_xml, encoding="utf-16")
        r2 = subprocess.run(
            ["schtasks", "/Create", "/TN", name, "/XML", str(wx_path), "/F"],
            capture_output=True, text=True
        )
        wake_tasks.append({"horario": h, "wake": wake, "ok": r2.returncode == 0})

    return {
        "ok":         True,
        "sistema":    "windows",
        "login_ok":   login_ok,
        "wake_tasks": wake_tasks,
    }


# ── Stream de logs (SSE) ──────────────────────────────────────────────────────
@app.route("/logs")
def logs():
    def stream():
        while True:
            try:
                msg = log_queue.get(timeout=30)
                yield f"data: {msg}\n\n"
            except queue.Empty:
                yield "data: __ping__\n\n"
    return Response(
        stream(),
        mimetype="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        }
    )


# ── Flag rápido (sem Supabase) ────────────────────────────────────────────────
@app.route("/rodando")
def rodando_flag():
    return jsonify({"rodando": rodando})


# ── Status atual ──────────────────────────────────────────────────────────────
@app.route("/status")
def status():
    config = {}
    if CONFIG_FILE.exists():
        config = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))

    cliente = ""
    if config.get("token"):
        try:
            cliente = validar_token(config["token"])
        except Exception:
            cliente = "Token inválido"

    _DIA_KEYS = ["seg", "ter", "qua", "qui", "sex", "sab", "dom"]
    _dia_hoje = _DIA_KEYS[datetime.now().weekday()]
    horarios_semanais = config.get("horarios_semanais", {})
    if horarios_semanais:
        horarios = horarios_semanais.get(_dia_hoje) or []
    else:
        horarios = config.get("horarios", [])
        if not horarios and config.get("hora_inicio"):
            horarios = _gerar_horarios(
                config.get("hora_inicio", "07:00"),
                config.get("hora_fim", "18:00"),
                int(config.get("intervalo_horas", 1)),
            )
    etapas  = config.get("etapas", {"importar": True, "picklist": True, "nfe": True, "envio": True})
    agora   = datetime.now().strftime("%H:%M")
    proxima = next((h for h in sorted(horarios) if h > agora), horarios[0] if horarios else "—")

    # Verifica sessão real: checa se o cookie de autenticação existe e não expirou
    sessao_ok = False
    if PASTA_SESSAO.exists():
        try:
            import sqlite3 as _sq3, time as _time
            db = PASTA_SESSAO / "Default" / "Cookies"
            if db.exists():
                con = _sq3.connect(str(db))
                agr = _time.time()
                # Cookies do upseller.com com expires_utc > agora (Chrome armazena em microssegundos desde 1601)
                chrome_epoch = 11644473600  # segundos entre 1601-01-01 e 1970-01-01
                agr_chrome = (agr + chrome_epoch) * 1_000_000
                rows = con.execute(
                    "SELECT COUNT(*) FROM cookies WHERE host_key LIKE '%upseller%' AND expires_utc > ?",
                    (agr_chrome,)
                ).fetchone()
                con.close()
                sessao_ok = (rows[0] > 0) if rows else False
        except Exception:
            sessao_ok = PASTA_SESSAO.exists() and any(PASTA_SESSAO.iterdir())

    try:
        versao = json.loads((PASTA_RAIZ / "version.json").read_text())["version"]
    except Exception:
        versao = "?"

    return jsonify({
        "configurado":    bool(config.get("token")),
        "token":          config.get("token", ""),
        "cliente":        cliente,
        "email":          config.get("email", ""),
        "horarios":          horarios,
        "horarios_semanais": config.get("horarios_semanais"),
        "etapas":            etapas,
        "proxima":           proxima,
        "nome_impressora": config.get("nome_impressora", ""),
        "auto_imprimir":  config.get("auto_imprimir", False),
        "rodando":     rodando or _pos_import_ativo,
        "etapa":       _etapa_atual,
        "sessao_ok":   sessao_ok,
        "versao":      versao,
        "ultima_rodada": _ultima_rodada,
    })


# ── Controles do processo ─────────────────────────────────────────────────────

@app.route("/parar", methods=["POST", "OPTIONS"], provide_automatic_options=False)
def parar():
    if request.method == "OPTIONS":
        return _cors_preflight()
    global rodando
    _parar.set()
    rodando = False
    log("⏹️ Robô parado pelo usuário")
    return _cors(jsonify({"ok": True}))


@app.route("/recarregar-config", methods=["POST", "OPTIONS"], provide_automatic_options=False)
def recarregar_config():
    if request.method == "OPTIONS":
        return _cors_preflight()
    log("🔄 Configuração recarregada pelo PCP")
    return _cors(jsonify({"ok": True}))


@app.route("/reiniciar", methods=["POST"])
def reiniciar():
    def _restart():
        time.sleep(1.5)
        os.execv(sys.executable, [sys.executable, str(PASTA_RAIZ / "app.py")])
    threading.Thread(target=_restart, daemon=True).start()
    log("🔄 Reiniciando em 2s...")
    return jsonify({"ok": True})


@app.route("/verificar-cancelados", methods=["POST", "OPTIONS"], provide_automatic_options=False)
def rota_verificar_cancelados():
    """Verifica pedidos sem etiqueta no UpSeller e marca cancelados no Supabase."""
    if request.method == "OPTIONS":
        return _cors_preflight()
    if not CONFIG_FILE.exists():
        return _cors(jsonify({"ok": False, "erro": "Config não encontrada"})), 400
    cfg = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))

    def _run():
        try:
            import asyncio
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            loop.run_until_complete(_verificar_cancelados_playwright(cfg))
            loop.close()
        except Exception as e:
            log(f"[cancelados] ❌ Erro: {e}")

    threading.Thread(target=_run, daemon=True).start()
    return _cors(jsonify({"ok": True, "msg": "Verificação de cancelados iniciada — acompanhe o log"}))


@app.route("/resumo-etiquetas", methods=["GET", "OPTIONS"], provide_automatic_options=False)
def resumo_etiquetas():
    if request.method == "OPTIONS":
        return _cors_preflight()
    try:
        hoje  = datetime.now().strftime("%Y-%m-%d")
        ontem = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
        supa  = create_client(*_supa())

        def _resumo_dia(data: str) -> dict:
            rows = supa.table("pedidos") \
                .select("order_number,label_url,plataforma") \
                .eq("data", data).execute().data or []
            # deduplica por order_number (uma linha por SKU)
            vistos: dict = {}
            for r in rows:
                on = r.get("order_number")
                if on not in vistos:
                    vistos[on] = r
            total    = len(vistos)
            com_lbl  = sum(1 for r in vistos.values() if r.get("label_url"))
            sem_lbl  = total - com_lbl
            # breakdown por plataforma para os sem etiqueta
            sem_por_plat: dict = {}
            for r in vistos.values():
                if not r.get("label_url"):
                    plat = r.get("plataforma") or "Outros"
                    sem_por_plat[plat] = sem_por_plat.get(plat, 0) + 1
            return {"total": total, "com_etiqueta": com_lbl, "sem_etiqueta": sem_lbl, "sem_por_plataforma": sem_por_plat}

        return _cors(jsonify({
            "hoje":  _resumo_dia(hoje),
            "ontem": _resumo_dia(ontem),
        }))
    except Exception as e:
        return _cors(jsonify({"erro": str(e)}))


@app.route("/verificar-atualizacao", methods=["GET", "OPTIONS"], provide_automatic_options=False)
def verificar_atualizacao():
    if request.method == "OPTIONS":
        return _cors_preflight()
    BASE = "https://qaqlaqlxxeilouvbwgiv.supabase.co/storage/v1/object/public/robo-upseller"
    try:
        with urllib.request.urlopen(f"{BASE}/version.json", timeout=10) as r:
            disponivel = json.loads(r.read())["version"]
        atual = json.loads((PASTA_RAIZ / "version.json").read_text())["version"]
        def _ver(v):
            return tuple(int(x) for x in v.split("."))
        tem = _ver(disponivel) > _ver(atual)
        return _cors(jsonify({"atual": atual, "disponivel": disponivel, "tem_atualizacao": tem}))
    except Exception as e:
        try:
            atual = json.loads((PASTA_RAIZ / "version.json").read_text())["version"]
        except Exception:
            atual = "?"
        return _cors(jsonify({"atual": atual, "disponivel": atual, "tem_atualizacao": False, "erro": str(e)}))


@app.route("/atualizar", methods=["POST", "OPTIONS"], provide_automatic_options=False)
def atualizar():
    if request.method == "OPTIONS":
        return _cors_preflight()
    BASE = "https://qaqlaqlxxeilouvbwgiv.supabase.co/storage/v1/object/public/robo-upseller"
    ARQUIVOS_UPDATE = [
        (f"{BASE}/version.json",         PASTA_RAIZ / "version.json"),
        (f"{BASE}/app.py",               PASTA_RAIZ / "app.py"),
        (f"{BASE}/templates/index.html", PASTA_RAIZ / "templates" / "index.html"),
    ]
    def _update():
        try:
            log("📥 Verificando atualização...")
            with urllib.request.urlopen(f"{BASE}/version.json", timeout=10) as r:
                nova_versao = json.loads(r.read())["version"]
            versao_atual = json.loads((PASTA_RAIZ / "version.json").read_text())["version"]
            def _ver(v):
                return tuple(int(x) for x in v.split("."))
            if _ver(nova_versao) <= _ver(versao_atual):
                log(f"✅ Já está na versão mais recente ({versao_atual})")
                return
            log(f"⬆️ Atualizando {versao_atual} → {nova_versao}...")
            for url, dest in ARQUIVOS_UPDATE:
                with urllib.request.urlopen(url, timeout=30) as r:
                    dest.write_bytes(r.read())
                log(f"  ✓ {dest.name}")
            log("✅ Atualização concluída — reiniciando...")
            time.sleep(1.5)
            os.execv(sys.executable, [sys.executable, str(PASTA_RAIZ / "app.py")])
        except Exception as e:
            log(f"❌ Erro na atualização: {e}")
    threading.Thread(target=_update, daemon=True).start()
    return _cors(jsonify({"ok": True}))


# ── Helpers ───────────────────────────────────────────────────────────────────

def log(msg: str):
    timestamp = datetime.now().strftime("%H:%M:%S")
    linha = f"[{timestamp}] {msg}"
    log_queue.put(linha)
    try:
        with open(PASTA_RAIZ / "robo_debug.log", "a", encoding="utf-8") as f:
            f.write(linha + "\n")
    except Exception:
        pass


def validar_token(token: str) -> str:
    supa = create_client(*_supa())
    resp = supa.table("clientes") \
        .select("nome, ativo").eq("token", token).single().execute()
    if not resp.data:
        raise Exception("Token não encontrado")
    if not resp.data.get("ativo", True):
        raise Exception("Token desativado")
    return resp.data["nome"]


def atualizar_ultima_execucao(token: str):
    try:
        supa = create_client(*_supa())
        supa.table("clientes").update(
            {"ultima_execucao": datetime.utcnow().isoformat()}
        ).eq("token", token).execute()
    except Exception:
        pass


_ultimo_retro: float = 0.0  # timestamp da última captura retroativa
_INTERVALO_RETRO: float = 5 * 60  # retroativo a cada 5 minutos


def _executar_captura(config: dict, aguardar: bool = True, background: bool = False):
    """Roda captura com lock — impede duas execuções simultâneas.
    Background: pula se já rodando. Executar principal: espera até 60s o background ceder.
    """
    global _ultimo_retro
    if background:
        if not _captura_lock.acquire(blocking=False):
            log("[captura] ⏭️ Já em execução — pulando")
            return
    else:
        # Executar principal: aguarda o background ceder (ele verifica `rodando` a cada pedido)
        adquiriu = _captura_lock.acquire(timeout=60)
        if not adquiriu:
            log("[captura] ⚠️ Background ainda rodando após 60s — pulando captura")
            return
    try:
        asyncio.run(_capturar_etiquetas_playwright(config, aguardar_se_vazio=aguardar, background=background))
        # Retroativo: tenta capturar pedidos sem label_url a cada 5 min (só background)
        if background:
            agora = time.time()
            if agora - _ultimo_retro >= _INTERVALO_RETRO:
                asyncio.run(_capturar_retroativo_playwright(config))
                _ultimo_retro = time.time()
    finally:
        _captura_lock.release()


SUPABASE_STORAGE_URL = "https://qaqlaqlxxeilouvbwgiv.supabase.co/storage/v1/object/public/robo-upseller"


def _upload_etiqueta_supabase(pdf_path: Path, order_number: str) -> str | None:
    """Faz upload do PDF para Supabase Storage e retorna a URL pública permanente.
    Retorna None em caso de falha — o fallback local continua funcionando.
    """
    for tentativa in range(2):
        try:
            supa = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)
            caminho = f"etiquetas/etiqueta_{order_number}.pdf"
            supa.storage.from_("robo-upseller").upload(
                caminho,
                pdf_path.read_bytes(),
                {"content-type": "application/pdf", "upsert": "true"},
            )
            return f"{SUPABASE_STORAGE_URL}/{caminho}"
        except Exception as e:
            if tentativa == 1:
                log(f"[storage] ⚠️ Upload {order_number} falhou: {e}")
                _ultima_rodada["storage_erros"] = _ultima_rodada.get("storage_erros", 0) + 1
    return None


def _limpar_etiquetas_storage(dias: int = 7):
    """Apaga do Supabase Storage etiquetas com mais de N dias."""
    try:
        from datetime import timezone
        supa = create_client(SUPABASE_URL, SUPABASE_SERVICE_KEY)
        arquivos = supa.storage.from_("robo-upseller").list("etiquetas", {"limit": 1000})
        limite = datetime.now(timezone.utc) - timedelta(days=dias)
        apagar = []
        for f in arquivos:
            criado_str = f.get("created_at") or f.get("updated_at") or ""
            if not criado_str:
                continue
            criado = datetime.fromisoformat(criado_str.replace("Z", "+00:00"))
            if criado < limite:
                apagar.append(f"etiquetas/{f['name']}")
        if apagar:
            supa.storage.from_("robo-upseller").remove(apagar)
            log(f"[storage] 🗑️ {len(apagar)} etiqueta(s) antigas removidas do Storage (>{dias} dias)")
        else:
            log(f"[storage] ✅ Nenhuma etiqueta com mais de {dias} dias")
    except Exception as e:
        log(f"[storage] ⚠️ Erro na limpeza: {e}")


def _atualizar_supabase_com_pdfs_locais(token: str):
    """Atualiza label_url no Supabase para todos os pedidos que já têm PDF em disco."""
    pdfs = list(PASTA_ETIQUETAS.glob("etiqueta_*.pdf"))
    if not pdfs:
        return
    try:
        supa = create_client(*_supa())
        atualizados = 0
        for pdf in pdfs:
            on = pdf.stem.replace("etiqueta_", "")
            if not on:
                continue
            # Tenta URL pública Supabase Storage; fallback para localhost se upload falhar
            url_publica = _upload_etiqueta_supabase(pdf, on)
            label_url = url_publica or f"http://127.0.0.1:5001/etiqueta/{on}"
            # Sempre atualiza — sem condição .is_null para não perder PDFs que já existem
            supa.table("pedidos").update({"label_url": label_url}) \
                .eq("order_number", on).execute()
            atualizados += 1
        if atualizados:
            log(f"[import] 🔗 Supabase: {atualizados} etiqueta(s) sincronizadas")
    except Exception as e:
        log(f"[import] ⚠️ Erro ao sincronizar PDFs com Supabase: {e}")


def _atualizar_json_com_urls_robot(token: str):
    """Reescreve o JSON mais recente com URLs do robô para todos os pedidos que têm PDF local."""
    tkn8 = token[:8]
    jsons = sorted(PASTA_DADOS.glob(f"lista_*_{tkn8}.json"),
                   key=lambda p: p.stat().st_mtime, reverse=True)
    if not jsons:
        return
    arq = jsons[0]
    try:
        pedidos = json.loads(arq.read_text(encoding="utf-8"))
        atualizados = 0
        for p in pedidos:
            on = p.get("order_number", "")
            if (PASTA_ETIQUETAS / f"etiqueta_{on}.pdf").exists():
                p["label_url"] = f"http://127.0.0.1:5001/etiqueta/{on}"
                atualizados += 1
        arq.write_text(json.dumps(pedidos, ensure_ascii=False, indent=2), encoding="utf-8")
        if atualizados:
            log(f"[capturar] 🔗 JSON atualizado: {atualizados} etiqueta(s) com URL do robô")
    except Exception as e:
        log(f"[capturar] ⚠️ Erro ao atualizar JSON: {e}")


def _pos_import(config: dict):
    """Etapas executadas após qualquer import, respeitando os toggles em config['etapas'].
    Ordem: captura etiquetas prontas → picklist → NF-e → Envio → captura novas"""
    global _picklist_hoje, _etapa_atual
    etapas = config.get("etapas", {})

    # 1. Sincroniza PDFs em disco com Supabase e captura o que já está em Para Imprimir
    _etapa_atual = "Capturando etiquetas..."
    _atualizar_supabase_com_pdfs_locais(config.get("token", ""))
    log("━━ Capturando etiquetas já prontas ━━")
    _executar_captura(config, aguardar=False)
    _atualizar_json_com_urls_robot(config.get("token", ""))

    if _parar.is_set():
        log("⏹️ Execução cancelada pelo usuário")
        return

    # 2. Picklist (gerada do Supabase — dedup via picklist_impresso_em)
    if etapas.get("picklist", True):
        _etapa_atual = "Gerando picklist..."
        log("━━ Gerando picklist ━━")
        _imprimir_picklist_supabase(config)

    if _parar.is_set():
        log("⏹️ Execução cancelada pelo usuário")
        return

    # 3. Emite NF-e
    if etapas.get("nfe", True):
        _etapa_atual = "Emitindo NF-e..."
        log("━━ Emitindo NF-e em massa ━━")
        asyncio.run(_emitir_nfe_massa_playwright(config))

    if _parar.is_set():
        log("⏹️ Execução cancelada pelo usuário")
        return

    # 4. Programa envio
    if etapas.get("envio", True):
        _etapa_atual = "Programando envio..."
        log("━━ Programando envio ━━")
        asyncio.run(_programar_envio_playwright(config))

    if _parar.is_set():
        log("⏹️ Execução cancelada pelo usuário")
        return

    # 5. Captura etiquetas novas pós NF-e+Envio
    _etapa_atual = "Capturando etiquetas novas..."
    log("━━ Capturando etiquetas novas ━━")
    _executar_captura(config, aguardar=True)
    _atualizar_json_com_urls_robot(config.get("token", ""))


def _rodar_em_thread(config: dict, data_inicio: datetime, data_fim: datetime):
    global rodando, _pos_import_ativo, _etapa_atual
    rodando = True
    _pos_import_ativo = True
    _etapa_atual = "Importando pedidos..."
    _parar.clear()
    _capturar_skip.clear()
    _ultima_rodada.update({"inseridos": 0, "atualizados": 0, "storage_erros": 0,
                           "etiquetas_ok": 0, "etiquetas_total": 0, "sem_etiqueta": 0, "em": None})
    # Aguarda o background capturar fechar o Chrome antes de abrir o Excel
    # (dois processos no mesmo perfil causam conflito)
    adquiriu = _captura_lock.acquire(timeout=45)
    if adquiriu:
        _captura_lock.release()
    else:
        log("⚠️ Background capturar demorou — prosseguindo mesmo assim")
    try:
        etapas = config.get("etapas", {})
        if etapas.get("importar", True):
            data_arquivo = data_fim.date().strftime("%Y-%m-%d")
            log("Importando pedidos via API UpSeller...")
            sucesso = asyncio.run(_importar_via_api(config, data_arquivo))
            if not sucesso:
                log("⚠️ API falhou — usando Excel como fallback...")
                sucesso = asyncio.run(_baixar_excel_playwright(config, data_inicio, data_fim))
                if not sucesso:
                    log("[excel] ⚠️ Tentativa 1 falhou — retentando em 10s...")
                    import time as _t; _t.sleep(10)
                    sucesso = asyncio.run(_baixar_excel_playwright(config, data_inicio, data_fim))
                if sucesso:
                    extrair(config, data_fim.date())

        # Limpeza de etiquetas antigas no Storage (só na rodada das 6h)
        from datetime import time as _time
        if datetime.now().time() < _time(7, 0):
            _etapa_atual = "Limpando etiquetas antigas..."
            _limpar_etiquetas_storage(dias=7)

        # Pedidos que já têm PDF em disco: marca label_url no Supabase antes de liberar o PCP
        _etapa_atual = "Sincronizando etiquetas..."
        _atualizar_supabase_com_pdfs_locais(config.get("token", ""))

        # Import concluído — PCP já pode carregar pedidos e imprimir etiquetas disponíveis
        rodando = False
        log("__fim__")

        # Captura, NF-e e envio continuam sem bloquear o PCP
        _pos_import(config)
    except Exception as e:
        log(f"❌ Erro inesperado: {e}")
        _etapa_atual = ""
    finally:
        rodando = False
        _pos_import_ativo = False
        _etapa_atual = ""


def _rodar_em_thread_duplo(config: dict,
                           ontem_ini: datetime, ontem_fim: datetime,
                           hoje_ini: datetime,  hoje_fim: datetime):
    """Puxa ontem completo, depois hoje até agora."""
    global rodando, _pos_import_ativo
    rodando = True
    _pos_import_ativo = True
    _parar.clear()
    _ultima_rodada.update({"inseridos": 0, "atualizados": 0, "storage_erros": 0,
                           "etiquetas_ok": 0, "etiquetas_total": 0, "sem_etiqueta": 0, "em": None})
    adquiriu = _captura_lock.acquire(timeout=45)
    if adquiriu:
        _captura_lock.release()
    else:
        log("⚠️ Background capturar demorou — prosseguindo mesmo assim")
    try:
        etapas = config.get("etapas", {})
        if etapas.get("importar", True):
            data_hoje = hoje_fim.date().strftime("%Y-%m-%d")
            log("Importando pedidos via API UpSeller...")
            ok = asyncio.run(_importar_via_api(config, data_hoje))
            if not ok:
                log("⚠️ API falhou — usando Excel como fallback...")
                log("━━ Carga do dia anterior ━━")
                ok = asyncio.run(_baixar_excel_playwright(config, ontem_ini, ontem_fim))
                if ok:
                    extrair(config, ontem_fim.date())
                log("━━ Atualização de hoje ━━")
                ok2 = asyncio.run(_baixar_excel_playwright(config, hoje_ini, hoje_fim))
                if ok2:
                    extrair(config, hoje_fim.date())

        # Import concluído — PCP já pode carregar pedidos e imprimir etiquetas disponíveis
        rodando = False
        log("__fim__")

        _pos_import(config)
    except Exception as e:
        log(f"❌ Erro inesperado: {e}")
    finally:
        rodando = False
        _pos_import_ativo = False


def _imprimir_picklist_thread(config: dict):
    """Imprime picklist em thread separada (chamado pelo agendador).
    Aguarda automaticamente se um import estiver em andamento."""
    # Pequena pausa para deixar o import setar rodando=True caso tenha disparado junto
    time.sleep(5)
    if rodando:
        log("━━ Picklist aguardando import terminar... ━━")
        while rodando:
            time.sleep(10)
    log("━━ Gerando picklist automático ━━")
    _imprimir_picklist_supabase(config)


# ── Playwright: download automático do Excel ──────────────────────────────────

async def _baixar_excel_playwright(config: dict, data_inicio: datetime, data_fim: datetime) -> bool:
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        log("❌ Playwright não instalado. Rode: pip install playwright && playwright install chromium")
        return False

    _limpar_crash_chrome()
    data_arquivo    = data_fim.strftime("%Y-%m-%d")
    ini_str         = data_inicio.strftime("%d/%m/%Y %H:%M:%S")
    fim_str         = data_fim.strftime("%d/%m/%Y %H:%M:%S")
    PASTA_ARQUIVOS  = PASTA_RAIZ / "arquivos"
    PASTA_ARQUIVOS.mkdir(exist_ok=True)
    caminho_destino = PASTA_ARQUIVOS / f"pedidos_{data_arquivo}.xlsx"

    log(f"Abrindo UpSeller — {ini_str} até {fim_str}...")

    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            user_data_dir=str(PASTA_SESSAO),
            headless=True,
            args=["--window-size=1280,800"],
        )
        try:
            page = context.pages[0] if context.pages else await context.new_page()

            await page.goto(
                "https://app.upseller.com/pt/order/all-orders",
                wait_until="domcontentloaded",
                timeout=60_000,
            )
            await page.wait_for_timeout(3000)

            log(f"URL atual: {page.url}")
            if "/login" in page.url:
                # Sessão expirada — reabre em modo visível para login manual
                log("⚠️ Sessão expirada. Abrindo browser para login manual...")
                await context.close()
                context = await p.chromium.launch_persistent_context(
                    user_data_dir=str(PASTA_SESSAO),
                    headless=False,
                    args=["--window-size=1280,800"],
                )
                page = context.pages[0] if context.pages else await context.new_page()
                await page.goto("https://app.upseller.com/login", wait_until="domcontentloaded", timeout=60_000)
                log("🔑 Faça login no UpSeller (janela aberta). Aguardando...")
                await page.wait_for_url(lambda url: "/login" not in url, timeout=600_000)
                log("✅ Login realizado! Salvando sessão...")
                await page.wait_for_load_state("networkidle", timeout=30_000)
                await page.wait_for_timeout(3000)
                log("✅ Sessão salva. Continuando importação...")
                await page.goto(
                    "https://app.upseller.com/pt/order/all-orders",
                    wait_until="domcontentloaded",
                    timeout=60_000,
                )
                await page.wait_for_timeout(3000)

            log("Conectado. Aplicando filtro de data...")

            await _fechar_popups_upseller(page)

            await page.locator("input[placeholder='Filtrar por data']").first.click(force=True)
            await page.wait_for_timeout(1500)

            start_input = page.locator("input[placeholder='Filtrar por data']").last
            await start_input.click(click_count=3, force=True)
            await start_input.type(ini_str, delay=30)
            await page.wait_for_timeout(300)
            await page.keyboard.press("Tab")
            await page.wait_for_timeout(300)
            await page.keyboard.type(fim_str, delay=30)
            await page.wait_for_timeout(300)

            ok_btn = page.locator(".ant-calendar-ok-btn")
            if await ok_btn.count() > 0:
                await ok_btn.click(force=True)
            await page.wait_for_timeout(2000)
            log(f"Filtro: {ini_str} → {fim_str}")

            log("Exportando Excel...")
            # Garante que o botão está visível antes de clicar (page pode estar scrollada)
            await page.evaluate("window.scrollTo(0, 0)")
            await page.wait_for_timeout(400)
            # Usa mouse.click() com coordenadas reais — dispara pointer events que o React/AntD precisa
            pos_exportar = await page.evaluate("""() => {
                const btn = Array.from(document.querySelectorAll('button'))
                    .find(b => b.innerText.trim().includes('Exportar') && b.offsetParent !== null
                          && !b.innerText.includes('Todos'));
                if (!btn) return null;
                btn.scrollIntoView({block: 'center'});
                const r = btn.getBoundingClientRect();
                return {x: r.x + r.width/2, y: r.y + r.height/2};
            }""")
            if pos_exportar:
                await page.mouse.click(pos_exportar['x'], pos_exportar['y'])
                log(f"[excel] Clicou Exportar em ({pos_exportar['x']:.0f},{pos_exportar['y']:.0f})")
            else:
                log("[excel] ⚠️ Botão Exportar não encontrado na página")
            await page.wait_for_timeout(1200)

            pos_todos = await page.evaluate("""() => {
                const el = Array.from(document.querySelectorAll('li, .ant-dropdown-menu-item, [role="menuitem"]'))
                    .find(e => e.innerText && e.innerText.trim().includes('Exportar Todos') && e.offsetParent !== null);
                if (!el) return null;
                const r = el.getBoundingClientRect();
                return {x: r.x + r.width/2, y: r.y + r.height/2};
            }""")
            if pos_todos:
                await page.mouse.click(pos_todos['x'], pos_todos['y'])
                log(f"[excel] Clicou 'Exportar Todos' em ({pos_todos['x']:.0f},{pos_todos['y']:.0f})")
            else:
                log("[excel] ⚠️ 'Exportar Todos os Pedidos' não encontrado no dropdown")

            # Aguarda modal abrir — não chamar _fechar_popups aqui (fecharia o modal)
            try:
                await page.wait_for_selector('.ant-modal-content', timeout=6000)
                log("[excel] Modal de exportar aberto")
            except Exception:
                log("[excel] ⚠️ Modal não apareceu após 6s")

            # Intercepta respostas HTTP para capturar o Excel diretamente da rede
            # (funciona mesmo quando UpSeller não usa download do browser)
            _excel_capturado: list = []
            _excel_event = asyncio.Event()

            async def _captura_resposta_excel(response):
                if _excel_event.is_set():
                    return
                ct = response.headers.get('content-type', '').lower()
                cd = response.headers.get('content-disposition', '').lower()
                if ('spreadsheet' in ct or 'excel' in ct or
                        ('attachment' in cd and ('xls' in cd or 'xlsx' in cd))):
                    try:
                        body = await response.body()
                        if len(body) > 5000 and body[:2] == b'PK':  # XLSX = ZIP
                            _excel_capturado.append(body)
                            _excel_event.set()
                            log(f"[excel] 📥 Excel capturado via rede ({len(body)//1024}KB)")
                    except Exception:
                        pass

            page.on('response', _captura_resposta_excel)

            # Também monitora requests para diagnóstico
            _reqs_exportar: list = []

            async def _log_req(request):
                url = request.url
                if 'upseller.com' in url and '/order/all-orders' not in url:
                    _reqs_exportar.append(f"{request.method} {url[:120]}")

            page.on('request', _log_req)

            # Envolve TODO o fluxo do modal num único expect_download (300s)
            async with page.expect_download(timeout=300_000) as dl_info:

                # Loga os botões disponíveis no modal para diagnóstico
                botoes_modal = await page.evaluate("""() => {
                    const modal = document.querySelector('.ant-modal-content');
                    if (!modal) return [];
                    return Array.from(modal.querySelectorAll('button'))
                        .filter(b => b.offsetParent !== null)
                        .map(b => ({text: b.innerText.trim(), disabled: b.disabled}));
                }""")
                log(f"[excel] 🔎 Botões no modal: {botoes_modal}")

                # Clica "Exportar" dentro do modal (abre sub-form de páginas)
                pos_modal_exp = await page.evaluate("""() => {
                    const modal = document.querySelector('.ant-modal-content');
                    if (!modal) return null;
                    const btn = Array.from(modal.querySelectorAll('button'))
                        .find(b => b.innerText.trim().includes('Exportar') && b.offsetParent !== null);
                    if (!btn) return null;
                    const r = btn.getBoundingClientRect();
                    return {x: r.x + r.width/2, y: r.y + r.height/2, disabled: btn.disabled};
                }""")
                if pos_modal_exp:
                    log(f"[excel] Clicando Exportar no modal (disabled={pos_modal_exp.get('disabled')})")
                    await page.mouse.click(pos_modal_exp['x'], pos_modal_exp['y'])
                    await page.wait_for_timeout(2000)  # aguarda form de páginas carregar

                # Preenche "até página 999" — aguarda até 5s pelo input aparecer
                inputs_pagina = page.locator(
                    ".ant-modal-content input[type='number'], "
                    ".ant-modal-content input[type='text'], "
                    ".ant-modal-content .ant-input-number-input"
                )
                n_inputs = await inputs_pagina.count()
                if n_inputs == 0:
                    await page.wait_for_timeout(2000)
                    n_inputs = await inputs_pagina.count()
                log(f"[excel] {n_inputs} input(s) encontrado(s) no modal")

                if n_inputs >= 1:
                    ultimo = inputs_pagina.nth(n_inputs - 1)
                    await ultimo.click(click_count=3, force=True)
                    await ultimo.fill("999")
                    await page.keyboard.press("Tab")
                    await page.wait_for_timeout(300)
                    valor_final = await ultimo.input_value()
                    log(f"[excel] Exportando páginas 1 a {valor_final}")

                    pos_modal_exp2 = await page.evaluate("""() => {
                        const modal = document.querySelector('.ant-modal-content');
                        if (!modal) return null;
                        const btn = Array.from(modal.querySelectorAll('button'))
                            .find(b => b.innerText.trim().includes('Exportar') && b.offsetParent !== null);
                        if (!btn) return null;
                        const r = btn.getBoundingClientRect();
                        return {x: r.x + r.width/2, y: r.y + r.height/2};
                    }""")
                    if pos_modal_exp2:
                        await page.mouse.click(pos_modal_exp2['x'], pos_modal_exp2['y'])
                        log("[excel] Exportar submetido")
                else:
                    # 0 inputs — loga requests ANTES do 2º clique para diagnóstico
                    log(f"[excel] 🔎 Requests após 1º clique: {_reqs_exportar}")

                    # Screenshot do estado do modal para diagnóstico
                    sc_modal = PASTA_RAIZ / "debug_excel_modal.png"
                    try:
                        await page.screenshot(path=str(sc_modal))
                        log(f"[excel] 📸 Screenshot: {sc_modal.name}")
                    except Exception:
                        pass

                    # Tenta JS click (força o handler React/Vue independente de estado)
                    clicou_js = await page.evaluate("""() => {
                        const modal = document.querySelector('.ant-modal-content');
                        if (!modal) return false;
                        const btn = Array.from(modal.querySelectorAll('button'))
                            .find(b => b.innerText.trim().includes('Exportar'));
                        if (!btn) return false;
                        btn.removeAttribute('disabled');
                        btn.click();
                        return true;
                    }""")
                    log(f"[excel] JS click Exportar: {clicou_js}")
                    await page.wait_for_timeout(2000)
                    log(f"[excel] 🔎 Requests após JS click: {_reqs_exportar}")

                log("Gerando arquivo, aguardando...")
                await page.wait_for_timeout(3000)

                # Tenta encontrar e clicar "Baixar"
                try:
                    baixar_btn = page.locator("button:has-text('Baixar'), a:has-text('Baixar')").first
                    await baixar_btn.wait_for(timeout=300_000)
                    await baixar_btn.click(force=True)
                    log("Baixando arquivo...")
                except Exception:
                    log("[excel] ℹ️ Botão Baixar não apareceu — verificando captura de rede...")
                    # Aguarda até 90s pela captura via interceptação HTTP
                    try:
                        await asyncio.wait_for(_excel_event.wait(), timeout=90)
                    except Exception:
                        pass
                    await page.wait_for_timeout(3000)

            # Tenta download do browser primeiro
            try:
                download = await dl_info.value
                await download.save_as(caminho_destino)
                log(f"✅ Excel salvo: pedidos_{data_arquivo}.xlsx")
                return True
            except Exception:
                pass

            # Fallback: arquivo capturado via interceptação HTTP
            if _excel_capturado:
                caminho_destino.write_bytes(_excel_capturado[0])
                log(f"✅ Excel salvo (via rede): pedidos_{data_arquivo}.xlsx")
                return True

            log("❌ Excel não capturado por nenhum método")
            return False

        except Exception as e:
            log(f"❌ Erro ao baixar Excel: {e}")
            return False
        finally:
            await context.close()


# ── Limpa estado de crash do perfil Chrome ───────────────────────────────────

def _limpar_crash_chrome():
    """Reseta exit_type no perfil Chrome para evitar o modo recuperação de crash."""
    import json as _json
    prefs_path = PASTA_SESSAO / "Default" / "Preferences"
    if not prefs_path.exists():
        return
    try:
        prefs = _json.loads(prefs_path.read_text(encoding="utf-8"))
        if prefs.get("profile", {}).get("exit_type") not in (None, "Normal"):
            prefs.setdefault("profile", {})["exit_type"] = "Normal"
            prefs_path.write_text(_json.dumps(prefs), encoding="utf-8")
    except Exception:
        pass


# ── Fecha popups/modais do UpSeller ──────────────────────────────────────────

async def _fechar_popups_upseller(page) -> None:
    """Fecha qualquer popup/aviso que o UpSeller exibe ao navegar."""
    # Driver.js tour — clica "Ignorar" (span, não button) e remove overlay do DOM
    try:
        fechou_driver = await page.evaluate("""() => {
            // Tenta clicar em qualquer elemento com texto "Ignorar" dentro do popover
            const els = Array.from(document.querySelectorAll(
                '.driver-popover-footer span, .driver-popover span, .driver-popover button, .driver-popover a'
            )).filter(el => el.innerText && el.innerText.trim() === 'Ignorar');
            if (els.length) { els[0].click(); return true; }
            // Fallback: remove toda a estrutura do driver.js do DOM
            const overlay = document.querySelectorAll('.driver-overlay, .driver-popover, #driver-page-overlay');
            if (overlay.length) { overlay.forEach(el => el.remove()); return true; }
            return false;
        }""")
        if fechou_driver:
            await page.wait_for_timeout(500)
    except Exception:
        pass

    # Modal com YouTube iframe (UpSeller promotional popup) — remove do DOM diretamente
    try:
        removeu = await page.evaluate("""() => {
            const modals = Array.from(document.querySelectorAll('.ant-modal-wrap, .ant-modal-root'));
            let removido = 0;
            modals.forEach(m => {
                if (m.querySelector('iframe[src*="youtube"]') || m.offsetParent !== null) {
                    m.remove();
                    removido++;
                }
            });
            // Remove também o overlay de fundo
            document.querySelectorAll('.ant-modal-mask').forEach(el => el.remove());
            return removido > 0;
        }""")
        if removeu:
            await page.wait_for_timeout(300)
    except Exception:
        pass

    for _ in range(8):
        try:
            # Botão X padrão do Ant Design Modal
            x_btn = page.locator(".ant-modal-close").first
            if await x_btn.count() > 0 and await x_btn.is_visible():
                await x_btn.click()
                await page.wait_for_timeout(600)
                continue
            # Botões de dismiss — role=button E texto genérico (cobre spans/divs também)
            fechou = False
            for nome in ["Ignorar", "Entendido", "Fechar", "OK", "Dispensar", "Não mostrar novamente"]:
                for loc in [
                    page.get_by_role("button", name=nome).first,
                    page.locator(f"text={nome}").first,
                ]:
                    if await loc.count() > 0 and await loc.is_visible():
                        await loc.click(force=True)
                        await page.wait_for_timeout(500)
                        fechou = True
                        break
                if fechou:
                    break
            if not fechou:
                break
        except Exception:
            break
    try:
        await page.keyboard.press("Escape")
        await page.wait_for_timeout(300)
    except Exception:
        pass


# ── Impressora / PDF ──────────────────────────────────────────────────────────

async def _imprimir_ou_pdf(popup, tem_impressora: bool, nome: str, prefixo: str):
    """Imprime via OS ou salva PDF. Para URLs diretas de PDF (print-label.upseller.cn),
    baixa o arquivo com urllib em vez de usar page.pdf()."""
    import urllib.request

    url = popup.url

    def _abrir(path: Path):
        if platform.system() == "Darwin":
            subprocess.Popen(["open", str(path)])
        elif platform.system() == "Windows":
            subprocess.Popen(["start", str(path)], shell=True)
        else:
            subprocess.Popen(["xdg-open", str(path)])

    # PDF hospedado externamente (ex: print-label.upseller.cn)
    pasta_pdf = PASTA_PICKLISTS if prefixo == "picklist" else PASTA_ETIQUETAS

    if url.lower().endswith(".pdf") or "print-label" in url.lower():
        pdf_path = pasta_pdf / f"{prefixo}_{nome}.pdf"
        try:
            urllib.request.urlretrieve(url, str(pdf_path))
            # Corrige dimensões (A4 → 100×150mm) e substitui o original para que o
            # arquivo salvo na pasta já esteja no tamanho certo para impressão manual
            if prefixo != "picklist":
                _corr = _corrigir_mediabox(pdf_path)
                if _corr != pdf_path and _corr.exists():
                    _corr.replace(pdf_path)
            log(f"[{prefixo}] 📄 PDF salvo: {pdf_path.name}")
            if prefixo != "picklist" and tem_impressora:
                if platform.system() == "Windows":
                    cfg = json.loads(CONFIG_FILE.read_text(encoding="utf-8")) if CONFIG_FILE.exists() else {}
                    _imprimir_windows(pdf_path, cfg.get("nome_impressora", "").strip())
                elif platform.system() == "Darwin":
                    subprocess.Popen(["lp", str(pdf_path)])
                log(f"[{prefixo}] 🖨️ Impresso direto da pasta")
            elif prefixo != "picklist":
                _abrir(pdf_path)
        except Exception as e:
            log(f"[{prefixo}] ⚠️ Erro ao baixar/imprimir PDF: {e}")
        return

    # Página HTML — gera PDF (headless obrigatório) e imprime ou abre
    pdf_path = pasta_pdf / f"{prefixo}_{nome}.pdf"
    # Picklist: 4x6 polegadas (impressora térmica Elgin/Epson). Demais: A4
    if prefixo == "picklist":
        await popup.pdf(path=str(pdf_path), print_background=True,
                        width="4in", height="6in",
                        margin={"top": "4mm", "bottom": "4mm", "left": "4mm", "right": "4mm"})
    else:
        await popup.pdf(path=str(pdf_path), print_background=True, format="A4")
    log(f"[{prefixo}] 📄 PDF gerado: {pdf_path.name}")

    # Picklist: não abre automaticamente — PCP serve e imprime
    if prefixo == "picklist":
        log("[picklist] 📋 PDF disponível na pasta — imprima pelo PCP")
        return

    if tem_impressora:
        try:
            if platform.system() == "Darwin":
                subprocess.Popen(["lp", str(pdf_path)])
                log(f"[{prefixo}] 🖨️ Enviado para impressora (lp)")
            elif platform.system() == "Windows":
                cfg2 = json.loads(CONFIG_FILE.read_text(encoding="utf-8")) if CONFIG_FILE.exists() else {}
                _imprimir_windows(_corrigir_mediabox(pdf_path), cfg2.get("nome_impressora", "").strip())
                log(f"[{prefixo}] 🖨️ Enviado para impressora (Windows)")
        except Exception as e:
            log(f"[{prefixo}] ⚠️ Erro ao imprimir: {e} — abrindo PDF")
            _abrir(pdf_path)
    else:
        _abrir(pdf_path)


# ── Helpers de PDF ────────────────────────────────────────────────────────────

def _contar_paginas_pdf(pdf_path: "Path") -> int:
    try:
        from pypdf import PdfReader
        return len(PdfReader(str(pdf_path), strict=False).pages)
    except Exception:
        return 1

def _garantir_pagina_unica(pdf_path: "Path") -> None:
    """Se o PDF tiver múltiplas páginas (batch), mantém só a página 1."""
    try:
        from pypdf import PdfReader, PdfWriter
        reader = PdfReader(str(pdf_path), strict=False)
        if len(reader.pages) <= 1:
            return
        writer = PdfWriter()
        writer.add_page(reader.pages[0])
        with open(str(pdf_path), "wb") as f:
            writer.write(f)
        # Remove versão _c para regenerar com dimensões corretas
        corr = pdf_path.with_stem(pdf_path.stem + "_c")
        corr.unlink(missing_ok=True)
    except Exception:
        pass

def _split_pdf_batch(pdf_path: "Path", ordens: list) -> list:
    """Divide PDF batch (N páginas) em um arquivo por pedido. Retorna lista de order_numbers salvos."""
    try:
        from pypdf import PdfReader, PdfWriter
        reader = PdfReader(str(pdf_path), strict=False)
        n_pages = len(reader.pages)
        salvos = []
        for i, on in enumerate(ordens):
            if i >= n_pages:
                break
            dest = PASTA_ETIQUETAS / f"etiqueta_{on}.pdf"
            if dest.exists():
                salvos.append(on)
                continue
            writer = PdfWriter()
            writer.add_page(reader.pages[i])
            with open(str(dest), "wb") as f:
                writer.write(f)
            c = _corrigir_mediabox(dest)
            if c != dest and c.exists():
                c.replace(dest)
            salvos.append(on)
        return salvos
    except Exception as e:
        log(f"[capturar] ⚠️ Split PDF falhou: {e}")
        return []


# ── SumatraPDF (impressão com escala no Windows) ─────────────────────────────

def _corrigir_mediabox(pdf_path: "Path") -> "Path":
    """Recorta o PDF para 100×150mm (papel térmico padrão).
    Para PDFs A4 de carrier (TikTok/iMile) a etiqueta fica no canto superior esquerdo."""
    import re, zlib

    out = pdf_path.with_stem(pdf_path.stem + "_c")
    if out.exists():
        return out
    try:
        data = pdf_path.read_bytes()

        # --- Lê o MediaBox original (bytes crus ou stream comprimido) ---
        streams_raw = re.findall(rb'stream\r?\n(.*?)\r?\nendstream', data, re.DOTALL)

        m_orig = re.search(rb'/MediaBox\s*\[([^\]]+)\]', data)
        if not m_orig:
            for s in streams_raw:
                try:
                    dec = zlib.decompress(s)
                    m_orig = re.search(rb'/MediaBox\s*\[([^\]]+)\]', dec)
                    if m_orig:
                        break
                except Exception:
                    pass

        if not m_orig:
            return pdf_path

        orig_vals = list(map(float, m_orig.group(1).split()))
        if len(orig_vals) != 4:
            return pdf_path

        orig_w = orig_vals[2] - orig_vals[0]
        orig_h = orig_vals[3] - orig_vals[1]

        # Se já é tamanho de etiqueta térmica (≤ 130mm × 170mm ≈ 368pt × 481pt), não corrige
        if orig_w <= 368 and orig_h <= 481:
            return pdf_path

        # --- Calcula novo MediaBox ---
        LABEL_W = 283.46   # 100mm em points (72pt/in)
        LABEL_H = 425.20   # 150mm em points
        x0_new = y0_new = x1_new = y1_new = None

        # Caso A4 portrait (etiqueta no canto superior esquerdo — confirmado para TikTok/iMile)
        if abs(orig_w - 595.28) < 20 and abs(orig_h - 841.89) < 20:
            x0_new, y0_new = 0.0, 841.89 - LABEL_H   # ≈ 416.69
            x1_new, y1_new = LABEL_W, 841.89

        # Caso A4 landscape
        elif abs(orig_w - 841.89) < 20 and abs(orig_h - 595.28) < 20:
            x0_new, y0_new = 0.0, 595.28 - LABEL_H
            x1_new, y1_new = LABEL_W, 595.28

        else:
            # Fallback: tenta detectar Form XObject + BBox (método original)
            tx = ty = bw = bh = None
            for s in streams_raw:
                for raw in [False, True]:
                    try:
                        text = (zlib.decompress(s) if raw else s).decode('latin-1', errors='ignore')
                    except Exception:
                        continue
                    m = re.search(
                        r'([\d.+-]+)\s+([\d.+-]+)\s+([\d.+-]+)\s+([\d.+-]+)\s+([\d.+-]+)\s+([\d.+-]+)\s+cm\s*/\w+\s+Do',
                        text)
                    if m:
                        a, b, c, d, e, f = (float(m.group(i)) for i in range(1, 7))
                        if abs(a - 1) < 0.1 and abs(d - 1) < 0.1 and abs(b) < 0.1 and abs(c) < 0.1:
                            tx, ty = e, f
                    m2 = re.search(r'/BBox\s*\[([^\]]+)\]', text)
                    if m2:
                        bb = list(map(float, m2.group(1).split()))
                        if len(bb) == 4:
                            bw, bh = bb[2] - bb[0], bb[3] - bb[1]
                    if tx is not None and bw is not None:
                        break
                if tx is not None and bw is not None:
                    break

            if tx is not None and bw is not None and bw >= 60 and bh >= 80:
                x0_new, y0_new = tx, ty
                x1_new, y1_new = tx + bw, ty + bh

        if x0_new is None:
            return pdf_path  # não conseguiu detectar

        new_mb_bytes = f'/MediaBox[{x0_new:.4f} {y0_new:.4f} {x1_new:.4f} {y1_new:.4f}]'.encode('latin-1')

        # Tenta patch nos bytes crus (MediaBox fora de stream comprimido)
        patched = re.sub(rb'/MediaBox\s*\[[^\]]+\]', new_mb_bytes, data)
        if patched != data:
            out.write_bytes(patched)
            log(f"[etiqueta] ✂️ MediaBox {orig_w/2.835:.0f}×{orig_h/2.835:.0f}mm → {(x1_new-x0_new)/2.835:.0f}×{(y1_new-y0_new)/2.835:.0f}mm")
            return out

        # Se MediaBox está dentro de stream comprimido, recomprime
        def patch_stream(m):
            hdr, length_s, mid, sdata, tail = m.group(1), m.group(2), m.group(3), m.group(4), m.group(5)
            try:
                dec = zlib.decompress(sdata)
                text = dec.decode('latin-1', errors='ignore')
                if 'MediaBox' not in text:
                    return m.group(0)
                new_text = re.sub(r'/MediaBox\[[^\]]+\]', new_mb_bytes.decode('latin-1'), text)
                if new_text == text:
                    return m.group(0)
                new_comp = zlib.compress(new_text.encode('latin-1'))
                new_len = str(len(new_comp)).encode()
                new_hdr = re.sub(rb'/Length\s+\d+', b'/Length ' + new_len, hdr + length_s + mid)
                return new_hdr + new_comp + tail
            except Exception:
                return m.group(0)

        pattern = re.compile(
            rb'(<<.*?/Filter/FlateDecode.*?/Length\s+)(\d+)(.*?>>\s*stream\r?\n)([\x00-\xff]+?)(\r?\nendstream)',
            re.DOTALL
        )
        result = pattern.sub(patch_stream, data)
        out.write_bytes(result)
        log(f"[etiqueta] ✂️ MediaBox {orig_w/2.835:.0f}×{orig_h/2.835:.0f}mm → {(x1_new-x0_new)/2.835:.0f}×{(y1_new-y0_new)/2.835:.0f}mm (stream)")
        return out
    except Exception as e:
        log(f"[etiqueta] ⚠️ _corrigir_mediabox falhou: {e}")
        return pdf_path


def _adicionar_margem_topo(pdf_path: "Path", topo_mm: float, esq_mm: float = 0.0) -> "Path":
    """Injeta um operador 'cm' de escala+centralização nos content streams do PDF
    usando apenas stdlib (re + zlib) — sem dependência de pypdf.
    Funciona independente de /Rotate (Shein/Shopee têm rotação 90°)."""
    if topo_mm <= 0 and esq_mm <= 0:
        return pdf_path
    try:
        import re as _re, zlib as _zlib
        t_tag = str(topo_mm).replace(".", "_")
        e_tag = str(esq_mm).replace(".", "_")
        out = pdf_path.with_stem(pdf_path.stem + f"_mz{t_tag}x{e_tag}")
        if out.exists():
            return out

        data = pdf_path.read_bytes()

        # Dimensões da página original
        mb = _re.search(rb'/MediaBox\s*\[([^\]]+)\]', data)
        if not mb:
            log(f"[etiqueta] ⚠️ margem: MediaBox nao encontrado em {pdf_path.name}")
            return pdf_path
        vals = list(map(float, mb.group(1).split()))
        if len(vals) != 4:
            return pdf_path
        page_w = vals[2] - vals[0]
        page_h = vals[3] - vals[1]

        pt_margin = max(topo_mm, esq_mm) * 72 / 25.4
        scale = min((page_w - 2 * pt_margin) / page_w,
                    (page_h - 2 * pt_margin) / page_h)
        scale = max(0.5, min(0.99, scale))
        tx = page_w * (1 - scale) / 2
        ty = page_h * (1 - scale) / 2

        # Operadores PDF a injetar: salva estado, aplica escala+translação
        cm_open  = (f"q\n{scale:.6f} 0 0 {scale:.6f} {tx:.4f} {ty:.4f} cm\n").encode()
        cm_close = b"\nQ\n"

        patched  = data
        n_patched = 0

        def _patch(m):
            nonlocal n_patched
            hdr, length_s, mid, sdata, tail = m.group(1), m.group(2), m.group(3), m.group(4), m.group(5)
            try:
                dec = _zlib.decompress(sdata)
            except Exception:
                return m.group(0)
            # Só modifica streams que parecem content streams (têm operadores PDF legíveis)
            sample = dec[:400].decode('latin-1', errors='replace')
            if not _re.search(r'\b(?:q|Q|BT|ET|cm|Do|Tf|Td|Tm|re|m\b|l\b|S\b|f\b|w\b|rg\b|RG\b|gs\b)\b', sample):
                return m.group(0)
            new_dec  = cm_open + dec + cm_close
            new_comp = _zlib.compress(new_dec)
            new_hdr  = _re.sub(rb'/Length\s+\d+',
                               b'/Length ' + str(len(new_comp)).encode(),
                               hdr + length_s + mid)
            n_patched += 1
            return new_hdr + new_comp + tail

        pattern = _re.compile(
            rb'(<<[^>]*?/Filter\s*/FlateDecode[^>]*?/Length\s+)(\d+)([^>]*?>>\s*stream\r?\n)'
            rb'([\x00-\xff]+?)(\r?\nendstream)',
            _re.DOTALL
        )
        patched = pattern.sub(_patch, patched)

        if n_patched == 0:
            log(f"[etiqueta] ⚠️ margem: nenhum content stream encontrado em {pdf_path.name}")
            return pdf_path

        out.write_bytes(patched)
        log(f"[etiqueta] 📐 margem={pt_margin/2.835:.1f}mm escala={scale:.3f} streams={n_patched} → {out.name}")
        return out
    except Exception as e:
        log(f"[etiqueta] ⚠️ _adicionar_margem_topo: {e}")
        return pdf_path


def _obter_sumatra() -> "Path | None":
    """Retorna path do SumatraPDF.exe — baixa portátil se não encontrar."""
    import shutil
    local = PASTA_RAIZ / "SumatraPDF.exe"
    if local.exists():
        return local
    sys_path = shutil.which("SumatraPDF")
    if sys_path:
        return Path(sys_path)
    try:
        url = ("https://qaqlaqlxxeilouvbwgiv.supabase.co"
               "/storage/v1/object/public/robo-upseller/SumatraPDF.exe")
        log("[sumatra] 📥 Baixando SumatraPDF...")
        import urllib.request as _ur
        _ur.urlretrieve(url, str(local))
        log("[sumatra] ✅ SumatraPDF pronto")
        return local
    except Exception as e:
        log(f"[sumatra] ⚠️ Não foi possível baixar: {e}")
        return None


def _imprimir_windows(pdf_path: "Path", impressora: str) -> bool:
    """Imprime PDF no Windows usando SumatraPDF (escala para preencher o papel)."""
    sumatra = _obter_sumatra()
    if sumatra:
        if impressora:
            cmd = [str(sumatra), "-print-to", impressora,
                   "-print-settings", "fit,color", str(pdf_path)]
        else:
            cmd = [str(sumatra), "-print-to-default",
                   "-print-settings", "fit,color", str(pdf_path)]
        try:
            subprocess.run(cmd, creationflags=subprocess.CREATE_NO_WINDOW, timeout=60)
            return True
        except Exception as e:
            log(f"[etiqueta] ⚠️ SumatraPDF erro: {e}")
    # Fallback: PowerShell
    try:
        ps = (f'Start-Process -FilePath "{pdf_path}" -Verb PrintTo -ArgumentList "{impressora}" -Wait'
              if impressora else
              f'Start-Process -FilePath "{pdf_path}" -Verb Print -Wait')
        subprocess.run(["powershell", "-Command", ps],
                       creationflags=subprocess.CREATE_NO_WINDOW, timeout=30)
    except Exception as e:
        log(f"[etiqueta] ⚠️ Fallback erro: {e}")
    return True


# ── Verificação de impressora ─────────────────────────────────────────────────

def _verificar_impressora() -> bool:
    """Retorna True se houver impressora padrão configurada no sistema."""
    try:
        if platform.system() == "Darwin":
            # Usa somente a impressora padrão do sistema. O fallback lpstat -a capturava
            # impressoras não-padrão (Epson, AirPrint) e mandava labels para o lugar errado.
            r = subprocess.run(["lpstat", "-d"], capture_output=True, text=True, timeout=5)
            output = r.stdout.strip()
            sem_default = ("nenhum" in output.lower() or "no system default" in output.lower() or not output)
            return r.returncode == 0 and not sem_default
        elif platform.system() == "Windows":
            # PowerShell — mais confiável que wmic no Windows 10/11
            try:
                ps = "(Get-Printer | Where-Object {$_.PrinterStatus -ne 'Offline'} | Measure-Object).Count"
                r = subprocess.run(
                    ["powershell", "-NoProfile", "-NonInteractive", "-Command", ps],
                    capture_output=True, text=True, timeout=10,
                    creationflags=subprocess.CREATE_NO_WINDOW,
                )
                if r.returncode == 0 and r.stdout.strip().isdigit() and int(r.stdout.strip()) > 0:
                    return True
            except Exception:
                pass
            # Fallback: wmic (Windows 10 ou anterior)
            r2 = subprocess.run(
                ["wmic", "printer", "where", "Default=TRUE", "get", "Name"],
                capture_output=True, text=True, timeout=5,
            )
            lines = [l.strip() for l in r2.stdout.splitlines() if l.strip() and l.strip() != "Name"]
            return len(lines) > 0
    except Exception:
        pass
    return False


def _nome_impressora_padrao() -> str:
    """Retorna o nome da impressora padrão do sistema (para usar com lp -d)."""
    try:
        if platform.system() == "Darwin":
            r = subprocess.run(["lpstat", "-d"], capture_output=True, text=True, timeout=5)
            output = r.stdout.strip()
            if output and "nenhum" not in output.lower() and "no system default" not in output.lower():
                # "destino padrão do sistema: NomeDaImpressora"
                partes = output.split(":")
                if len(partes) > 1:
                    return partes[-1].strip()
    except Exception:
        pass
    return ""


# ── Picklist gerada do Supabase (sem Playwright) ─────────────────────────────

def _gerar_picklist_pdf(pedidos: list, margem_topo_mm: float = 5.0) -> "Path":
    """Gera PDF de picklist em papel térmico 10x5cm — capa resumo + uma etiqueta por pedido."""
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.units import mm as rl_mm
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.colors import HexColor
    from collections import Counter

    PASTA_PICKLISTS.mkdir(exist_ok=True)
    nome = f"picklist_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    out  = PASTA_PICKLISTS / nome

    LW = 100 * rl_mm
    LH = 50  * rl_mm
    offset_y = margem_topo_mm * rl_mm
    topo = LH - offset_y - 2 * rl_mm

    ts = datetime.now().strftime("%d/%m/%Y %H:%M")

    def _cat(p):
        st  = (p.get("status") or "").strip().lower()
        sup = (p.get("status_upseller") or "").strip().lower()
        if st == "cancelado":          return "cancelado"
        if sup == "para reservar":     return "reservar"
        if sup == "para retirada":     return "retirada"
        return "normal"

    # Ordena: normal → retirada → reservar → cancelado
    _ORD = {"normal": 0, "retirada": 1, "reservar": 2, "cancelado": 3}
    pedidos = sorted(pedidos, key=lambda p: _ORD.get(_cat(p), 0))
    cats = Counter(_cat(p) for p in pedidos)

    _BADGE = {
        "retirada":  (HexColor("#1D4ED8"), HexColor("#DBEAFE"), "RETIRADA"),
        "reservar":  (HexColor("#92400E"), HexColor("#FEF3C7"), "RESERVAR"),
        "cancelado": (HexColor("#991B1B"), HexColor("#FEE2E2"), "CANCELADO"),
    }

    c = rl_canvas.Canvas(str(out), pagesize=(LW, LH))

    # ── CAPA RESUMO ───────────────────────────────────────────────────────────
    c.setFont("Helvetica-Bold", 9)
    c.setFillColor(rl_colors.black)
    c.drawString(4 * rl_mm, topo, f"PICKLIST — {ts}")
    c.setFont("Helvetica-Bold", 8)
    c.drawString(4 * rl_mm, topo - 8 * rl_mm, f"Total: {len(pedidos)} pedidos")

    y = topo - 16 * rl_mm
    for label, key, color in [
        ("Para Processar", "normal",    rl_colors.black),
        ("Para Retirada",  "retirada",  HexColor("#1D4ED8")),
        ("Para Reservar",  "reservar",  HexColor("#92400E")),
        ("Cancelados",     "cancelado", HexColor("#991B1B")),
    ]:
        n = cats.get(key, 0)
        if n:
            c.setFillColor(color)
            c.setFont("Helvetica-Bold" if key == "cancelado" else "Helvetica", 7)
            c.drawString(6 * rl_mm, y, f"{n}x  {label}")
            y -= 6 * rl_mm

    c.setStrokeColor(HexColor("#cccccc"))
    c.setLineWidth(0.4)
    c.line(0, 2 * rl_mm, LW, 2 * rl_mm)
    c.showPage()

    # ── UMA ETIQUETA POR PEDIDO ───────────────────────────────────────────────
    for p in pedidos:
        cat     = _cat(p)
        qtd     = int(p.get("quantidade") or 1)
        on      = p.get("order_number", "")
        produto = (p.get("product_name") or "")[:38]
        sku     = p.get("sku", "")

        # Fundo levemente colorido para situações especiais
        if cat == "cancelado":
            c.setFillColor(HexColor("#FFF5F5"))
            c.rect(0, 0, LW, LH, fill=1, stroke=0)
        elif cat == "reservar":
            c.setFillColor(HexColor("#FFFBEB"))
            c.rect(0, 0, LW, LH, fill=1, stroke=0)

        # Badge topo-direito
        badge = _BADGE.get(cat)
        if badge:
            fg, bg, txt = badge
            bw, bh = 26 * rl_mm, 6 * rl_mm
            bx = LW - bw - 3 * rl_mm
            by = topo - 1 * rl_mm
            c.setFillColor(bg)
            c.roundRect(bx, by, bw, bh, 1.5 * rl_mm, fill=1, stroke=0)
            c.setFillColor(fg)
            c.setFont("Helvetica-Bold", 6.5)
            c.drawCentredString(bx + bw / 2, by + 1.8 * rl_mm, txt)

        # Cabeçalho
        c.setFont("Helvetica", 5.5)
        c.setFillColor(rl_colors.grey)
        c.drawString(4 * rl_mm, topo, f"PICKLIST  {ts}")

        # Número do pedido
        c.setFont("Helvetica-Bold", 12)
        c.setFillColor(HexColor("#991B1B") if cat == "cancelado" else rl_colors.black)
        c.drawString(4 * rl_mm, topo - 8 * rl_mm, on)

        # Linha de aviso (cancelado / reservar) ou produto+sku (normal/retirada)
        if cat == "cancelado":
            c.setFont("Helvetica-Bold", 7)
            c.setFillColor(HexColor("#991B1B"))
            c.drawString(4 * rl_mm, topo - 17 * rl_mm, "NAO PROCESSAR — PEDIDO CANCELADO")
        elif cat == "reservar":
            c.setFont("Helvetica", 7)
            c.setFillColor(HexColor("#92400E"))
            c.drawString(4 * rl_mm, topo - 17 * rl_mm, "Aguardando liberacao — nao processar ainda")
        else:
            c.setFont("Helvetica", 8)
            c.setFillColor(rl_colors.black)
            c.drawString(4 * rl_mm, topo - 17 * rl_mm, produto)
            c.setFont("Helvetica", 7)
            c.setFillColor(rl_colors.grey)
            c.drawString(4 * rl_mm, topo - 24 * rl_mm, f"SKU: {sku}")

        # Quantidade
        if qtd > 1:
            c.setFont("Helvetica-Bold", 9)
            c.setFillColor(HexColor("#b45309"))
            c.drawRightString(LW - 4 * rl_mm, topo - 8 * rl_mm, f"QTD {qtd}x")

        # Linha separadora
        c.setStrokeColor(HexColor("#cccccc"))
        c.setLineWidth(0.4)
        c.line(0, 2 * rl_mm, LW, 2 * rl_mm)
        c.showPage()

    c.save()
    log(f"[picklist] 📄 PDF gerado: {out.name} ({len(pedidos)} pedidos, 10x5cm)")
    return out


def _imprimir_picklist_supabase(config: dict) -> bool:
    """Gera e imprime a picklist da rodada atual a partir dos dados do Supabase.
    Pedidos ativos: incluídos uma vez (marcados com picklist_impresso_em).
    Para Reservar: incluídos em toda picklist como aviso (não marcados — status muda quando ML liberar).
    Cancelados: incluídos uma vez como aviso (marcados)."""
    try:
        supa  = create_client(*_supa())
        token = config.get("token", "")

        # Pedidos ativos ainda não na picklist
        r = (supa.table("pedidos")
             .select("order_number,product_name,sku,quantidade,status_upseller,label_url")
             .eq("status", "ativo")
             .eq("cliente", token)
             .is_("picklist_impresso_em", "null")
             .order("data", desc=False)
             .execute())
        ativos = r.data or []

        # Cancelados ainda não avisados na picklist
        r_c = (supa.table("pedidos")
               .select("order_number,product_name,sku,quantidade,status_upseller")
               .eq("status", "cancelado")
               .eq("cliente", token)
               .is_("picklist_impresso_em", "null")
               .order("data", desc=False)
               .execute())
        cancelados = r_c.data or []
        for p in cancelados:
            p["status"] = "cancelado"

        # Para Reservar sempre aparece (não marcados) — ficam até ML liberar
        para_reservar = [p for p in ativos
                         if (p.get("status_upseller") or "").strip().lower() == "para reservar"]
        # Demais ativos: marcados após impressão
        para_marcar_ativos = [p for p in ativos
                              if (p.get("status_upseller") or "").strip().lower() != "para reservar"]

        pedidos = ativos + cancelados  # tudo junto — _gerar_picklist_pdf ordena por categoria
    except Exception as e:
        log(f"[picklist] ⚠️ Erro ao consultar Supabase: {e}")
        return False

    if not pedidos:
        log("[picklist] ℹ️ Nenhum pedido novo — picklist não gerada")
        return True

    n_reservar   = len(para_reservar)
    n_cancelados = len(cancelados)
    n_processar  = len(pedidos) - n_reservar - n_cancelados
    log(f"[picklist] 📋 {len(pedidos)} pedido(s): {n_processar} processar, "
        f"{n_reservar} reservar, {n_cancelados} cancelados")

    margem_topo_mm = float(config.get("margem_topo_mm", 5) or 0)
    try:
        pdf_pick = _gerar_picklist_pdf(pedidos, margem_topo_mm=margem_topo_mm)
    except Exception as e:
        log(f"[picklist] ❌ Erro ao gerar PDF: {e}")
        return False

    # Imprime se houver impressora; caso contrário fica disponível via /picklist-pdf
    sistema        = platform.system()
    nome_impressora = config.get("nome_impressora", "").strip()
    tem_impressora  = bool(nome_impressora) or _verificar_impressora()
    impressora_usar = nome_impressora or (_nome_impressora_padrao() if sistema == "Darwin" else "")

    if tem_impressora:
        try:
            if sistema == "Darwin":
                cmd = (["lp", "-d", impressora_usar, str(pdf_pick)]
                       if impressora_usar else ["lp", str(pdf_pick)])
                subprocess.run(cmd, timeout=30, check=True)
                log(f"[picklist] 🖨️ Enviado para {impressora_usar or 'impressora padrão'}")
            elif sistema == "Windows":
                _imprimir_windows(pdf_pick, impressora_usar)
                log(f"[picklist] 🖨️ Enviado para {impressora_usar or 'impressora padrão'}")
        except Exception as e:
            log(f"[picklist] ⚠️ Erro ao imprimir: {e} — PDF disponível no PCP")
    else:
        log("[picklist] 📋 Sem impressora — PDF disponível via PCP (/picklist-pdf)")

    # Marca ativos (exceto Para Reservar) + cancelados — Para Reservar fica sem marca
    # para reaparecer na próxima picklist quando o ML liberar
    try:
        agora    = datetime.utcnow().isoformat()
        order_ns = [p["order_number"] for p in para_marcar_ativos + cancelados]
        for i in range(0, len(order_ns), 50):
            (supa.table("pedidos")
             .update({"picklist_impresso_em": agora})
             .in_("order_number", order_ns[i:i+50])
             .execute())
        log(f"[picklist] ✅ {len(order_ns)} pedidos marcados ({n_reservar} reservar não marcados) — {agora[:16]}")
    except Exception as e:
        log(f"[picklist] ⚠️ Erro ao marcar Supabase: {e}")

    return True


def _gerar_pdf_multicopia(pdf_path: "Path", n: int) -> "Path":
    """Cria um PDF com N cópias da mesma página (para pedidos multi-unidade)."""
    if n <= 1:
        return pdf_path
    out = pdf_path.with_stem(pdf_path.stem + f"_x{n}")
    if out.exists():
        return out
    try:
        from pypdf import PdfReader, PdfWriter
        reader = PdfReader(str(pdf_path))
        writer = PdfWriter()
        pagina = reader.pages[0]
        for _ in range(n):
            writer.add_page(pagina)
        with open(str(out), "wb") as f:
            writer.write(f)
        log(f"[etiqueta] 🗂️ {n}x cópias geradas → {out.name}")
        return out
    except Exception as e:
        log(f"[etiqueta] ⚠️ _gerar_pdf_multicopia: {e}")
        return pdf_path


# ── Playwright: imprimir picklist (mantido como fallback) ─────────────────────

async def _imprimir_picklist_playwright(config: dict):
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        log("[picklist] ❌ Playwright não instalado")
        return False

    _limpar_crash_chrome()
    log("[picklist] 🗒️ Abrindo UpSeller — Para Emitir...")

    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            user_data_dir=str(PASTA_SESSAO),
            headless=True,
            args=["--window-size=1280,900"],
        )
        try:
            page = context.pages[0] if context.pages else await context.new_page()

            await page.goto(
                "https://app.upseller.com/pt/order/pending-invoice",
                wait_until="domcontentloaded",
                timeout=60_000,
            )
            await page.wait_for_timeout(3000)

            if "/login" in page.url:
                log("[picklist] ⚠️ Sessão expirada. Abrindo para login...")
                await page.goto("https://app.upseller.com/login", wait_until="domcontentloaded", timeout=60_000)
                log("[picklist] 🔑 Faça login no UpSeller. Aguardando...")
                await page.wait_for_url(lambda url: "/login" not in url, timeout=600_000)
                log("[picklist] ✅ Login realizado!")
                await page.wait_for_load_state("networkidle", timeout=30_000)
                await page.wait_for_timeout(3000)
                await page.goto(
                    "https://app.upseller.com/pt/order/pending-invoice",
                    wait_until="domcontentloaded",
                    timeout=60_000,
                )
                await page.wait_for_timeout(3000)

            # Fecha qualquer popup/aviso (Avisos, NF-e, etc.)
            await _fechar_popups_upseller(page)

            # Verifica se há pedidos antes de tentar imprimir (aguarda a tabela carregar)
            await page.wait_for_timeout(2000)
            linhas_pick = await page.locator("tbody tr").count()
            if linhas_pick == 0:
                log("[picklist] ℹ️ Para Emitir vazio — picklist não gerada")
                return True
            log(f"[picklist] 📋 {linhas_pick} pedido(s) para imprimir")

            # Seleciona todos — o UpSeller já filtra: só mostra o que ainda está em "Para Emitir"
            # Pedidos já processados saem da lista automaticamente
            log("[picklist] ☑️ Selecionando todos os pedidos...")
            header_cb = page.locator("th .ant-checkbox-input").first
            await header_cb.click()
            await page.wait_for_timeout(1500)

            # Detecta impressora padrão no sistema
            tem_impressora = _verificar_impressora()
            if tem_impressora:
                log(f"[picklist] 🖨️ Impressora detectada — imprimindo direto...")
            else:
                log("[picklist] 📄 Sem impressora — gerando PDF...")

            # Intercepta o popup que o UpSeller abre com o conteúdo da picklist
            async with context.expect_page() as popup_info:
                await page.get_by_text("Imprimir em Massa").first.click()
                await page.wait_for_timeout(800)
                await page.get_by_text("Imprimir Lista de Separação").first.click()

            popup = await popup_info.value
            await popup.wait_for_load_state("networkidle", timeout=30_000)
            await popup.wait_for_timeout(1500)

            nome_pick = datetime.now().strftime("%Y%m%d_%H%M")
            await _imprimir_ou_pdf(popup, tem_impressora, nome_pick, "picklist")
            log("[picklist] ✅ Picklist processado!")

            # Marca todos os pedidos não impressos do cliente como impressos agora
            try:
                supa  = create_client(*_supa())
                agora = datetime.utcnow().isoformat()
                supa.from_("pedidos").update({"picklist_impresso_em": agora}) \
                    .is_("picklist_impresso_em", "null") \
                    .eq("cliente", config["token"]).execute()
                log(f"[picklist] 📋 Supabase atualizado — todos os pendentes marcados: {agora[:16]}")
            except Exception as ex:
                log(f"[picklist] ⚠️ Não foi possível registrar no Supabase: {ex}")

            return True

        except Exception as e:
            log(f"[picklist] ❌ Erro: {e}")
            return False
        finally:
            await context.close()


# ── Rota: marcar picklist como impressa (chamada pelo PCP após imprimir) ──────
@app.route("/marcar-picklist-impresso", methods=["POST", "OPTIONS"])
def marcar_picklist_impresso():
    if request.method == "OPTIONS":
        resp = jsonify({})
        resp.headers["Access-Control-Allow-Origin"]  = "*"
        resp.headers["Access-Control-Allow-Methods"] = "POST"
        resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
        return resp
    data         = request.get_json() or {}
    order_numbers = data.get("order_numbers", [])
    if not order_numbers:
        r = jsonify({"ok": False, "erro": "Nenhum pedido informado"})
        r.headers["Access-Control-Allow-Origin"] = "*"
        return r
    try:
        supa  = create_client(*_supa())
        agora = datetime.utcnow().isoformat()
        for i in range(0, len(order_numbers), 50):
            supa.table("pedidos").update({"picklist_impresso_em": agora}) \
                .in_("order_number", order_numbers[i:i+50]).execute()
        log(f"[picklist] ✅ {len(order_numbers)} pedido(s) marcados via PCP")
        r = jsonify({"ok": True, "marcados": len(order_numbers)})
        r.headers["Access-Control-Allow-Origin"] = "*"
        return r
    except Exception as e:
        r = jsonify({"ok": False, "erro": str(e)})
        r.headers["Access-Control-Allow-Origin"] = "*"
        return r


# ── Rota: imprimir picklist (legado — mantida para compatibilidade) ────────────
@app.route("/imprimir-picklist", methods=["POST", "OPTIONS"])
def imprimir_picklist():
    if request.method == "OPTIONS":
        resp = jsonify({})
        resp.headers["Access-Control-Allow-Origin"]  = "*"
        resp.headers["Access-Control-Allow-Methods"] = "POST"
        resp.headers["Access-Control-Allow-Headers"] = "Content-Type"
        return resp

    if not CONFIG_FILE.exists():
        r = jsonify({"ok": False, "erro": "Configure o robô primeiro"})
        r.headers["Access-Control-Allow-Origin"] = "*"
        return r

    config = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
    threading.Thread(
        target=_imprimir_picklist_thread,
        args=(config,),
        daemon=True
    ).start()

    r = jsonify({"ok": True})
    r.headers["Access-Control-Allow-Origin"] = "*"
    return r


@app.route("/picklist-disponivel", methods=["GET"])
def picklist_disponivel():
    arquivos = sorted(PASTA_PICKLISTS.glob("picklist_*.pdf"))
    r = jsonify({"disponivel": len(arquivos) > 0, "total": len(arquivos)})
    r.headers["Access-Control-Allow-Origin"] = "*"
    return r


@app.route("/etiqueta/<order_number>", methods=["GET"])
def servir_etiqueta(order_number):
    """Serve ou imprime diretamente a etiqueta da subpasta local / Supabase."""
    from flask import Response, request as _req
    import urllib.request as _ul
    pdf_path = PASTA_ETIQUETAS / f"etiqueta_{order_number}.pdf"

    # ?raw=1 → serve o PDF binário diretamente (para browser abrir/imprimir)
    if _req.args.get("raw"):
        if pdf_path.exists():
            conteudo = pdf_path.read_bytes()
            resp = Response(conteudo, mimetype="application/pdf")
            resp.headers["Access-Control-Allow-Origin"] = "*"
            return resp
        return jsonify({"erro": "PDF não encontrado"}), 404

    # Log imediato — aparece ANTES de qualquer verificação, inclusive para 404
    log(f"[etiqueta] 🔔 PCP solicitou: {order_number} | existe={pdf_path.exists()}")

    config = json.loads(CONFIG_FILE.read_text(encoding="utf-8")) if CONFIG_FILE.exists() else {}
    nome_impressora = config.get("nome_impressora", "").strip()

    def _obter_pdf() -> bool:
        if pdf_path.exists():
            _garantir_pagina_unica(pdf_path)
            return True
        try:
            supa = create_client(*_supa())
            r = (supa.table("pedidos")
                 .select("label_url")
                 .eq("order_number", order_number)
                 .limit(1).execute())
            rows = r.data or []
            url = rows[0].get("label_url") if rows else None
            if url and "127.0.0.1:5001" not in url and "localhost:5001" not in url:
                _ul.urlretrieve(url, str(pdf_path))
                _garantir_pagina_unica(pdf_path)
                return True
        except Exception:
            pass
        return False

    if not _obter_pdf():
        r = jsonify({"erro": "Etiqueta não disponível — execute o robô primeiro"})
        r.headers["Access-Control-Allow-Origin"] = "*"
        return r, 404

    # Valida que o arquivo é realmente um PDF
    try:
        header_bytes = pdf_path.read_bytes()[:8]
        if not header_bytes.startswith(b'%PDF'):
            log(f"[etiqueta] ❌ {order_number} — arquivo não é PDF (inicia com: {header_bytes[:20]})")
            pdf_path.unlink(missing_ok=True)
            r = jsonify({"erro": "Arquivo de etiqueta inválido — será baixado novamente"})
            r.headers["Access-Control-Allow-Origin"] = "*"
            return r, 500
    except Exception:
        pass

    # Tenta imprimir direto; se não conseguir, serve o PDF para o browser
    sistema = platform.system()
    # Se nome_impressora configurado manualmente, assume que existe sem detecção automática
    tem_impressora = bool(nome_impressora) or _verificar_impressora()
    # Resolve o nome da impressora a usar (configurado > padrão do sistema)
    impressora_usar = nome_impressora or (
        _nome_impressora_padrao() if sistema == "Darwin" else ""
    )
    log(f"[etiqueta] 📄 {order_number} | impressora={impressora_usar or '(padrão)'} | tem_impressora={tem_impressora} | sistema={sistema}")

    margem_topo_mm = float(config.get("margem_topo_mm", 4) or 0)
    margem_esq_mm  = float(config.get("margem_esq_mm",  4) or 0)

    if tem_impressora:
        try:
            pdf_imprimir = _corrigir_mediabox(pdf_path)
            pdf_imprimir = _adicionar_margem_topo(pdf_imprimir, margem_topo_mm, margem_esq_mm)
            log(f"[etiqueta] 📐 {order_number} | pdf={pdf_imprimir.name} | tamanho={pdf_imprimir.stat().st_size}B")
            if sistema == "Darwin":
                cmd = ["lp", "-d", impressora_usar, str(pdf_imprimir)] if impressora_usar else ["lp", str(pdf_imprimir)]
                log(f"[etiqueta] 🔄 {order_number} → lp cmd: {' '.join(cmd)}")
                result = subprocess.run(cmd, timeout=30, capture_output=True, text=True)
                if result.returncode != 0:
                    raise RuntimeError(f"lp falhou (code {result.returncode}): {result.stderr.strip()}")
                log(f"[etiqueta] 🖨️ {order_number} → {impressora_usar or 'impressora padrão'}")
            elif sistema == "Windows":
                log(f"[etiqueta] 🔄 {order_number} → SumatraPDF → {impressora_usar or 'padrão'}")
                _imprimir_windows(pdf_imprimir, impressora_usar)
                log(f"[etiqueta] 🖨️ {order_number} → {impressora_usar or 'impressora padrão'}")
            r = jsonify({"ok": True, "impresso": True, "impressora": impressora_usar})
            r.headers["Access-Control-Allow-Origin"] = "*"
            return r
        except Exception as e:
            log(f"[etiqueta] ⚠️ Erro ao imprimir: {e!r} — abrindo PDF no viewer")

    # Sem impressora: abre o PDF no sistema e retorna JSON para PCP confirmar
    log(f"[etiqueta] 📋 {order_number} — abrindo PDF no sistema (sem impressora)")
    try:
        pdf_imprimir = _corrigir_mediabox(pdf_path)
        if sistema == "Darwin":
            subprocess.Popen(["open", str(pdf_imprimir)])
        elif sistema == "Windows":
            subprocess.Popen(["start", "", str(pdf_imprimir)], shell=True)
    except Exception as e:
        log(f"[etiqueta] ⚠️ Não foi possível abrir PDF: {e}")
    r = jsonify({"ok": True, "impresso": False, "motivo": "sem_impressora",
                 "pdf": f"http://127.0.0.1:5001/etiqueta/{order_number}?raw=1"})
    r.headers["Access-Control-Allow-Origin"] = "*"
    return r


@app.route("/picklist-pdf", methods=["GET"])
def picklist_pdf():
    arquivos = sorted(PASTA_PICKLISTS.glob("picklist_*.pdf"))
    if not arquivos:
        r = jsonify({"erro": "Nenhum picklist disponível"})
        r.headers["Access-Control-Allow-Origin"] = "*"
        return r, 404
    pdf = arquivos[0]  # mais antigo primeiro
    conteudo = pdf.read_bytes()
    try:
        pdf.unlink()
    except Exception:
        pass
    from flask import Response
    resp = Response(conteudo, mimetype="application/pdf")
    resp.headers["Access-Control-Allow-Origin"] = "*"
    return resp


# ── Sincroniza pedidos novos (capturados entre imports) no Supabase ──────────

async def _sincronizar_pedidos_novos_supabase(page, order_numbers: list, config: dict):
    """Insere no Supabase pedidos que estão em Para Imprimir mas não foram importados ainda.
    Ocorre quando o background captura um pedido que chegou após o último import do dia.
    """
    import re as _re
    try:
        supa = create_client(*_supa())
        # Descobre quais ainda não estão no Supabase
        r = supa.table("pedidos").select("order_number").in_("order_number", order_numbers).execute()
        existentes = {row["order_number"] for row in (r.data or [])}
        faltando = [on for on in order_numbers if on not in existentes]

        # Existentes sem numero_plataforma (para backfill)
        sem_num: set = set()
        if existentes:
            r2 = supa.table("pedidos").select("order_number").in_("order_number", list(existentes)).is_("numero_plataforma", "null").execute()
            sem_num = {row["order_number"] for row in (r2.data or [])}

        if not faltando and not sem_num:
            return

        if faltando:
            log(f"[capturar] 📥 {len(faltando)} pedido(s) novos — importando para Supabase...")
        if sem_num:
            log(f"[capturar] 🔢 {len(sem_num)} pedido(s) sem numero_plataforma — atualizando...")

        PLAT = {"shopee": "Shopee", "shein": "Shein", "mercado": "Mercado Livre",
                "tiktok": "TikTok", "kwai": "Kwai"}
        token = config.get("token", "")
        hoje  = __import__("datetime").date.today().isoformat()

        # Busca detalhes de todos os Para Imprimir atuais via API (uma só chamada)
        all_orders = await page.evaluate("""async () => {
            let todos = [], page = 1;
            while (true) {
                const r = await fetch('/api/order/index', {
                    method:'POST', headers:{'Content-Type':'application/x-www-form-urlencoded'},
                    body:'timeType=0&searchType=0&sortName=1&sortValue=1&orderState=in_process&labelStatus=success&warehouseType=0&pageNum='+page+'&pageSize=50'
                });
                const j = await r.json();
                const list = (j.data||{}).list||[];
                todos = todos.concat(list);
                if (page * 50 >= ((j.data||{}).total||0) || !list.length) break;
                page++;
            }
            return todos;
        }""")

        novos_pedidos  = []
        novos_itens    = []
        atualizar_num  = []
        for order in all_orders:
            on = (order.get("orderNumber") or "").strip()
            num_plat = str(order.get("orderId") or order.get("extendedId") or "") or None
            if on in faltando:
                items = order.get("orderItemList") or []
                first = items[0] if items else {}
                plat  = PLAT.get((order.get("platform") or "").lower(), order.get("platform") or "")
                try:
                    valor = float(order.get("orderAmount") or 0) or None
                except Exception:
                    valor = None
                qtd = sum(int(i.get("productCount") or 1) for i in items) or 1
                novos_pedidos.append({
                    "order_number":      on,
                    "numero_plataforma": num_plat,
                    "sku":               (first.get("variationSku") or first.get("productSku") or "").strip(),
                    "product_name":      (first.get("productName") or "").strip(),
                    "image_url":         (first.get("productImg")  or "").strip(),
                    "quantidade":        qtd,
                    "plataforma":        plat,
                    "valor":             valor,
                    "data":              hoje,
                    "cliente":           token,
                    "label_url":         None,
                })
                for item in items:
                    novos_itens.append({
                        "order_number": on,
                        "sku":          (item.get("variationSku") or item.get("productSku") or "").strip(),
                        "product_name": (item.get("productName") or "").strip(),
                        "image_url":    (item.get("productImg")  or "").strip(),
                        "quantidade":   int(item.get("productCount") or 1),
                        "cliente":      token,
                        "data":         hoje,
                    })
            elif on in sem_num and num_plat:
                atualizar_num.append({"order_number": on, "numero_plataforma": num_plat})

        if novos_pedidos:
            supa.table("pedidos").upsert(novos_pedidos, on_conflict="order_number").execute()
        if novos_itens:
            supa.table("pedido_itens").insert(novos_itens).execute()
        if atualizar_num:
            for row in atualizar_num:
                try:
                    supa.table("pedidos").update({"numero_plataforma": row["numero_plataforma"]}).eq("order_number", row["order_number"]).execute()
                except Exception:
                    pass
            log(f"[capturar] ✅ {len(atualizar_num)} numero_plataforma preenchido(s)")
        log(f"[capturar] ✅ {len(novos_pedidos)} pedido(s) sincronizados com Supabase")
    except Exception as e:
        log(f"[capturar] ⚠️ Sync Supabase falhou: {e}")


# ── Playwright: capturar etiquetas retroativas (pedidos já processados) ──────

async def _capturar_retroativo_playwright(config: dict) -> None:
    """Captura etiquetas de pedidos (ontem/hoje) processados manualmente, sem label_url no Supabase.
    Navega por Para Imprimir e all-orders, pesquisa cada pedido e aciona Mais → Imprimir Etiqueta."""
    import re as _re_r, urllib.request as _ur_r
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        return
    try:
        from datetime import date as _date_r, timedelta as _td_r
        supa_r = create_client(*_supa())
        ontem_r = (_date_r.today() - _td_r(days=1)).isoformat()
        res_r = supa_r.table("pedidos").select("order_number") \
            .is_("label_url", "null") \
            .gte("data", ontem_r) \
            .execute()
        sem_label = [
            r["order_number"] for r in (res_r.data or [])
            if r.get("order_number")
            and not (PASTA_ETIQUETAS / f"etiqueta_{r['order_number']}.pdf").exists()
            and r["order_number"] not in _capturar_skip
        ]
        if not sem_label:
            return
        log(f"[retro] 🔍 {len(sem_label)} pedido(s) sem etiqueta — capturando via all-orders...")
    except Exception as e_q:
        log(f"[retro] ⚠️ Supabase query: {e_q}")
        return

    _limpar_crash_chrome()
    _cdn_r: dict = {}

    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            user_data_dir=str(PASTA_SESSAO),
            headless=True,
            args=["--window-size=1920,1080", "--disable-popup-blocking"],
            viewport={"width": 1920, "height": 1080},
        )
        try:
            page = context.pages[0] if context.pages else await context.new_page()
            await page.set_viewport_size({"width": 1920, "height": 1080})

            # Listener de CDN no contexto (captura URLs de qualquer popup aberto)
            async def _cdn_listener_r(pg):
                async def _on_resp_r(resp):
                    u = resp.url
                    if "print-label.upseller.cn" in u and ".pdf" in u.lower():
                        _cdn_r[id(pg)] = u
                pg.on("response", _on_resp_r)
            context.on("page", _cdn_listener_r)

            # Abas a tentar (em ordem de probabilidade)
            abas_r = [
                "https://app.upseller.com/pt/order/all-orders",
                "https://app.upseller.com/pt/order/to-ship",
                "https://app.upseller.com/pt/order/in-process",
            ]

            await page.goto(abas_r[0], wait_until="domcontentloaded", timeout=60_000)
            await page.wait_for_timeout(3000)

            if "/login" in page.url:
                log("[retro] ⚠️ Sessão expirada — faça login")
                return

            await _fechar_popups_upseller(page)
            capturados_r = 0

            # ── Fase 2 (PRIMEIRO): batch via Para Retirada / all-orders ──────────
            # Mais rápido: 20 pedidos por lote em vez de 1 por vez
            LOTE_B = 20
            cap_b = 0
            pendentes_b = list(sem_label)

            abas_batch = [
                ("to-pickup",  "https://app.upseller.com/pt/order/to-pickup"),
                ("pickup",     "https://app.upseller.com/pt/order/pickup"),
                ("all-orders", "https://app.upseller.com/pt/order/all-orders"),
            ]

            for nome_aba_b, url_aba_b in abas_batch:
                if not pendentes_b or _parar.is_set():
                    break
                try:
                    await page.goto(url_aba_b, wait_until="domcontentloaded", timeout=30_000)
                    await page.wait_for_timeout(2000)
                    await _fechar_popups_upseller(page)
                except Exception:
                    continue

                inp_b = page.locator(
                    "input[placeholder*='Pedido'], input[placeholder*='vírgulas'], input[placeholder*='Número']"
                ).first
                if not await inp_b.count() or await page.locator("tbody tr").count() == 0:
                    continue

                for i_b in range(0, len(pendentes_b), LOTE_B):
                    if _parar.is_set():
                        break
                    lote_b = [
                        on for on in pendentes_b[i_b:i_b + LOTE_B]
                        if not (PASTA_ETIQUETAS / f"etiqueta_{on}.pdf").exists()
                    ]
                    if not lote_b:
                        continue

                    await page.goto(url_aba_b, wait_until="domcontentloaded", timeout=30_000)
                    await page.wait_for_timeout(1500)
                    await _fechar_popups_upseller(page)

                    inp_b = page.locator(
                        "input[placeholder*='Pedido'], input[placeholder*='vírgulas'], input[placeholder*='Número']"
                    ).first
                    if not await inp_b.count():
                        break
                    await inp_b.click()
                    await inp_b.fill(",".join(lote_b))
                    await inp_b.press("Enter")
                    await page.wait_for_timeout(2000)

                    visiveis_b = [
                        on for on in lote_b
                        if await page.locator("tbody tr").filter(has_text=on).count() > 0
                    ]
                    if not visiveis_b:
                        continue

                    log(f"[retro] 📋 {nome_aba_b}: {len(visiveis_b)} pedido(s) — batch...")

                    for on in visiveis_b:
                        cb_b = page.locator("tbody tr").filter(has_text=on).first \
                            .locator(".ant-checkbox-input, input[type='checkbox']").first
                        if await cb_b.count() > 0:
                            await cb_b.click()
                            await page.wait_for_timeout(60)
                    await page.wait_for_timeout(200)

                    pgs_b = set(id(pg) for pg in context.pages)
                    clicou_b = False
                    for txt_bulk_b in ["Imprimir em Massa", "Imprimir Etiquetas"]:
                        btn_bulk_b = page.locator("button, span").filter(has_text=txt_bulk_b).first
                        if await btn_bulk_b.count() > 0 and await btn_bulk_b.is_visible():
                            await btn_bulk_b.click(force=True)
                            await page.wait_for_timeout(400)
                            for sub_txt_b in ["Imprimir Etiquetas", "Imprimir Etiqueta"]:
                                sub_b = page.locator("li, .ant-dropdown-menu-item").filter(has_text=sub_txt_b).first
                                if await sub_b.count() > 0 and await sub_b.is_visible():
                                    await sub_b.click(force=True)
                                    clicou_b = True
                                    break
                            if not clicou_b:
                                await page.keyboard.press("Escape")
                            break

                    if not clicou_b:
                        await page.evaluate(
                            "() => document.querySelectorAll('tbody .ant-checkbox-checked "
                            ".ant-checkbox-input').forEach(c=>c.click())"
                        )
                        continue

                    novas_b = []
                    for _ in range(50):
                        novas_b = [pg for pg in context.pages if id(pg) not in pgs_b]
                        if novas_b:
                            break
                        await page.wait_for_timeout(200)

                    if not novas_b:
                        log(f"[retro] ⚠️ {nome_aba_b}: batch sem popup")
                        continue

                    if len(novas_b) == 1 and len(visiveis_b) > 1:
                        popup_b = novas_b[0]
                        for _ in range(20):
                            if id(popup_b) in _cdn_r or (popup_b.url and popup_b.url != "about:blank"):
                                break
                            await page.wait_for_timeout(300)
                        cdn_b = _cdn_r.pop(id(popup_b), None)
                        dir_b = popup_b.url if popup_b.url and popup_b.url != "about:blank" else None
                        url_batch_b = cdn_b or dir_b
                        await popup_b.close()
                        if url_batch_b:
                            tmp_b = PASTA_ETIQUETAS / f"_retro_batch_{i_b}.pdf"
                            try:
                                _ur_r.urlretrieve(url_batch_b, str(tmp_b))
                                salvos_b = _split_pdf_batch(tmp_b, visiveis_b)
                                for on_s in salvos_b:
                                    pp_s = PASTA_ETIQUETAS / f"etiqueta_{on_s}.pdf"
                                    if pp_s.exists():
                                        log(f"[retro] 📄 {on_s} — batch {nome_aba_b}")
                                        cap_b += 1
                                        if on_s in pendentes_b:
                                            pendentes_b.remove(on_s)
                                        try:
                                            pub_s = _upload_etiqueta_supabase(pp_s, on_s)
                                            lbl_s = pub_s or f"http://127.0.0.1:5001/etiqueta/{on_s}"
                                            create_client(*_supa()).table("pedidos") \
                                                .update({"label_url": lbl_s}).eq("order_number", on_s).execute()
                                        except Exception:
                                            pass
                            except Exception as e_split_b:
                                log(f"[retro] ⚠️ split {nome_aba_b}: {e_split_b}")
                            finally:
                                try:
                                    tmp_b.unlink(missing_ok=True)
                                except Exception:
                                    pass
                    else:
                        for popup_b, on_b in zip(novas_b, visiveis_b):
                            for _ in range(15):
                                if id(popup_b) in _cdn_r or (popup_b.url and popup_b.url != "about:blank"):
                                    break
                                await page.wait_for_timeout(300)
                            cdn_bi = _cdn_r.pop(id(popup_b), None)
                            dir_bi = popup_b.url if popup_b.url and popup_b.url != "about:blank" else None
                            url_bi = cdn_bi or dir_bi
                            await popup_b.close()
                            if url_bi:
                                pp_bi = PASTA_ETIQUETAS / f"etiqueta_{on_b}.pdf"
                                try:
                                    _ur_r.urlretrieve(url_bi, str(pp_bi))
                                    c_bi = _corrigir_mediabox(pp_bi)
                                    if c_bi != pp_bi and c_bi.exists():
                                        c_bi.replace(pp_bi)
                                    log(f"[retro] 📄 {on_b} — individual {nome_aba_b}")
                                    cap_b += 1
                                    if on_b in pendentes_b:
                                        pendentes_b.remove(on_b)
                                    try:
                                        pub_bi = _upload_etiqueta_supabase(pp_bi, on_b)
                                        lbl_bi = pub_bi or f"http://127.0.0.1:5001/etiqueta/{on_b}"
                                        create_client(*_supa()).table("pedidos") \
                                            .update({"label_url": lbl_bi}).eq("order_number", on_b).execute()
                                    except Exception:
                                        pass
                                except Exception as e_bi:
                                    log(f"[retro] ⚠️ {on_b}: {e_bi}")

            if cap_b > 0:
                log(f"[retro] ✅ {cap_b} etiqueta(s) capturadas via batch")

            # ── Fase 1: individual via dropdown (para os que restaram do batch) ──
            # Atualiza lista com os que ainda faltam após o batch
            sem_label_r1 = [
                on for on in sem_label
                if not (PASTA_ETIQUETAS / f"etiqueta_{on}.pdf").exists()
                and on not in _capturar_skip
            ]
            aba_atual = abas_r[0]
            if sem_label_r1:
                log(f"[retro] 🔍 {len(sem_label_r1)} restantes — tentando dropdown individual...")

            for on in sem_label_r1[:30]:
                if _parar.is_set():
                    break
                if (PASTA_ETIQUETAS / f"etiqueta_{on}.pdf").exists():
                    continue

                # Pesquisa o pedido em cada aba até encontrar
                row_r = None
                for aba in abas_r:
                    if aba != aba_atual:
                        await page.goto(aba, wait_until="domcontentloaded", timeout=30_000)
                        await page.wait_for_timeout(2000)
                        await _fechar_popups_upseller(page)
                        aba_atual = aba

                    # Tenta preencher o campo de busca
                    for sel in ["input[placeholder*='Pedido']", "input[placeholder*='pedido']",
                                ".ant-input-search input", "input[type='search']"]:
                        s = page.locator(sel).first
                        if await s.count() > 0:
                            await s.click(click_count=3)
                            await s.fill(on)
                            await page.keyboard.press("Enter")
                            await page.wait_for_timeout(1500)
                            break

                    row_r = page.locator("tbody tr").filter(has_text=on).first
                    if await row_r.count() > 0:
                        break

                if not row_r or await row_r.count() == 0:
                    log(f"[retro] ⚠️ {on} — não encontrado em nenhuma aba")
                    _capturar_skip.add(on)
                    continue

                pgs_r = set(id(pg) for pg in context.pages)
                try:
                    await row_r.hover(timeout=10_000)
                    await page.wait_for_timeout(500)

                    # Localiza o botão "Mais" da coluna de ação — usa o mais à direita na viewport
                    mais_btn = None
                    cands = page.locator("span, button").filter(has_text=_re_r.compile(r'^Mais'))
                    cnt = await cands.count()
                    max_x = -1
                    for i in range(cnt):
                        btn = cands.nth(i)
                        bb = await btn.bounding_box()
                        if bb and bb['x'] > max_x:
                            max_x = bb['x']
                            mais_btn = btn

                    if not mais_btn:
                        log(f"[retro] ⚠️ {on} — botão Mais não encontrado (cnt={cnt})")
                        _capturar_skip.add(on)
                        continue
                    await mais_btn.click(force=True)
                    await page.wait_for_timeout(600)

                    clicou_r = False
                    for txt_r in ["Imprimir Etiqueta", "Imprimir Etiq"]:
                        item_r = page.locator("li, .ant-dropdown-menu-item").filter(has_text=txt_r).first
                        if await item_r.count() > 0 and await item_r.is_visible():
                            await item_r.click(force=True)
                            clicou_r = True
                            break
                    if not clicou_r:
                        itens_vis = page.locator("li.ant-dropdown-menu-item")
                        txts = await itens_vis.all_inner_texts() if await itens_vis.count() > 0 else []
                        log(f"[retro] ⚠️ {on} — dropdown: {txts[:3]} — tentando detalhe do pedido...")
                        await page.keyboard.press("Escape")
                        await page.wait_for_timeout(400)

                        # Fallback: abre o detalhe do pedido clicando no número/link
                        # O painel de detalhe pode ter botão de imprimir diferente do dropdown
                        pgs_det = set(id(pg) for pg in context.pages)
                        abriu_detalhe = False
                        try:
                            # Clica no link com o número do pedido para abrir o painel lateral
                            link_det = page.locator(f"a:has-text('{on}'), td:has-text('{on}') a").first
                            if not await link_det.count():
                                link_det = row_r.locator("a").nth(1)  # segunda ancora (1a pode ser checkbox)
                            if await link_det.count():
                                await link_det.click(force=True)
                                # Aguarda painel lateral (.ant-drawer ou .ant-modal) aparecer
                                try:
                                    await page.wait_for_selector(".ant-drawer-open, .ant-modal-wrap:visible", timeout=4000)
                                except Exception:
                                    pass
                                await page.wait_for_timeout(1000)
                                abriu_detalhe = True
                        except Exception:
                            pass

                        if abriu_detalhe:
                            # Loga conteúdo do painel lateral para diagnóstico
                            painel = page.locator(".ant-drawer-open, .ant-modal-wrap:visible, .ant-drawer-content").first
                            if await painel.count() == 0:
                                painel = page
                            # Loga todos os botões
                            btns_vis = painel.locator("button, .ant-btn, a[class*='btn'], li.ant-dropdown-menu-item").filter(has_not_text="")
                            btns_txts = []
                            for _bi in range(min(await btns_vis.count(), 15)):
                                try:
                                    _bt = await btns_vis.nth(_bi).inner_text()
                                    if _bt.strip():
                                        btns_txts.append(_bt.strip()[:25])
                                except Exception:
                                    pass
                            # Loga título/header do painel para confirmar o que está aberto
                            titulo_painel = ""
                            for sel_tit in [".ant-drawer-title", ".ant-modal-title", "h1, h2, h3, h4"]:
                                el_tit = painel.locator(sel_tit).first
                                if await el_tit.count() > 0:
                                    titulo_painel = (await el_tit.inner_text())[:40]
                                    break
                            log(f"[retro] 🔎 {on} painel='{titulo_painel}' botões={btns_txts[:8]}")

                            # Tenta "Mais Ações" no painel (pode ter "Imprimir Etiqueta" no dropdown)
                            mais_acoes = painel.locator("button, .ant-btn, span").filter(has_text="Mais Ações").first
                            if await mais_acoes.count() > 0 and await mais_acoes.is_visible():
                                await mais_acoes.click(force=True)
                                await page.wait_for_timeout(600)
                                itens_ma = page.locator("li.ant-dropdown-menu-item, .ant-menu-item")
                                txts_ma = []
                                for _mi in range(min(await itens_ma.count(), 8)):
                                    try:
                                        _mt = await itens_ma.nth(_mi).inner_text()
                                        if _mt.strip():
                                            txts_ma.append(_mt.strip()[:25])
                                    except Exception:
                                        pass
                                log(f"[retro] 🔎 {on} Mais Ações={txts_ma}")
                                await page.keyboard.press("Escape")
                                await page.wait_for_timeout(300)

                            # Procura botão de imprimir dentro do painel
                            clicou_det = False
                            for txt_det in ["Imprimir Etiqueta", "Imprimir Etiq", "Print Label", "Etiqueta", "Imprimir"]:
                                btn_det = painel.locator("button, a, span").filter(has_text=txt_det).last
                                if await btn_det.count() > 0 and await btn_det.is_visible():
                                    await btn_det.click(force=True)
                                    await page.wait_for_timeout(1500)
                                    clicou_det = True
                                    break

                            if clicou_det:
                                # Verifica se abriu popup com PDF
                                novas_det = [pg for pg in context.pages if id(pg) not in pgs_det]
                                if novas_det:
                                    popup_det = novas_det[-1]
                                    for _ in range(15):
                                        if id(popup_det) in _cdn_r or (popup_det.url and popup_det.url != "about:blank"):
                                            break
                                        await page.wait_for_timeout(300)
                                    cdn_det = _cdn_r.pop(id(popup_det), None)
                                    dir_det = popup_det.url if popup_det.url and popup_det.url != "about:blank" else None
                                    url_det = cdn_det or dir_det
                                    await popup_det.close()
                                    if url_det:
                                        pp_det = PASTA_ETIQUETAS / f"etiqueta_{on}.pdf"
                                        try:
                                            _ur_r.urlretrieve(url_det, str(pp_det))
                                            c_det = _corrigir_mediabox(pp_det)
                                            if c_det != pp_det and c_det.exists():
                                                c_det.replace(pp_det)
                                            log(f"[retro] 📄 {on} — etiqueta capturada via detalhe")
                                            capturados_r += 1
                                            try:
                                                pub_det = _upload_etiqueta_supabase(pp_det, on)
                                                lbl_det = pub_det or f"http://127.0.0.1:5001/etiqueta/{on}"
                                                create_client(*_supa()).table("pedidos") \
                                                    .update({"label_url": lbl_det}).eq("order_number", on).execute()
                                            except Exception:
                                                pass
                                            # Fecha detalhe e volta para lista
                                            await _fechar_popups_upseller(page)
                                            continue
                                        except Exception as e_det:
                                            log(f"[retro] ⚠️ {on} detalhe download: {e_det}")

                            # Fecha painel de detalhe e força re-navegação no próximo pedido
                            await _fechar_popups_upseller(page)
                            await page.keyboard.press("Escape")
                            await page.wait_for_timeout(500)
                            aba_atual = ""

                        _capturar_skip.add(on)
                        continue

                    popup_r = None
                    for _ in range(20):
                        await page.wait_for_timeout(400)
                        novas_r = [pg for pg in context.pages if id(pg) not in pgs_r]
                        if novas_r:
                            popup_r = novas_r[-1]
                            break

                    if popup_r:
                        for _ in range(15):
                            if id(popup_r) in _cdn_r:
                                break
                            await page.wait_for_timeout(300)
                        cdn_url_r = _cdn_r.pop(id(popup_r), None)
                        dir_url_r = popup_r.url if popup_r.url and popup_r.url != "about:blank" else None
                        url_r = cdn_url_r or dir_url_r
                        await popup_r.close()
                        if url_r:
                            pp_r = PASTA_ETIQUETAS / f"etiqueta_{on}.pdf"
                            try:
                                _ur_r.urlretrieve(url_r, str(pp_r))
                                c_r = _corrigir_mediabox(pp_r)
                                if c_r != pp_r and c_r.exists():
                                    c_r.replace(pp_r)
                                log(f"[retro] 📄 {on} — etiqueta capturada")
                                capturados_r += 1
                                try:
                                    pub_r = _upload_etiqueta_supabase(pp_r, on)
                                    lbl_r = pub_r or f"http://127.0.0.1:5001/etiqueta/{on}"
                                    create_client(*_supa()).table("pedidos") \
                                        .update({"label_url": lbl_r}).eq("order_number", on).execute()
                                except Exception:
                                    pass
                            except Exception as e_dl:
                                log(f"[retro] ⚠️ {on} download: {e_dl}")
                        else:
                            log(f"[retro] ⚠️ {on} — popup sem URL CDN")
                            _capturar_skip.add(on)
                    else:
                        log(f"[retro] ⚠️ {on} — popup não abriu")
                        _capturar_skip.add(on)

                except Exception as e_r:
                    log(f"[retro] ⚠️ {on}: {e_r}")
                    _capturar_skip.add(on)
                    await _fechar_popups_upseller(page)
                    aba_atual = ""  # força re-navegação no próximo pedido

            if capturados_r > 0:
                log(f"[retro] ✅ {capturados_r} etiqueta(s) retroativas capturadas")

            # ── Fase 3: re-verifica o que ainda falta após Fase 2 + Fase 1 ──────
            try:
                res_r2 = create_client(*_supa()).table("pedidos").select("order_number") \
                    .is_("label_url", "null").gte("data", ontem_r).execute()
                faltam_r2 = [
                    r["order_number"] for r in (res_r2.data or [])
                    if r.get("order_number")
                    and not (PASTA_ETIQUETAS / f"etiqueta_{r['order_number']}.pdf").exists()
                ]
            except Exception:
                faltam_r2 = []

            if faltam_r2:
                log(f"[retro] ℹ️ {len(faltam_r2)} pedido(s) ainda aguardando etiqueta da plataforma")
                _ultima_rodada["sem_etiqueta"] = len(faltam_r2)

        except Exception as e_main:
            log(f"[retro] ❌ {e_main}")
        finally:
            await context.close()


# ── Playwright: capturar etiquetas em lote (pré-download) ────────────────────

async def _capturar_etiquetas_playwright(config: dict, aguardar_se_vazio: bool = True, background: bool = False) -> bool:
    import re, urllib.request
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        log("[capturar] ❌ Playwright não instalado")
        return False

    _limpar_crash_chrome()

    # Corrige etiquetas A4 existentes na pasta (baixadas antes da correção automática)
    try:
        import re as _re2, zlib as _zlib2
        corrigidos_lote = 0
        for _pdf in PASTA_ETIQUETAS.glob("etiqueta_*.pdf"):
            if "_c" in _pdf.stem:
                continue
            try:
                _d = _pdf.read_bytes()
                _m = re.search(rb'/MediaBox\s*\[([^\]]+)\]', _d)
                if not _m:
                    for _s in re.findall(rb'stream\r?\n(.*?)\r?\nendstream', _d, re.DOTALL):
                        try:
                            _dec = zlib.decompress(_s)
                            _m = re.search(rb'/MediaBox\s*\[([^\]]+)\]', _dec)
                            if _m: break
                        except Exception: pass
                if _m:
                    _v = list(map(float, _m.group(1).split()))
                    _w = _v[2] - _v[0]
                    if _w > 368:  # maior que 130mm → precisa corrigir
                        _corr = _corrigir_mediabox(_pdf)
                        if _corr != _pdf and _corr.exists():
                            _corr.replace(_pdf)
                            corrigidos_lote += 1
            except Exception:
                pass
        if corrigidos_lote > 0:
            log(f"[capturar] ✂️ {corrigidos_lote} etiqueta(s) A4 corrigidas para 100×150mm")
    except Exception:
        pass

    log("[capturar] 🏷️ Verificando etiquetas para pré-download...")

    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            user_data_dir=str(PASTA_SESSAO),
            headless=True,
            args=["--window-size=1280,900", "--disable-popup-blocking"],
        )
        try:
            page = context.pages[0] if context.pages else await context.new_page()

            # Intercepta PDF do CDN UpSeller no popup — evita "Dados em processamento..."
            # O popup faz GET https://print-label.upseller.cn/pdf/YYYY-MM-DD/hash.pdf (resp 200)
            # Basta capturar essa URL e baixar diretamente.
            _cdn_pdfs = {}  # id(popup_page) → cdn_url

            async def _cdn_listener(pg):
                async def _on_resp(resp):
                    u = resp.url
                    if "print-label.upseller.cn" in u and ".pdf" in u.lower():
                        _cdn_pdfs[id(pg)] = u
                pg.on("response", _on_resp)

            context.on("page", _cdn_listener)

            await page.goto(
                "https://app.upseller.com/pt/order/in-process",
                wait_until="domcontentloaded",
                timeout=60_000,
            )
            await page.wait_for_timeout(3000)

            if "/login" in page.url:
                log("[capturar] ⚠️ Sessão expirada — faça login manualmente")
                return False

            await _fechar_popups_upseller(page)

            # Retry: aguarda até 90s após NFe+Envio (pedidos aparecem em <30s normalmente)
            #        ou retorna imediatamente se vazio (captura inicial ou background)
            linhas = 0
            max_tentativas = 3 if aguardar_se_vazio else 1
            for tentativa in range(max_tentativas):
                linhas = await page.locator("tbody tr").count()
                if linhas > 0:
                    break
                if tentativa < max_tentativas - 1:
                    log(f"[capturar] ⏳ Para Imprimir vazio — aguardando 30s ({tentativa+1}/{max_tentativas-1})...")
                    await page.wait_for_timeout(30_000)
                    await page.reload(wait_until="domcontentloaded")
                    await page.wait_for_timeout(2000)
                    await _fechar_popups_upseller(page)

            if linhas == 0:
                log("[capturar] ℹ️ Nenhum pedido em Para Imprimir após aguardar")
                return True

            # Busca TODOS os números via API (sem limite de paginação do browser)
            api_orders = await page.evaluate("""async () => {
                let todos = [], pg = 1;
                while (true) {
                    const r = await fetch('/api/order/index', {
                        method: 'POST',
                        headers: {'Content-Type': 'application/x-www-form-urlencoded'},
                        body: 'timeType=0&searchType=0&sortName=1&sortValue=1&orderState=in_process&pageNum=' + pg + '&pageSize=200'
                    });
                    const j = await r.json();
                    const total = (j.data || {}).total || 0;
                    const list = (j.data || {}).list || [];
                    todos = todos.concat(list.map(o => o.orderNumber || '').filter(Boolean));
                    if (!list.length || todos.length >= total) break;
                    pg++;
                }
                return todos;
            }""")

            # Extrai também da tabela visível (garante que inclui pedidos visíveis mesmo se API diferir)
            all_texts = await page.locator("tbody tr").all_inner_texts()
            order_numbers = list(api_orders) if api_orders else []
            for txt in all_texts:
                m = re.search(r'UP[0-9][A-Z]{3}[0-9]{6}', txt)
                if m and m.group() not in order_numbers:
                    order_numbers.append(m.group())
            log(f"[capturar] 📋 API: {len(api_orders)} pedido(s) em Para Imprimir")

            if not order_numbers:
                log("[capturar] ⚠️ Não foi possível extrair números de pedido da tabela")
                return False

            # Garante que todos os pedidos visíveis estão no Supabase
            # (pedidos capturados pelo background entre imports ficam de fora)
            await _sincronizar_pedidos_novos_supabase(page, order_numbers, config)

            # Separa já capturados (cache) dos que faltam
            pendentes = [on for on in order_numbers
                         if not (PASTA_ETIQUETAS / f"etiqueta_{on}.pdf").exists()
                         and (not background or on not in _capturar_skip)]
            capturados = len(order_numbers) - len(pendentes)
            if capturados:
                log(f"[capturar] ✅ {capturados} já em cache")

            log(f"[capturar] 📋 {len(pendentes)} pedido(s) para capturar")

            async def _salvar_pdf(on: str, url: str, cdn: bool):
                nonlocal capturados
                pp = PASTA_ETIQUETAS / f"etiqueta_{on}.pdf"
                try:
                    urllib.request.urlretrieve(url, str(pp))
                    _garantir_pagina_unica(pp)  # descarta páginas extras de PDFs batch individuais
                    c = _corrigir_mediabox(pp)
                    if c != pp and c.exists():
                        c.replace(pp)
                    log(f"[capturar] 📄 {on} — PDF salvo ({'CDN' if cdn else 'URL'})")
                    capturados += 1
                    if on in pendentes:
                        pendentes.remove(on)
                    try:
                        url_pub = _upload_etiqueta_supabase(pp, on)
                        label_url = url_pub or f"http://127.0.0.1:5001/etiqueta/{on}"
                        create_client(*_supa()).table("pedidos").update({"label_url": label_url}) \
                            .eq("order_number", on).execute()
                    except Exception:
                        pass
                    return True
                except Exception as e:
                    log(f"[capturar] ⚠️ {on} — erro ao baixar: {e}")
                    return False

            async def _clicar_bulk():
                bulk_btn = page.locator("button, span").filter(has_text="Imprimir Etiquetas").first
                if not await bulk_btn.count():
                    return False
                await bulk_btn.click(force=True)
                await page.wait_for_timeout(400)
                for opcao in ["Imprimir Etiquetas", "Imprimir Etiqueta"]:
                    item = page.locator("li, .ant-dropdown-menu-item").filter(has_text=opcao).first
                    if await item.count() > 0 and await item.is_visible():
                        await item.click(force=True)
                        return True
                return False

            async def _desmarcar_todos():
                await page.evaluate(
                    "() => document.querySelectorAll('tbody .ant-checkbox-checked .ant-checkbox-input').forEach(c=>c.click())"
                )
                await page.wait_for_timeout(100)

            async def _filtrar_por_numeros(pg, numeros: list):
                """Filtra a tabela de Para Imprimir pelos números de pedido (campo de busca)."""
                try:
                    inp = pg.locator("input[placeholder*='Pedido'], input[placeholder*='vírgulas']").first
                    if not await inp.count():
                        return False
                    await inp.click()
                    await inp.fill(",".join(numeros))
                    await inp.press("Enter")
                    await pg.wait_for_timeout(1500)
                    return True
                except Exception:
                    return False

            async def _limpar_filtro_numeros(pg):
                try:
                    # Limpa o campo de busca e volta para a lista completa
                    inp = pg.locator("input[placeholder*='Pedido'], input[placeholder*='vírgulas']").first
                    if await inp.count():
                        await inp.fill("")
                        await inp.press("Enter")
                        await pg.wait_for_timeout(800)
                except Exception:
                    pass

            # Loop: itera em lotes de pendentes filtrando pelo campo de busca
            rodada_cap = 0
            BATCH = 50  # UpSeller aceita ~50 números no campo de busca
            while pendentes and not _parar.is_set():
                if background and rodando:
                    log("[capturar] ⏸️ Executar iniciado — pausando captura automática")
                    break

                rodada_cap += 1
                fez_algo = False

                lote = pendentes[:BATCH]

                await page.goto(
                    "https://app.upseller.com/pt/order/in-process",
                    wait_until="domcontentloaded", timeout=60_000,
                )
                await page.wait_for_timeout(1500)
                await _fechar_popups_upseller(page)

                # Remove pendentes cujo PDF já existe
                for on in list(pendentes):
                    if (PASTA_ETIQUETAS / f"etiqueta_{on}.pdf").exists():
                        pendentes.remove(on)
                        capturados += 1
                        fez_algo = True

                # Filtra tabela pelos pendentes do lote atual
                filtrou = await _filtrar_por_numeros(page, lote)
                if filtrou:
                    log(f"[capturar] 🔍 Filtrando {len(lote)} pedidos pendentes...")

                # Identifica quais pendentes estão visíveis nesta carga
                visiveis = []
                for on in list(pendentes):
                    if _parar.is_set() or (background and rodando):
                        break
                    row_loc = page.locator("tbody tr").filter(has_text=on).first
                    if await row_loc.count() > 0:
                        visiveis.append(on)

                if not visiveis:
                    if fez_algo:
                        continue
                    break

                # ── MODO BATCH: seleciona todos, 1 clique bulk ────────────────
                for on in visiveis:
                    row_loc = page.locator("tbody tr").filter(has_text=on).first
                    cb = row_loc.locator(".ant-checkbox-input, input[type='checkbox']").first
                    if await cb.count() > 0:
                        await cb.click()
                        await page.wait_for_timeout(80)
                await page.wait_for_timeout(200)

                paginas_antes = set(id(pg) for pg in context.pages)
                batch_ok = await _clicar_bulk()

                if not batch_ok:
                    await _desmarcar_todos()
                    for on in visiveis:
                        _capturar_skip.add(on)
                        pendentes.remove(on)
                    break

                # Aguarda popups — espera N ou até 12s
                novas_pgs = []
                for _ in range(60):
                    novas_pgs = [pg for pg in context.pages if id(pg) not in paginas_antes]
                    if len(novas_pgs) >= len(visiveis):
                        break
                    await page.wait_for_timeout(200)

                if not novas_pgs:
                    # Batch falhou — processa um por um nesta mesma rodada
                    log(f"[capturar] ⚠️ Batch sem popup — modo individual ({len(visiveis)} pedidos)")
                    await _desmarcar_todos()
                    for on in visiveis:
                        if _parar.is_set() or (background and rodando):
                            break
                        pp = PASTA_ETIQUETAS / f"etiqueta_{on}.pdf"
                        if pp.exists():
                            pendentes.remove(on)
                            capturados += 1
                            fez_algo = True
                            continue
                        # Fechar modal pendente antes de clicar no checkbox
                        modal = page.locator("div[role='dialog'].ant-modal-wrap, .ant-modal-wrap")
                        if await modal.count() > 0:
                            await page.keyboard.press("Escape")
                            await page.wait_for_timeout(500)
                        row_loc = page.locator("tbody tr").filter(has_text=on).first
                        cb = row_loc.locator(".ant-checkbox-input, input[type='checkbox']").first
                        if await cb.count() > 0:
                            await cb.click()
                            await page.wait_for_timeout(200)
                        pgs_antes_ind = set(id(pg) for pg in context.pages)
                        if not await _clicar_bulk():
                            _capturar_skip.add(on)
                            pendentes.remove(on)
                            continue
                        popup = None
                        for _ in range(25):
                            await page.wait_for_timeout(400)
                            novas_ind = [pg for pg in context.pages if id(pg) not in pgs_antes_ind]
                            if novas_ind:
                                popup = novas_ind[-1]
                                break
                        if not popup:
                            log(f"[capturar] ⚠️ {on} — popup não abriu")
                            await _desmarcar_todos()
                            continue
                        for _ in range(20):
                            if id(popup) in _cdn_pdfs or (popup.url and popup.url != "about:blank"):
                                break
                            await page.wait_for_timeout(300)
                        cdn_url    = _cdn_pdfs.pop(id(popup), None)
                        direct_url = popup.url if popup.url and popup.url != "about:blank" else None
                        url        = cdn_url or direct_url
                        await popup.close()
                        await _desmarcar_todos()
                        if url:
                            ok = await _salvar_pdf(on, url, bool(cdn_url))
                            if ok:
                                fez_algo = True
                        else:
                            log(f"[capturar] ⏳ {on} — sem CDN URL")
                            pendentes.remove(on)
                            _capturar_skip.add(on)
                else:
                    # Batch OK — dá mais 500ms para CDN capturar todas as URLs
                    await page.wait_for_timeout(500)
                    log(f"[capturar] ⚡ Batch: {len(novas_pgs)} popup(s) aberto(s)")

                    if len(novas_pgs) == 1 and len(visiveis) > 1:
                        # UpSeller gerou 1 PDF combinado (batch multi-página) — split por pedido
                        popup = novas_pgs[0]
                        cdn_url    = _cdn_pdfs.pop(id(popup), None)
                        direct_url = popup.url if popup.url and popup.url != "about:blank" else None
                        url        = cdn_url or direct_url
                        await popup.close()
                        if url:
                            tmp = PASTA_ETIQUETAS / "_batch_tmp.pdf"
                            try:
                                urllib.request.urlretrieve(url, str(tmp))
                                n_pgs = _contar_paginas_pdf(tmp)
                                log(f"[capturar] 📦 Batch PDF: {n_pgs} página(s) → {len(visiveis)} pedido(s)")
                                salvos = _split_pdf_batch(tmp, visiveis)
                                for on in salvos:
                                    pp = PASTA_ETIQUETAS / f"etiqueta_{on}.pdf"
                                    if pp.exists():
                                        log(f"[capturar] 📄 {on} — PDF split do batch")
                                        capturados += 1
                                        if on in pendentes:
                                            pendentes.remove(on)
                                        fez_algo = True
                                        try:
                                            url_pub  = _upload_etiqueta_supabase(pp, on)
                                            label_url = url_pub or f"http://127.0.0.1:5001/etiqueta/{on}"
                                            create_client(*_supa()).table("pedidos") \
                                                .update({"label_url": label_url}) \
                                                .eq("order_number", on).execute()
                                        except Exception:
                                            pass
                            except Exception as e:
                                log(f"[capturar] ⚠️ Erro ao baixar batch PDF combinado: {e}")
                            finally:
                                try:
                                    tmp.unlink(missing_ok=True)
                                except Exception:
                                    pass
                        else:
                            log(f"[capturar] ⏳ Batch combinado — sem URL")
                    else:
                        for i, popup in enumerate(novas_pgs):
                            on = visiveis[i] if i < len(visiveis) else None
                            cdn_url    = _cdn_pdfs.pop(id(popup), None)
                            direct_url = popup.url if popup.url and popup.url != "about:blank" else None
                            url        = cdn_url or direct_url
                            await popup.close()
                            if not on:
                                continue
                            if url:
                                ok = await _salvar_pdf(on, url, bool(cdn_url))
                                if ok:
                                    fez_algo = True
                            else:
                                log(f"[capturar] ⏳ {on} — sem CDN URL")
                                if on in pendentes:
                                    pendentes.remove(on)
                                _capturar_skip.add(on)
                    await _desmarcar_todos()

                if not fez_algo and not visiveis:
                    break

            log(f"[capturar] ✅ {capturados}/{len(order_numbers)} etiqueta(s) prontas")
            _ultima_rodada.update({"etiquetas_ok": capturados, "etiquetas_total": len(order_numbers)})

            # ── Download etiquetas Shopee/ML via label_url no Supabase ──────────
            try:
                import urllib.request as _ur2
                supa2 = create_client(*_supa())
                # Busca pedidos com label_url que ainda não têm PDF local
                offset2 = 0
                baixados = 0
                erros = 0
                while True:
                    lote = supa2.table("pedidos").select("order_number,label_url,plataforma") \
                        .not_.is_("label_url", "null") \
                        .range(offset2, offset2 + 199).execute()
                    if not lote.data:
                        break
                    for row in lote.data:
                        on = row.get("order_number", "")
                        url_label = row.get("label_url", "")
                        if not on or not url_label:
                            continue
                        pdf_p = PASTA_ETIQUETAS / f"etiqueta_{on}.pdf"
                        if pdf_p.exists():
                            continue  # já baixada
                        try:
                            _ur2.urlretrieve(url_label, str(pdf_p))
                            _corr3 = _corrigir_mediabox(pdf_p)
                            if _corr3 != pdf_p and _corr3.exists():
                                _corr3.replace(pdf_p)
                            baixados += 1
                        except Exception:
                            erros += 1
                    if len(lote.data) < 200:
                        break
                    offset2 += 200
                if baixados > 0 or erros > 0:
                    log(f"[capturar] 📦 Shopee/ML: {baixados} etiqueta(s) baixadas, {erros} erro(s)")
            except Exception as e:
                log(f"[capturar] ⚠️ Shopee label_url download falhou: {e}")

            return True

        except Exception as e:
            log(f"[capturar] ❌ Erro: {e}")
            return False
        finally:
            await context.close()


# ── Playwright: marcar pedido como impresso (background) ─────────────────────

async def _marcar_impresso_playwright(config: dict, order_number: str) -> bool:
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        return False

    log(f"[marcar] 📋 Marcando {order_number} como impresso...")

    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            user_data_dir=str(PASTA_SESSAO),
            headless=True,
            args=["--window-size=1280,900"],
        )
        try:
            page = context.pages[0] if context.pages else await context.new_page()

            await page.goto(
                "https://app.upseller.com/pt/order/in-process",
                wait_until="domcontentloaded", timeout=60_000,
            )
            await page.wait_for_timeout(2000)

            if "/login" in page.url:
                return False

            await _fechar_popups_upseller(page)

            search = None
            for sel in ["input[placeholder*='Pedido']", "input[placeholder*='pedido']",
                        ".ant-input-search input", ".ant-input"]:
                loc = page.locator(sel).first
                if await loc.count() > 0:
                    search = loc
                    break
            if search is None:
                return False

            await search.click(force=True)
            await search.fill(order_number)
            await page.keyboard.press("Enter")
            await page.wait_for_timeout(2000)

            row = page.locator("tr").filter(has_text=order_number).first
            if await row.count() == 0:
                log(f"[marcar] ℹ️ {order_number} não está em Para Imprimir — já avançado")
                return True

            # Abre o popup da etiqueta para exibir o botão "Marcar como Impresso"
            imprimir = row.locator("a, button, span").filter(has_text="Imprimir Etiq").first
            if await imprimir.count() > 0:
                paginas_antes = set(id(pg) for pg in context.pages)
                await imprimir.click()
                for _ in range(25):
                    await page.wait_for_timeout(300)
                    novas = [pg for pg in context.pages if id(pg) not in paginas_antes]
                    if novas:
                        await novas[-1].close()
                        break

            await page.wait_for_timeout(600)
            marcar = page.locator("button", has_text="Marcar como Impresso").first
            if await marcar.count() > 0 and await marcar.is_visible():
                await marcar.click()
                await page.wait_for_timeout(500)
                log(f"[marcar] ✅ {order_number} → Para Retirada")
            return True

        except Exception as e:
            log(f"[marcar] ❌ Erro: {e}")
            return False
        finally:
            await context.close()


# ── Playwright: emitir etiqueta ──────────────────────────────────────────────

def _imprimir_pdf_local(pdf_path: Path, order_number: str, config: dict) -> dict:
    """Imprime PDF local, deleta após impressão e avança pedido em background."""
    tem_impressora = _verificar_impressora()
    try:
        if tem_impressora:
            if platform.system() == "Darwin":
                subprocess.Popen(["lp", str(pdf_path)])
                log(f"[etiqueta] 🖨️ {order_number} enviado para impressora")
            elif platform.system() == "Windows":
                # run() bloqueante: espera o spool terminar antes de deletar o arquivo
                subprocess.run(
                    ["powershell", "-Command",
                     f'Start-Process -FilePath "{pdf_path}" -Verb Print -Wait'],
                    creationflags=subprocess.CREATE_NO_WINDOW,
                    timeout=30
                )
                log(f"[etiqueta] 🖨️ {order_number} enviado para impressora")
        else:
            if platform.system() == "Darwin":
                subprocess.Popen(["open", str(pdf_path)])
            elif platform.system() == "Windows":
                subprocess.Popen(["start", str(pdf_path)], shell=True)
            else:
                subprocess.Popen(["xdg-open", str(pdf_path)])
    except Exception as e:
        log(f"[etiqueta] ⚠️ Erro ao imprimir: {e}")
    if tem_impressora:
        # lp espoola na hora — pode deletar imediatamente
        try:
            pdf_path.unlink()
        except Exception:
            pass
    else:
        # open/start é async — aguarda 10s para o leitor carregar antes de deletar
        def _deletar_apos_delay(p: Path):
            import time; time.sleep(10)
            try: p.unlink()
            except Exception: pass
        threading.Thread(target=_deletar_apos_delay, args=(pdf_path,), daemon=True).start()
    threading.Thread(
        target=lambda: asyncio.run(_marcar_impresso_playwright(config, order_number)),
        daemon=True
    ).start()
    return {"ok": True}


async def _emitir_etiqueta_playwright(config: dict, order_number: str) -> dict:
    import urllib.request as _urllib

    pdf_path = PASTA_ETIQUETAS / f"etiqueta_{order_number}.pdf"

    # ── Caminho 1: PDF já em cache local ─────────────────────────────────────
    if pdf_path.exists():
        log(f"[etiqueta] ⚡ Cache local — imprimindo {order_number} instantaneamente...")
        return _imprimir_pdf_local(pdf_path, order_number, config)

    # ── Caminho 2: label_url no Supabase → baixa direto sem abrir browser ────
    try:
        supa = create_client(*_supa())
        # limit(1) em vez de .single() — pedidos tem múltiplas linhas por order_number (1 por SKU)
        r = supa.table("pedidos").select("label_url").eq("order_number", order_number).limit(1).execute()
        rows = r.data or []
        url = rows[0].get("label_url") if rows else None
        if url:
            log(f"[etiqueta] ⚡ URL Supabase — baixando e imprimindo {order_number}...")
            _urllib.urlretrieve(url, str(pdf_path))
            _corr2 = _corrigir_mediabox(pdf_path)
            if _corr2 != pdf_path and _corr2.exists():
                _corr2.replace(pdf_path)
            return _imprimir_pdf_local(pdf_path, order_number, config)
    except Exception:
        pass

    # ── Caminho 3: Playwright (pedido sem etiqueta ainda) ─────────────────────
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        log(f"[etiqueta] ❌ Playwright não instalado")
        return {"ok": False, "erro": "Playwright não instalado"}

    log(f"[etiqueta] 🏷️ Emitindo: {order_number}...")

    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            user_data_dir=str(PASTA_SESSAO),
            headless=True,
            args=["--window-size=1920,1080"],
            viewport={"width": 1920, "height": 1080},
        )
        try:
            page = context.pages[0] if context.pages else await context.new_page()
            await page.set_viewport_size({"width": 1920, "height": 1080})

            await page.goto(
                "https://app.upseller.com/pt/order/in-process",
                wait_until="domcontentloaded",
                timeout=60_000,
            )
            await page.wait_for_timeout(1500)

            if "/login" in page.url:
                log("[etiqueta] ⚠️ Sessão expirada. Abrindo para login...")
                await context.close()
                context = await p.chromium.launch_persistent_context(
                    user_data_dir=str(PASTA_SESSAO),
                    headless=False,
                    args=["--window-size=1280,900"],
                )
                page = context.pages[0] if context.pages else await context.new_page()
                await page.goto("https://app.upseller.com/login", wait_until="domcontentloaded", timeout=60_000)
                log("[etiqueta] 🔑 Faça login no UpSeller. Aguardando...")
                await page.wait_for_url(lambda url: "/login" not in url, timeout=600_000)
                log("[etiqueta] ✅ Login realizado!")
                await page.wait_for_load_state("networkidle", timeout=30_000)
                await page.wait_for_timeout(3000)
                await page.goto(
                    "https://app.upseller.com/pt/order/in-process",
                    wait_until="domcontentloaded",
                    timeout=60_000,
                )
                await page.wait_for_timeout(3000)

            await _fechar_popups_upseller(page)

            # Busca o pedido na fila de impressão (Para Imprimir)
            log(f"[etiqueta] 🔍 Buscando {order_number} em Para Imprimir...")
            search = None
            for tentativa in range(3):
                for sel in [
                    "input[placeholder*='Pedido']",
                    "input[placeholder*='pedido']",
                    "input[placeholder*='Buscar']",
                    "input[placeholder*='buscar']",
                    ".ant-input-search input",
                    ".ant-input",
                ]:
                    loc = page.locator(sel).first
                    if await loc.count() > 0:
                        search = loc
                        break
                if search is None:
                    screenshot = PASTA_RAIZ / f"debug_etiqueta_{order_number}.png"
                    await page.screenshot(path=str(screenshot))
                    log(f"[etiqueta] ⚠️ Campo de busca não encontrado — screenshot: {screenshot.name}")
                    return {"ok": False, "erro": "Campo de busca não encontrado"}
                try:
                    await search.click(timeout=5_000)
                    break  # clicou com sucesso
                except Exception:
                    if tentativa == 2:
                        # Última tentativa: força o click ignorando overlays
                        try:
                            await search.click(force=True)
                            break
                        except Exception:
                            pass
                    log(f"[etiqueta] ⚠️ Click bloqueado (tentativa {tentativa+1}/3) — fechando popups...")
                    await _fechar_popups_upseller(page)
                    await page.wait_for_timeout(1000)
                    search = None  # reinicia busca do selector
            else:
                screenshot = PASTA_RAIZ / f"debug_etiqueta_{order_number}.png"
                await page.screenshot(path=str(screenshot))
                log(f"[etiqueta] ❌ Não foi possível clicar no campo de busca — screenshot: {screenshot.name}")
                return {"ok": False, "erro": "Campo de busca bloqueado por popup — screenshot salvo"}

            await search.fill(order_number)
            await page.keyboard.press("Enter")
            await page.wait_for_timeout(1500)

            # Verifica cancelamento APENAS na linha do pedido (evita falso positivo de outras linhas)
            row = page.locator("tr").filter(has_text=order_number).first
            if await row.count() == 0:
                log(f"[etiqueta] ⚠️ {order_number} não está em Para Imprimir — buscando em outras abas...")
                for url_fallback in [
                    "https://app.upseller.com/pt/order/pending",
                    "https://app.upseller.com/pt/order/to-ship",
                    "https://app.upseller.com/pt/order/all-orders",
                ]:
                    await page.goto(url_fallback, wait_until="domcontentloaded", timeout=30_000)
                    await page.wait_for_timeout(2000)
                    await _fechar_popups_upseller(page)
                    # Tenta pesquisar pelo número
                    for sel in ["input[placeholder*='Pedido']", "input[placeholder*='pedido']",
                                "input[placeholder*='Buscar']", ".ant-input-search input", "input[type='search']"]:
                        s = page.locator(sel).first
                        if await s.count() > 0:
                            await s.click()
                            await s.fill(order_number)
                            await page.keyboard.press("Enter")
                            await page.wait_for_timeout(1500)
                            break
                    row = page.locator("tr").filter(has_text=order_number).first
                    if await row.count() > 0:
                        log(f"[etiqueta] ✅ {order_number} encontrado em {url_fallback.split('/')[-1]}")
                        break
                if await row.count() == 0:
                    return {"ok": False, "erro": "Pedido não encontrado em nenhuma aba do UpSeller"}

            # Verifica cancelamento só na linha encontrada
            if await row.count() > 0:
                row_txt = await row.inner_text()
                if "Cancelado" in row_txt or "Estornado" in row_txt:
                    status_txt = "Cancelado" if "Cancelado" in row_txt else "Estornado"
                    log(f"[etiqueta] ⛔ Pedido {order_number} está {status_txt}")
                    try:
                        create_client(*_supa()).table("pedidos").update({"status": "cancelado"}).eq("order_number", order_number).execute()
                    except Exception:
                        pass
                    return {"ok": False, "cancelado": True, "erro": f"Pedido {status_txt.lower()} no UpSeller"}

            tem_impressora = _verificar_impressora()
            log(f"[etiqueta] {'🖨️ Impressora detectada' if tem_impressora else '📄 Sem impressora — gerando PDF'}...")

            # Clica "Imprimir Etiq..." diretamente na linha do pedido
            imprimir = row.locator("a, button, span").filter(has_text="Imprimir Etiq").first
            if await imprimir.count() == 0:
                imprimir = page.locator("a, button").filter(has_text="Imprimir Etiq").first

            # Pedido em "Enviado": etiqueta só disponível em Para Imprimir
            # No fluxo de 6h, pedidos estarão em Para Emitir e serão processados normalmente
            if await imprimir.count() == 0:
                log(f"[etiqueta] ℹ️ Pedido {order_number} não está em Para Imprimir — verifique o estado no UpSeller")
                return {"ok": False, "sem_etiqueta": True, "erro": "Etiqueta não disponível no UpSeller — pedido TikTok/iMile requer impressão manual"}

            if await imprimir.count() == 0:
                screenshot = PASTA_RAIZ / f"debug_etiqueta_{order_number}.png"
                await page.screenshot(path=str(screenshot))
                log(f"[etiqueta] ⚠️ Botão de etiqueta não encontrado — screenshot: {screenshot.name}")
                return {"ok": False, "erro": "Botão de impressão não encontrado"}

            log("[etiqueta] ✅ Clicando botão de etiqueta — aguardando PDF...")
            paginas_antes = set(id(pg) for pg in context.pages)
            await imprimir.click()

            # Polling: aguarda nova aba com o PDF (até 60s)
            popup = None
            for _ in range(60):
                await page.wait_for_timeout(1000)
                novas = [pg for pg in context.pages if id(pg) not in paginas_antes]
                if novas:
                    popup = novas[-1]
                    break

            if not popup:
                log("[etiqueta] ⚠️ PDF não abriu após 60s")
                return {"ok": False, "erro": "PDF não abriu — tente novamente"}

            await popup.wait_for_load_state("networkidle", timeout=30_000)
            await popup.wait_for_timeout(1000)
            await _imprimir_ou_pdf(popup, tem_impressora, order_number, "etiqueta")

            # Confirma impressão no modal → move pedido para Para Retirada
            await page.wait_for_timeout(1500)
            await _fechar_popups_upseller(page)  # fecha avisos antes de procurar o modal
            marcar = page.locator("button", has_text="Marcar como Impresso").first
            if await marcar.count() > 0 and await marcar.is_visible():
                log("[etiqueta] ✅ Confirmando impressão — movendo para Para Retirada...")
                await marcar.click()
                await page.wait_for_timeout(1000)

            log(f"[etiqueta] ✅ Etiqueta emitida: {order_number}")
            return {"ok": True}

        except Exception as e:
            log(f"[etiqueta] ❌ Erro ao emitir {order_number}: {e}")
            return {"ok": False, "erro": str(e)}
        finally:
            await context.close()


# ── Playwright: reimprimir etiqueta/NF-e ─────────────────────────────────────

async def _reimprimir_etiqueta_playwright(config: dict, order_number: str) -> dict:
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        return {"ok": False, "erro": "Playwright não instalado"}

    log(f"[reimprimir] 🖨️ Reimprimindo: {order_number}...")

    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            user_data_dir=str(PASTA_SESSAO),
            headless=True,
            args=["--window-size=1280,900"],
        )
        try:
            page = context.pages[0] if context.pages else await context.new_page()

            await page.goto(
                "https://app.upseller.com/pt/order/all-orders",
                wait_until="domcontentloaded",
                timeout=60_000,
            )
            await page.wait_for_timeout(3000)

            if "/login" in page.url:
                log("[reimprimir] ⚠️ Sessão expirada. Abrindo para login...")
                await context.close()
                context = await p.chromium.launch_persistent_context(
                    user_data_dir=str(PASTA_SESSAO),
                    headless=False,
                    args=["--window-size=1280,900"],
                )
                page = context.pages[0] if context.pages else await context.new_page()
                await page.goto("https://app.upseller.com/login", wait_until="domcontentloaded", timeout=60_000)
                log("[reimprimir] 🔑 Faça login no UpSeller. Aguardando...")
                await page.wait_for_url(lambda url: "/login" not in url, timeout=600_000)
                log("[reimprimir] ✅ Login realizado!")
                await page.wait_for_load_state("networkidle", timeout=30_000)
                await page.wait_for_timeout(3000)
                await page.goto(
                    "https://app.upseller.com/pt/order/all-orders",
                    wait_until="domcontentloaded",
                    timeout=60_000,
                )
                await page.wait_for_timeout(3000)

            await _fechar_popups_upseller(page)

            # Busca o pedido
            log(f"[reimprimir] 🔍 Buscando {order_number}...")
            search = page.locator("input[placeholder*='Pedido']").first
            if await search.count() == 0:
                search = page.locator(".ant-input").first
            await search.click(timeout=15_000)
            await search.fill(order_number)
            await page.keyboard.press("Enter")
            await page.wait_for_timeout(1500)

            # Abre o drawer
            link = page.locator(f"text=#{order_number}").first
            if await link.count() == 0:
                link = page.locator(f"a:has-text('{order_number}')").first
            if await link.count() == 0:
                log(f"[reimprimir] ⚠️ Pedido {order_number} não encontrado")
                return {"ok": False, "erro": f"Pedido {order_number} não encontrado no UpSeller"}

            await link.click()
            await page.wait_for_timeout(3000)

            tem_impressora = _verificar_impressora()
            log(f"[reimprimir] {'🖨️ Impressora detectada' if tem_impressora else '📄 Sem impressora — gerando PDF'}...")

            # Tenta botões diretos no drawer
            for nome_btn in ["Reimprimir Etiqueta", "Reimprimir NF-e", "Reimprimir", "Imprimir Etiqueta", "Imprimir"]:
                btn = page.locator("button", has_text=nome_btn).first
                if await btn.count() > 0 and await btn.is_visible():
                    log(f"[reimprimir] ✅ Clicando '{nome_btn}' — aguardando documento...")
                    paginas_antes = set(id(p) for p in context.pages)
                    await btn.click()

                    popup = None
                    for _ in range(60):
                        await page.wait_for_timeout(1000)
                        novas = [p for p in context.pages if id(p) not in paginas_antes]
                        if novas:
                            popup = novas[-1]
                            break

                    if popup:
                        log("[reimprimir] 📄 Documento aberto — gerando PDF...")
                        await popup.wait_for_load_state("networkidle", timeout=30_000)
                        await popup.wait_for_timeout(1500)
                        await _imprimir_ou_pdf(popup, tem_impressora, order_number, "reimprimir")
                        log(f"[reimprimir] ✅ Reimpresso: {order_number}")
                        return {"ok": True}
                    else:
                        log(f"[reimprimir] ⚠️ Documento não abriu após 60s")
                        return {"ok": False, "erro": "Documento não abriu — tente novamente"}

            # Tenta via dropdown "Mais Ações"
            clicou_mais = await page.evaluate("""() => {
                const els = Array.from(document.querySelectorAll('button, a, [role="button"], .ant-btn'))
                    .filter(e => e.innerText.trim().includes('Mais Ações') && e.offsetParent !== null);
                if (!els.length) return false;
                const target = els.sort((a, b) =>
                    b.getBoundingClientRect().left - a.getBoundingClientRect().left)[0];
                target.click();
                return true;
            }""")

            if clicou_mais:
                log("[reimprimir] 🔽 Abrindo 'Mais Ações'...")
                await page.wait_for_timeout(1500)

                for opcao in ["Reimprimir Etiqueta", "Reimprimir NF-e", "Reimprimir", "Imprimir Etiqueta", "Imprimir", "Etiqueta"]:
                    item = page.locator("li").filter(has_text=opcao).first
                    if await item.count() > 0 and await item.is_visible():
                        log(f"[reimprimir] ✅ Clicando '{opcao}' — aguardando documento...")
                        paginas_antes = set(id(p) for p in context.pages)
                        await item.click()

                        # Polling: aguarda nova aba/popup abrir (até 60s)
                        popup = None
                        for _ in range(60):
                            await page.wait_for_timeout(1000)
                            novas = [p for p in context.pages if id(p) not in paginas_antes]
                            if novas:
                                popup = novas[-1]
                                break

                        if popup:
                            log("[reimprimir] 📄 Documento aberto — gerando PDF...")
                            await popup.wait_for_load_state("networkidle", timeout=30_000)
                            await popup.wait_for_timeout(1500)
                            await _imprimir_ou_pdf(popup, tem_impressora, order_number, "reimprimir")
                            log(f"[reimprimir] ✅ Reimpresso: {order_number}")
                            return {"ok": True}
                        else:
                            log(f"[reimprimir] ⚠️ Documento não abriu após 60s")
                            return {"ok": False, "erro": "Documento não abriu — tente novamente"}

                # Log dos itens encontrados para diagnóstico
                itens = await page.locator("li").all_inner_texts()
                itens_visiveis = [t.strip() for t in itens if t.strip()]
                log(f"[reimprimir] 🔎 Itens li visíveis: {itens_visiveis[:20]}")

            screenshot = PASTA_RAIZ / f"debug_reimprimir_{order_number}.png"
            await page.screenshot(path=str(screenshot))
            log(f"[reimprimir] ⚠️ Botão não encontrado — screenshot: {screenshot.name}")
            return {"ok": False, "erro": "Botão de reimpressão não encontrado — ação manual necessária"}

        except Exception as e:
            log(f"[reimprimir] ❌ Erro ao reimprimir {order_number}: {e}")
            return {"ok": False, "erro": str(e)}
        finally:
            await context.close()


# ── Playwright: emitir NF-e em massa ─────────────────────────────────────────

async def _emitir_nfe_massa_playwright(config: dict) -> bool:
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        log("[nfe] ❌ Playwright não instalado")
        return False

    _limpar_crash_chrome()
    log("[nfe] 📋 Verificando pedidos para emitir NF-e...")

    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            user_data_dir=str(PASTA_SESSAO),
            headless=True,
            args=["--window-size=1280,900"],
        )
        try:
            page = context.pages[0] if context.pages else await context.new_page()

            await page.goto(
                "https://app.upseller.com/pt/order/pending-invoice",
                wait_until="domcontentloaded",
                timeout=60_000,
            )
            await page.wait_for_timeout(2000)

            if "/login" in page.url:
                log("[nfe] ⚠️ Sessão expirada — faça login manualmente")
                return False

            await _fechar_popups_upseller(page)

            log("[nfe] 🔧 NFe engine v1.5.5")
            rodada = 0
            linhas_anterior = None
            sem_progresso = 0
            while True:
                rodada += 1
                await page.reload(wait_until="domcontentloaded")
                await page.wait_for_timeout(1500)
                await _fechar_popups_upseller(page)

                # Garante sub-aba "Para Emitir" — tenta CSS e JS como fallback
                clicou_tab = False
                for texto_tab in ["Para Emitir", "Para emitir"]:
                    tab = page.locator(".ant-tabs-tab").filter(has_text=texto_tab).first
                    if await tab.count() > 0:
                        await tab.click()
                        clicou_tab = True
                    else:
                        clicou_tab = await page.evaluate("""(txt) => {
                            const els = Array.from(document.querySelectorAll('*')).filter(el =>
                                el.innerText && el.innerText.trim().startsWith(txt)
                                && el.offsetParent !== null && !['BODY','HTML','MAIN'].includes(el.tagName)
                            );
                            if (els.length) { els[0].click(); return true; }
                            return false;
                        }""", texto_tab)
                    if clicou_tab:
                        await page.wait_for_timeout(800)
                        log(f"[nfe] 🔖 Sub-aba 'Para Emitir' ativa")
                        break
                if not clicou_tab:
                    log("[nfe] ⚠️ Sub-aba não encontrada — usando visão padrão")

                await _fechar_popups_upseller(page)

                # Critério de parada: usa API para contar total real (corrige bug >100 linhas)
                total_api = await page.evaluate("""async () => {
                    try {
                        const r = await fetch('/api/order/index', {
                            method: 'POST',
                            headers: {'Content-Type': 'application/x-www-form-urlencoded'},
                            body: 'timeType=0&searchType=0&sortName=1&sortValue=1&orderState=invoice_pending&invoiceStatus=to_issue&isVoided=0&pageNum=1&pageSize=1'
                        });
                        const j = await r.json();
                        return (j.data || {}).total || 0;
                    } catch(e) { return -1; }
                }""")

                if total_api == 0:
                    log("[nfe] ✅ Para Emitir vazio (API)")
                    break
                if total_api < 0:
                    # API falhou — fallback para contagem visual
                    emitir_btn = page.locator("button", has_text="Emitir Nota Fiscal").first
                    if await emitir_btn.count() == 0 or not await emitir_btn.is_visible():
                        log("[nfe] ℹ️ Nenhum pedido pendente de NF-e")
                        break
                    total_api = await page.locator("tbody tr").count()

                # Detecta loop travado (3 rodadas sem redução)
                if linhas_anterior is not None and total_api >= linhas_anterior:
                    sem_progresso += 1
                    log(f"[nfe] ⚠️ Sem redução ({linhas_anterior}→{total_api}) — {sem_progresso}/3")
                    if sem_progresso >= 3:
                        log("[nfe] ⚠️ Sem progresso após 3 rodadas — verifique 'Falha na Emissão'")
                        break
                else:
                    sem_progresso = 0
                linhas_anterior = total_api
                linhas = await page.locator("tbody tr").count()

                log(f"[nfe] 📝 {total_api} pedido(s) total ({linhas} na página) — rodada {rodada}")
                # JS click no checkbox de seleção global (evita overlay)
                await page.evaluate("""() => {
                    const cb = document.querySelector('th .ant-checkbox-input');
                    if (cb) cb.click();
                }""")
                await page.wait_for_timeout(800)

                # JS click no "Emitir Nota Fiscal" (evita timeout por overlay)
                clicou_emitir = await page.evaluate("""() => {
                    const btn = Array.from(document.querySelectorAll('button'))
                        .find(b => b.innerText.includes('Emitir Nota Fiscal') && b.offsetParent !== null);
                    if (btn) { btn.click(); return true; }
                    return false;
                }""")
                if not clicou_emitir:
                    log("[nfe] ⚠️ Botão 'Emitir Nota Fiscal' não encontrado via JS — tentando locator")
                    await emitir_btn.click(timeout=10000)
                await page.wait_for_timeout(1500)

                # Detecta formulário de Tributação (abre quando NCM está faltando)
                confirmou = False
                await page.wait_for_timeout(500)

                tem_tributacao = await page.evaluate("""() => {
                    return Array.from(document.querySelectorAll('*'))
                        .some(el => el.innerText && el.innerText.trim() === 'Tributação'
                              && el.offsetParent !== null);
                }""")

                if tem_tributacao:
                    log("[nfe] 📋 Formulário de Tributação — preenchendo NCM...")
                    # Screenshot de diagnóstico do formulário
                    sc_trib = PASTA_RAIZ / f"debug_tributacao_rodada{rodada}.png"
                    try:
                        await page.screenshot(path=str(sc_trib))
                        log(f"[nfe] 📸 Screenshot: {sc_trib.name}")
                    except Exception:
                        pass
                    ncm_map    = config.get("ncm_produtos", {})
                    ncm_padrao = config.get("ncm_padrao", "6109.10.00")

                    # Aguarda form renderizar completamente antes de extrair linhas
                    await page.wait_for_timeout(1500)

                    # Extrai nomes dos produtos sem usar locators (evita stale)
                    nomes_prod = await page.evaluate("""() => {
                        return Array.from(document.querySelectorAll('tbody tr'))
                            .filter(tr => tr.querySelector('[class*="info_item_ncm"]'))
                            .map(tr => {
                                const td = tr.querySelector('td');
                                const all = Array.from((td || tr).querySelectorAll('*'))
                                    .filter(e => !e.children.length
                                             && e.innerText
                                             && e.innerText.trim().length > 1);
                                return all.length ? all[0].innerText.trim() : '';
                            });
                    }""")
                    log(f"[nfe] 🔎 {len(nomes_prod)} linha(s) com NCM faltando: {nomes_prod}")

                    for idx, nome_prod in enumerate(nomes_prod):
                        ncm_val = ncm_map.get(nome_prod, ncm_padrao)
                        if not ncm_val:
                            log(f"[nfe] ⚠️ NCM não resolvido para '{nome_prod}' — pulando linha")
                            continue
                        log(f"[nfe] 🏷️ '{nome_prod}' → NCM {ncm_val}")

                        # Posição do campo NCM via page.evaluate com índice (sem locator stale)
                        pos = await page.evaluate("""(idx) => {
                            const rows = Array.from(document.querySelectorAll('tbody tr'))
                                .filter(tr => tr.querySelector('[class*="info_item_ncm"]'));
                            if (idx >= rows.length) return null;
                            const sels = Array.from(rows[idx].querySelectorAll('.ant-select'))
                                .filter(s => s.offsetParent !== null)
                                .sort((a, b) => a.getBoundingClientRect().x - b.getBoundingClientRect().x);
                            for (const s of sels) {
                                const v = s.querySelector('.ant-select-selection-selected-value');
                                if (!v || !v.innerText.trim()) {
                                    const r = s.getBoundingClientRect();
                                    return {x: Math.round(r.x + r.width/2), y: Math.round(r.y + r.height/2)};
                                }
                            }
                            return null;
                        }""", idx)

                        if not pos:
                            log(f"[nfe] ⚠️ Campo NCM não localizado para '{nome_prod}'")
                            continue
                        await page.mouse.click(pos['x'], pos['y'])
                        await page.wait_for_timeout(400)
                        await page.keyboard.type(ncm_val)
                        await page.wait_for_timeout(1000)  # aguarda dropdown aparecer
                        clicou_dd = await page.evaluate("""() => {
                            const li = document.querySelector(
                                '.ant-select-dropdown:not(.ant-select-dropdown-hidden) li'
                            );
                            if (li) { li.click(); return true; }
                            return false;
                        }""")
                        if clicou_dd:
                            log(f"[nfe] ✅ Dropdown NCM selecionado para '{nome_prod}'")
                        else:
                            await page.keyboard.press("Enter")
                            log(f"[nfe] ⚠️ Dropdown não apareceu — pressionou Enter para '{nome_prod}'")
                        await page.wait_for_timeout(500)

                    # JS click no Emitir — zero locator, zero timeout
                    confirmou = await page.evaluate("""() => {
                        const btn = Array.from(document.querySelectorAll('button'))
                            .find(b => b.innerText.trim() === 'Emitir');
                        if (btn) { btn.click(); return true; }
                        return false;
                    }""")
                    if confirmou:
                        log("[nfe] ✅ Tributação submetida!")
                    else:
                        log("[nfe] ⚠️ Botão Emitir não encontrado na Tributação")
                else:
                    # Caso 1: dialog "Submetido" — UpSeller já iniciou a emissão
                    submetido = await page.evaluate("""() => {
                        return Array.from(document.querySelectorAll('*'))
                            .some(el => el.innerText
                                  && el.innerText.includes('Submetido a Estar Emitindo')
                                  && el.offsetParent !== null);
                    }""")
                    if submetido:
                        log("[nfe] ✅ Emissão submetida pelo UpSeller")
                        await page.evaluate("""() => {
                            const btn = Array.from(document.querySelectorAll('button'))
                                .find(b => b.innerText.trim() === 'Fechar' && b.offsetParent !== null);
                            if (btn) btn.click();
                        }""")
                        await page.wait_for_timeout(500)
                        confirmou = True
                    else:
                        # Caso 2: modal de confirmação com botão Emitir/Confirmar
                        clicou_conf = await page.evaluate("""() => {
                            const nomes = ['Confirmar', 'OK', 'Sim', 'Emitir', 'Confirmar Emissão'];
                            for (const nome of nomes) {
                                const btn = Array.from(document.querySelectorAll('button'))
                                    .find(b => b.innerText.trim() === nome && b.offsetParent !== null);
                                if (btn) { btn.click(); return nome; }
                            }
                            return null;
                        }""")
                        if clicou_conf:
                            log(f"[nfe] ✅ Confirmação '{clicou_conf}' via JS")
                            confirmou = True

                if not confirmou:
                    log("[nfe] ℹ️ Sem confirmação necessária — emissão já em andamento")

                espera_max = min(max(linhas * 2, 30), 120)
                log(f"[nfe] ⏳ Aguardando lote {rodada} (até {espera_max}s)...")
                for _i in range(espera_max // 5):
                    await page.wait_for_timeout(5_000)
                    t_novo = await page.evaluate("""async () => {
                        try {
                            const r = await fetch('/api/order/index', {
                                method:'POST', headers:{'Content-Type':'application/x-www-form-urlencoded'},
                                body:'timeType=0&searchType=0&sortName=1&sortValue=1&orderState=invoice_pending&invoiceStatus=to_issue&isVoided=0&pageNum=1&pageSize=1'
                            });
                            const j = await r.json();
                            return (j.data||{}).total ?? -1;
                        } catch(e) { return -1; }
                    }""")
                    if t_novo >= 0 and t_novo < total_api:
                        log(f"[nfe] ⚡ Lote concluído em {(_i+1)*5}s ({total_api}→{t_novo})")
                        break
                log(f"[nfe] ✅ Lote {rodada} iniciado")

            return True

        except Exception as e:
            log(f"[nfe] ❌ Erro: {e}")
            return False
        finally:
            await context.close()


# ── Playwright: programar envio em massa ──────────────────────────────────────

async def _programar_envio_playwright(config: dict) -> bool:
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        log("[envio] ❌ Playwright não instalado")
        return False

    _limpar_crash_chrome()
    log("[envio] 🚚 Verificando pedidos para programar envio...")

    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            user_data_dir=str(PASTA_SESSAO),
            headless=True,
            args=["--window-size=1280,900"],
        )
        try:
            page = context.pages[0] if context.pages else await context.new_page()

            await page.goto(
                "https://app.upseller.com/pt/order/to-ship",
                wait_until="domcontentloaded",
                timeout=60_000,
            )
            await page.wait_for_timeout(3000)

            if "/login" in page.url:
                log("[envio] ⚠️ Sessão expirada — faça login manualmente")
                return False

            await _fechar_popups_upseller(page)

            PLATAFORMAS_UI = ["Mercado Libre", "Shopee", "Shein", "TikTok Shop"]

            async def _aplicar_filtro_plat(plat_nome: str) -> bool:
                try:
                    plat_sel = page.locator('.ant-select').filter(has_text='Plataformas')
                    if await plat_sel.count() == 0:
                        return False
                    await plat_sel.first.click()
                    await page.wait_for_timeout(800)
                    opcao = page.locator('.ant-select-dropdown-menu-item').filter(has_text=plat_nome)
                    if await opcao.count() == 0:
                        await page.keyboard.press('Escape')
                        return False
                    await opcao.first.click()
                    await page.wait_for_timeout(2000)
                    return True
                except Exception:
                    return False

            async def _get_total_pag() -> int:
                val = await page.evaluate("""() => {
                    const el = document.querySelector('.ant-pagination-total-text');
                    if (el) { const m = el.innerText.match(/\\d+/); return m ? parseInt(m[0]) : null; }
                    return null;
                }""")
                if val is not None:
                    return val
                return await page.locator("tbody tr").count()

            for plat_nome in PLATAFORMAS_UI:
                await page.reload(wait_until="domcontentloaded")
                await page.wait_for_timeout(2000)
                await _fechar_popups_upseller(page)

                filtrado = await _aplicar_filtro_plat(plat_nome)
                if not filtrado:
                    log(f"[envio] ⚠️ Filtro '{plat_nome}' não disponível — pulando")
                    continue

                linhas = await page.locator("tbody tr").count()
                if linhas == 0:
                    log(f"[envio] ✅ {plat_nome}: sem pedidos pendentes")
                    continue

                programar_btn_chk = page.locator("button", has_text="Programar Envio").first
                if await programar_btn_chk.count() == 0 or not await programar_btn_chk.is_visible():
                    log(f"[envio] ✅ {plat_nome}: botão 'Programar Envio' não disponível")
                    continue

                total_inicial = await _get_total_pag()
                log(f"[envio] 📊 {plat_nome}: {total_inicial} pedido(s) — iniciando agendamento")

                rodada = 0
                total_anterior = None
                sem_progresso = 0

                while True:
                    rodada += 1

                    if rodada > 1:
                        await page.reload(wait_until="domcontentloaded")
                        await page.wait_for_timeout(2000)
                        await _fechar_popups_upseller(page)
                        await _aplicar_filtro_plat(plat_nome)
                        await page.wait_for_timeout(1000)

                    linhas = await page.locator("tbody tr").count()
                    if linhas == 0:
                        log(f"[envio] ✅ {plat_nome}: todos agendados!")
                        break

                    programar_btn = page.locator("button", has_text="Programar Envio").first
                    if await programar_btn.count() == 0 or not await programar_btn.is_visible():
                        log(f"[envio] ✅ {plat_nome}: fila limpa!")
                        break

                    total = await _get_total_pag()
                    log(f"[envio] 📊 {plat_nome}: {total} total ({linhas} na página) — rodada {rodada}")

                    if total_anterior is not None and total >= total_anterior:
                        sem_progresso += 1
                        log(f"[envio] ⚠️ {plat_nome}: sem redução ({total_anterior}→{total}) — {sem_progresso}/3")
                        if sem_progresso >= 3:
                            log(f"[envio] ⚠️ {plat_nome}: sem progresso após 3 rodadas — próxima plataforma")
                            break
                    else:
                        sem_progresso = 0
                    total_anterior = total

                    log(f"[envio] 📦 {plat_nome}: rodada {rodada} — selecionando todos...")

                    # Clica checkbox select-all via JS (mesmo método do NF-e que funciona)
                    cb_selecionado = await page.evaluate("""() => {
                        const sels = [
                            'th .ant-checkbox-input',
                            'thead .ant-checkbox-input',
                            '.ant-table-thead .ant-checkbox-input',
                            'th input[type="checkbox"]',
                            'thead input[type="checkbox"]',
                        ];
                        for (const s of sels) {
                            const cb = document.querySelector(s);
                            if (cb) { cb.click(); return s; }
                        }
                        return null;
                    }""")
                    if cb_selecionado:
                        log(f"[envio] ☑️ Checkbox selecionado via JS '{cb_selecionado}'")
                    else:
                        log("[envio] ⚠️ Checkbox select-all não encontrado — tentando locator")
                        for cb_sel in ["th .ant-checkbox-wrapper", "thead .ant-checkbox-wrapper"]:
                            cb_loc = page.locator(cb_sel).first
                            if await cb_loc.count() > 0:
                                await cb_loc.click(force=True)
                                log(f"[envio] ☑️ Checkbox selecionado via locator '{cb_sel}'")
                                break
                    await page.wait_for_timeout(1500)

                    qtd_sel = await page.evaluate("""() => {
                        const checks = [
                            'tbody .ant-checkbox-checked',
                            'tr.ant-table-row-selected',
                            'tbody input[type="checkbox"]:checked',
                            'tbody .ant-checkbox-input:checked',
                        ];
                        for (const s of checks) {
                            const n = document.querySelectorAll(s).length;
                            if (n > 0) return n;
                        }
                        return 0;
                    }""")
                    log(f"[envio] ✔️ {qtd_sel} pedidos selecionados")
                    if qtd_sel == 0:
                        log("[envio] ⚠️ Nenhum pedido selecionado — pulando rodada")
                        sem_progresso += 1
                        if sem_progresso >= 3:
                            break
                        continue

                    # Clica "Programar Envio" via JS direto (evita interceptação por overlay
                    # e funciona mesmo quando o botão fica abaixo do viewport em modais longos)
                    clicou_toolbar = await page.evaluate("""() => {
                        // Procura em elementos genéricos (SPAN, A, BUTTON) — o action bar
                        // usa SPAN dentro de BUTTON.ant-btn-link, não um <button> comum
                        const todos = Array.from(document.querySelectorAll('*'))
                            .filter(el => el.offsetParent !== null &&
                                          el.innerText && el.innerText.trim() === 'Programar Envio' &&
                                          !el.children.length);
                        // Prefere o elemento no action bar (y < 400, x > 100)
                        const action = todos.find(el => {
                            const r = el.getBoundingClientRect();
                            return r.y < 400 && r.x > 100;
                        });
                        if (action) { action.click(); return action.tagName + '/' + (action.className||'action'); }
                        // Fallback: qualquer button com texto "Programar Envio"
                        const btn = Array.from(document.querySelectorAll('button'))
                            .find(b => b.innerText.trim().includes('Programar Envio') && b.offsetParent !== null);
                        if (btn) { btn.click(); return 'BUTTON/btn'; }
                        return null;
                    }""")
                    if clicou_toolbar:
                        log(f"[envio] 🖱️ Clicou 'Programar Envio' via JS ({clicou_toolbar})")
                    else:
                        await programar_btn.click(force=True)
                        log("[envio] 🖱️ Clicou 'Programar Envio' (locator fallback)")
                    await page.wait_for_timeout(3000)

                    # Confirma modal via JS direto (funciona mesmo com botão abaixo do viewport)
                    clicou_modal = await page.evaluate("""() => {
                        const textos = ['Programar Envio', 'Programar', 'Confirmar', 'Confirm', 'OK', 'Ok', 'Sim', 'Yes'];
                        const seletores = [
                            '.ant-modal-content button',
                            '.ant-popover-inner button',
                            '.ant-popconfirm button',
                            '.ant-dropdown-menu-item',
                        ];
                        for (const sel of seletores) {
                            const btns = Array.from(document.querySelectorAll(sel));
                            for (const txt of textos) {
                                const btn = btns.find(b => b.innerText.trim().includes(txt) && b.offsetParent !== null);
                                if (btn) { btn.click(); return btn.innerText.trim().substring(0, 30); }
                            }
                        }
                        return null;
                    }""")
                    if clicou_modal:
                        log(f"[envio] ✅ Confirmação clicada '{clicou_modal}' via JS")
                        await page.wait_for_timeout(2000)
                    else:
                        log(f"[envio] ℹ️ {plat_nome}: sem modal de confirmação — processando direto")

                    # Detecta "Programado com sucesso" (Shopee Xpress agenda coleta de forma
                    # assíncrona — pedidos só saem de Para Enviar quando o rastreio é gerado,
                    # então não há redução imediata de contagem; o sucesso é detectado via modal)
                    await page.wait_for_timeout(1500)
                    coleta_async = await page.evaluate("""() => {
                        const m = document.querySelector('.ant-modal-content');
                        if (!m || m.offsetParent === null) return null;
                        const txt = m.innerText || '';
                        if (txt.includes('Programado com sucesso') || txt.includes('sucesso')) {
                            const fechar = Array.from(m.querySelectorAll('button'))
                                .find(b => b.innerText.trim() === 'Fechar' && b.offsetParent !== null);
                            if (fechar) { fechar.click(); return 'fechado'; }
                            return 'sucesso';
                        }
                        return null;
                    }""")

                    if coleta_async:
                        log(f"[envio] ✅ {plat_nome}: coleta agendada (Shopee Xpress async — pedidos vão para Para Imprimir após rastreio)")
                        progresso = True
                        await page.wait_for_timeout(1000)
                        # Verifica quantos ainda faltam via aba "Para Programar"
                        faltam = await page.evaluate("""() => {
                            const tabs = Array.from(document.querySelectorAll(
                                '.ant-tabs-tab, [class*="tab"], [role="tab"]'
                            ));
                            for (const t of tabs) {
                                if (t.innerText.includes('Para Programar')) {
                                    const m = t.innerText.match(/\\d+/);
                                    return m ? parseInt(m[0]) : null;
                                }
                            }
                            return null;
                        }""")
                        if faltam is not None:
                            log(f"[envio] 📊 {plat_nome}: {faltam} pedido(s) ainda em 'Para Programar'")
                            if faltam == 0:
                                log(f"[envio] ✅ {plat_nome}: todos agendados!")
                                break
                        log(f"[envio] ✅ {plat_nome}: lote {rodada} processado")
                        continue  # próxima rodada sem esperar 90s

                    # Aguarda redução de contagem (TikTok, Shein, ML — processamento síncrono)
                    total_antes = total
                    log(f"[envio] ⏳ {plat_nome}: aguardando processamento... (total: {total_antes})")
                    progresso = False
                    for i in range(18):
                        await page.wait_for_timeout(5_000)
                        if i == 10:
                            await page.reload(wait_until="domcontentloaded")
                            await page.wait_for_timeout(1500)
                            await _fechar_popups_upseller(page)
                            await _aplicar_filtro_plat(plat_nome)
                            await page.wait_for_timeout(1000)
                        restantes = await page.locator("tbody tr").count()
                        total_agora = await page.evaluate("""() => {
                            const el = document.querySelector('.ant-pagination-total-text');
                            if (el) { const m = el.innerText.match(/\\d+/); return m ? parseInt(m[0]) : null; }
                            return null;
                        }""")
                        if total_agora is not None and total_agora < total_antes:
                            log(f"[envio] ⏳ {plat_nome}: {total_agora} restando (era {total_antes})")
                            progresso = True
                            if total_agora == 0:
                                break
                        elif restantes < linhas:
                            log(f"[envio] ⏳ {plat_nome}: {restantes} rows restando")
                            progresso = True
                            if restantes == 0:
                                break

                    if not progresso:
                        log(f"[envio] ℹ️ {plat_nome}: sem redução após 90s — próxima plataforma")

                    log(f"[envio] ✅ {plat_nome}: lote {rodada} processado")

            log("[envio] ✅ Todas as plataformas processadas!")
            return True

        except Exception as e:
            log(f"[envio] ❌ Erro: {e}")
            return False
        finally:
            await context.close()


# ── Playwright: login manual ──────────────────────────────────────────────────

async def _login_upseller_playwright(config: dict):
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        log("❌ Playwright não instalado. Rode: pip install playwright && playwright install chromium")
        return

    email = config.get("upseller_email", "")
    senha = config.get("upseller_senha", "")

    log("🌐 Abrindo browser para autenticação no UpSeller...")
    PASTA_SESSAO.mkdir(exist_ok=True)

    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            user_data_dir=str(PASTA_SESSAO),
            headless=False,
            args=["--window-size=1280,800"],
        )
        try:
            page = context.pages[0] if context.pages else await context.new_page()
            await page.goto(
                "https://app.upseller.com/login",
                wait_until="domcontentloaded",
                timeout=60_000,
            )
            await page.wait_for_timeout(2000)

            log(f"🔑 Digite seu e-mail e senha no UpSeller: {email}")
            log("Aguardando login (até 10 minutos)...")

            await page.wait_for_url(lambda url: "/login" not in url, timeout=600_000)
            log("✅ Login realizado! Aguardando sessão ser salva...")
            # Aguarda a página carregar completamente e o UpSeller salvar os tokens
            await page.wait_for_load_state("networkidle", timeout=30_000)
            await page.wait_for_timeout(4000)
            log("✅ Sessão salva! Pode fechar o browser.")
        except Exception as e:
            log(f"❌ Erro no login: {e}")
        finally:
            await context.close()


# ── Robô de extração ──────────────────────────────────────────────────────────

def extrair(config: dict, data_alvo: date):
    import openpyxl

    token        = config["token"]
    data_arquivo = data_alvo.strftime("%Y-%m-%d")
    data_display = data_alvo.strftime("%d/%m/%Y")

    PASTA_DADOS.mkdir(exist_ok=True)

    xlsx = PASTA_RAIZ / "arquivos" / f"pedidos_{data_arquivo}.xlsx"
    if not xlsx.exists():
        log(f"❌ Arquivo não encontrado: pedidos_{data_arquivo}.xlsx")
        return

    log(f"Lendo arquivo: {xlsx.name}")
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.active
    headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
    log(f"📋 Todas as colunas do Excel: {[h for h in headers if h]}")

    def achar(nome):
        for i, h in enumerate(headers):   # exact match primeiro
            if h == nome:
                return i + 1
        for i, h in enumerate(headers):   # substring como fallback
            if nome in h:
                return i + 1
        raise ValueError(f"Coluna '{nome}' não encontrada")

    try:
        col_pedido     = achar("Nº de Pedido")
        col_sku        = achar("SKU")
        col_imagem     = achar("Link da Imagem")
        col_nome       = achar("Nome do Anúncio")
        col_qtd        = achar("Qtd. do Produto")
        col_plataforma = achar("Plataformas")
        col_valor      = achar("Valor do Pedido")
        log(f"Colunas: pedido={col_pedido} sku={col_sku} img={col_imagem} nome={col_nome} qtd={col_qtd} plataforma={col_plataforma} valor={col_valor}")
    except ValueError as e:
        log(f"❌ {e}")
        log(f"Colunas disponíveis: {[h for h in headers if h][:10]}")
        return

    # Coluna opcional — URL da etiqueta já gerada pelo UpSeller
    try:
        col_etiqueta = achar("Etiqueta")
    except ValueError:
        col_etiqueta = None

    pedidos_dict  = {}
    itens_list    = []
    ultimo_order  = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        order = str(row[col_pedido - 1] or "").strip()
        if order:
            ultimo_order = order
        elif ultimo_order:
            order = ultimo_order   # linha de continuação do mesmo pedido
        else:
            continue

        sku        = str(row[col_sku        - 1] or "").strip()
        nome       = str(row[col_nome       - 1] or "").strip()
        imagem     = str(row[col_imagem     - 1] or "").strip()
        qtd        = int(float(str(row[col_qtd - 1] or 1) or 1)) if col_qtd else 1
        plataforma  = str(row[col_plataforma  - 1] or "").strip()
        label_url   = str(row[col_etiqueta   - 1] or "").strip() if col_etiqueta else ""
        try:
            valor_raw = row[col_valor - 1]
            valor = float(valor_raw) if valor_raw not in (None, "") else None
        except (TypeError, ValueError):
            valor = None

        if not sku and not nome and not imagem:
            continue   # linha completamente vazia, ignora

        itens_list.append({
            "order_number": order,
            "sku":          sku,
            "product_name": nome,
            "image_url":    imagem,
            "quantidade":   qtd,
            "cliente":      token,
            "data":         data_arquivo,
        })

        if order not in pedidos_dict:
            pedidos_dict[order] = {
                "order_number": order,
                "sku":          sku,
                "product_name": nome,
                "image_url":    imagem,
                "quantidade":   qtd,
                "plataforma":   plataforma,
                "valor":        valor,
                "label_url":    label_url or None,
            }
        else:
            pedidos_dict[order]["quantidade"] += qtd
            if label_url and not pedidos_dict[order].get("label_url"):
                pedidos_dict[order]["label_url"] = label_url

    pedidos = list(pedidos_dict.values())
    wb.close()

    multi = sum(1 for p in pedidos if p.get("quantidade", 1) > 1)
    log(f"Pedidos encontrados: {len(pedidos)} ({multi} com 2+ itens, {len(itens_list)} itens no total)")

    # ── 1. Baixa etiquetas do Excel ANTES de salvar o JSON ───────────────────
    # label_url no Excel é a URL externa real — baixa para a pasta e MANTÉM a URL original
    import urllib.request as _urllib
    com_url = [(p["order_number"], p["label_url"]) for p in pedidos if p.get("label_url")]
    if com_url:
        log(f"📥 Baixando {len(com_url)} etiqueta(s) do Excel...")
        baixados = 0
        for order_num, url in com_url:
            if "127.0.0.1:5001" in url or "localhost:5001" in url:
                continue  # URL já aponta pro robô (registro antigo) — pula
            pdf_path = PASTA_ETIQUETAS / f"etiqueta_{order_num}.pdf"
            if not pdf_path.exists():
                try:
                    _urllib.urlretrieve(url, str(pdf_path))
                    _corr = _corrigir_mediabox(pdf_path)
                    if _corr != pdf_path and _corr.exists():
                        _corr.replace(pdf_path)
                    baixados += 1
                except Exception:
                    pass
        if baixados:
            log(f"✅ {baixados} etiqueta(s) salvas em cache")

    # ── 2. JSON para o PCP: label_url aponta para o robô local ───────────────
    # Constrói cópia dos dados para o JSON sem modificar a lista `pedidos`
    # (que vai para o Supabase com a URL original — usada como fallback em /etiqueta/)
    import copy as _copy
    pedidos_json = _copy.deepcopy(pedidos)
    for p in pedidos_json:
        on = p["order_number"]
        if p.get("label_url") and "127.0.0.1:5001" not in p["label_url"]:
            if (PASTA_ETIQUETAS / f"etiqueta_{on}.pdf").exists():
                p["label_url"] = f"http://127.0.0.1:5001/etiqueta/{on}"

    # ── 3. Salva JSON (PCP lê daqui) ─────────────────────────────────────────
    arq = PASTA_DADOS / f"lista_{data_arquivo}_{token[:8]}.json"
    arq.write_text(json.dumps(pedidos_json, ensure_ascii=False, indent=2), encoding="utf-8")
    log(f"JSON salvo: {arq.name}")

    # ── 4. Envia para Supabase com URL ORIGINAL (não a URL do robô) ──────────
    log("Enviando para Supabase...")
    try:
        supa   = create_client(*_supa())
        linhas = [{
            "order_number": p["order_number"],
            "sku":          p["sku"],
            "product_name": p["product_name"],
            "image_url":    p["image_url"],
            "data":         data_arquivo,
            "cliente":      token,
            "quantidade":   p.get("quantidade", 1),
            "plataforma":   p.get("plataforma"),
            "valor":        p.get("valor"),
            "label_url":    p.get("label_url"),  # URL original — /etiqueta/ usa como fallback
        } for p in pedidos]

        if linhas:
            # Separa novos (inserir com data) de existentes (atualizar sem sobrescrever data)
            # Faz em lotes de 50 para evitar limite de URL do Supabase
            ons = [l["order_number"] for l in linhas]
            existentes_set: set = set()
            for i in range(0, len(ons), 50):
                ex_res = supa.table("pedidos").select("order_number").in_("order_number", ons[i:i+50]).execute()
                existentes_set.update(r["order_number"] for r in (ex_res.data or []))
            novos    = [l for l in linhas if l["order_number"] not in existentes_set]
            updates  = [{k: v for k, v in l.items() if k != "data"}
                        for l in linhas if l["order_number"] in existentes_set]

            colunas_opcionais = {"plataforma", "valor", "label_url", "numero_plataforma", "nome_cliente", "status_upseller"}

            def _upsert_com_fallback(lote_orig, label):
                excluir: set = set()
                tentativa = lote_orig
                while True:
                    try:
                        supa.table("pedidos").upsert(tentativa, on_conflict="order_number").execute()
                        log(f"✅ {len(tentativa)} pedidos {label} ao Supabase!")
                        _ultima_rodada["inseridos"] = _ultima_rodada.get("inseridos", 0) + len(tentativa)
                        return
                    except Exception as e_col:
                        col_faltando = next((c for c in colunas_opcionais - excluir if c in str(e_col)), None)
                        if col_faltando:
                            excluir.add(col_faltando)
                            tentativa = [{k: v for k, v in l.items() if k not in excluir} for l in lote_orig]
                            log(f"⚠️ Coluna '{col_faltando}' ausente — tentando sem ela...")
                        else:
                            raise

            def _update_existentes(lote):
                for linha in lote:
                    on = linha.get("order_number")
                    if not on:
                        continue
                    campos = {k: v for k, v in linha.items() if k != "order_number"}
                    try:
                        supa.table("pedidos").update(campos).eq("order_number", on).execute()
                    except Exception as e_upd:
                        log(f"⚠️ Erro ao atualizar {on}: {e_upd}")
                log(f"✅ {len(lote)} pedidos atualizados (data preservada)")
                _ultima_rodada["atualizados"] = _ultima_rodada.get("atualizados", 0) + len(lote)

            if novos:
                _upsert_com_fallback(novos, "inseridos")
            if updates:
                _update_existentes(updates)
        else:
            log("Nenhum pedido para enviar")

    except Exception as e:
        log(f"Erro Supabase (pedidos): {e}")

    try:
        if itens_list:
            order_numbers = list(pedidos_dict.keys())
            supa.table("pedido_itens").delete().in_("order_number", order_numbers).execute()
            batch = 200
            for i in range(0, len(itens_list), batch):
                supa.table("pedido_itens").insert(itens_list[i:i+batch]).execute()
            log(f"✅ {len(itens_list)} itens enviados a pedido_itens!")
    except Exception as e:
        log(f"Erro Supabase (pedido_itens): {e}")

    atualizar_ultima_execucao(token)
    _ultima_rodada.update({"em": datetime.now().strftime("%d/%m %H:%M"), "pedidos": len(pedidos)})
    log(f"✅ Concluído! {len(pedidos)} pedidos | {data_display}")

    try:
        xlsx.unlink()
        log(f"🗑 Arquivo Excel removido.")
    except Exception:
        pass


# ── Importação via API interna do UpSeller ────────────────────────────────────

async def _importar_via_api(config: dict, data_arquivo: str) -> bool:
    """
    Importa pedidos diretamente da API interna do UpSeller (sem Excel).
    Busca Para Imprimir (in_process) e Para Emitir (invoice_pending).
    """
    import copy as _copy, urllib.request as _urllib
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        log("[api] ❌ Playwright não instalado")
        return False

    token = config["token"]
    PASTA_DADOS.mkdir(exist_ok=True)
    _limpar_crash_chrome()

    todos_raw = []

    async with async_playwright() as p:
        context = await p.chromium.launch_persistent_context(
            user_data_dir=str(PASTA_SESSAO),
            headless=True,
            args=["--window-size=1280,900"],
        )
        try:
            page = context.pages[0] if context.pages else await context.new_page()

            await page.goto(
                "https://app.upseller.com/pt/order/in-process",
                wait_until="domcontentloaded",
                timeout=60_000,
            )
            await page.wait_for_timeout(2000)

            if "/login" in page.url:
                log("[api] ⚠️ Sessão expirada — faça login manualmente")
                return False

            # Para Reservar usa orderState=allocate com allocateStatus=pending_review
            # (descoberto via interceptação da página /pt/order/to-allocate)
            SECOES = [
                ("allocate",        "allocateStatus=pending_review",                "Para Reservar", False),
                ("in_process",      "labelStatus=success&warehouseType=0",          "Para Imprimir", False),
                ("invoice_pending", "invoiceStatus=to_issue&isVoided=0",            "Para Emitir",   False),
                ("to_ship",         "",                                              "Para Enviar",   False),
                ("to_pickup",       "",                                              "Para Retirada", False),
            ]

            for state, extra, label_secao, filtrar_por_data in SECOES:
                page_num = 1
                total_secao = None
                contado = 0
                while True:
                    body_str = (
                        f"timeType=0&searchType=0&searchValue=&sortName=1&sortValue=1"
                        f"&orderState={state}&isVoided=0"
                        + (f"&{extra}" if extra else "")
                        + f"&pageNum={page_num}&pageSize=50"
                    )
                    result = await page.evaluate(f"""async () => {{
                        try {{
                            const r = await fetch('/api/order/index', {{
                                method: 'POST',
                                headers: {{'Content-Type': 'application/x-www-form-urlencoded'}},
                                body: '{body_str}'
                            }});
                            return await r.json();
                        }} catch(e) {{ return null; }}
                    }}""")

                    data = (result or {}).get("data") or {}
                    lista = data.get("list") or []
                    total = data.get("total") or 0

                    if total_secao is None:
                        total_secao = total

                    # Para Enviar: só pedidos com NF-e emitida hoje
                    if filtrar_por_data:
                        lista = [
                            o for o in lista
                            if (o.get("invoiceTime") or "").startswith(data_arquivo)
                        ]

                    if page_num == 1 and not filtrar_por_data:
                        if total > 0:
                            log(f"[api] {label_secao}: {total} pedido(s)")
                        else:
                            log(f"[api] {label_secao}: 0 — raw={str(result)[:200]}")
                        _ultima_rodada[label_secao.lower().replace(" ", "_")] = total

                    for order in lista:
                        order["_state"] = state

                    contado += len(lista)
                    todos_raw.extend(lista)

                    if page_num * 50 >= total_secao or not data.get("list"):
                        break
                    page_num += 1

                if filtrar_por_data:
                    log(f"[api] {label_secao} (hoje): {contado} pedido(s)")

            # Detecta pedidos cancelados (voided) no UpSeller e atualiza Supabase
            try:
                nums_ativos = {(o.get("orderNumber") or "").strip() for o in todos_raw}
                pg_v = 1
                cancelados_api: list = []
                while True:
                    r_v = await page.evaluate(f"""async () => {{
                        const r = await fetch('/api/order/index', {{
                            method:'POST', headers:{{'Content-Type':'application/x-www-form-urlencoded'}},
                            body:'timeType=0&isVoided=1&searchType=0&sortName=1&sortValue=1&pageNum={pg_v}&pageSize=50'
                        }});
                        return await r.json();
                    }}""")
                    lista_v = (r_v.get("data") or {}).get("list") or []
                    total_v = (r_v.get("data") or {}).get("total") or 0
                    cancelados_api.extend((o.get("orderNumber") or "").strip() for o in lista_v)
                    if pg_v * 50 >= total_v or not lista_v:
                        break
                    pg_v += 1
                cancelados_api_set = {n for n in cancelados_api if n}
                if cancelados_api_set:
                    supa_c = create_client(*_supa())
                    for i in range(0, len(cancelados_api), 50):
                        lote_c = list(cancelados_api_set)[i:i+50]
                        supa_c.table("pedidos").update({"status": "cancelado", "status_upseller": "Cancelado"}).in_("order_number", lote_c).eq("status", "ativo").execute()
                    log(f"[api] ⛔ {len(cancelados_api_set)} pedido(s) voided verificado(s) no UpSeller")
                    _ultima_rodada["cancelados"] = _ultima_rodada.get("cancelados", 0) + len(cancelados_api_set)
            except Exception as e_v:
                log(f"[api] ⚠️ Verificação de cancelados falhou: {e_v}")

            # Cruza sem-etiqueta do Supabase com filas ativas — detecta pedidos que saíram de todas as filas
            try:
                _ativos_set = {(o.get("orderNumber") or "").strip() for o in todos_raw if o.get("orderNumber")}
                supa_d = create_client(*_supa())
                res_d = supa_d.table("pedidos").select("order_number") \
                    .eq("status", "ativo").is_("label_url", "null").execute()
                nums_sem = {(r.get("order_number") or "").strip() for r in (res_d.data or []) if r.get("order_number")}
                desap = nums_sem - _ativos_set
                if desap:
                    log(f"[api] ⚠️ {len(desap)} pedido(s) sem etiqueta fora de todas as filas (cancelados pela plataforma?): "
                        + ", ".join(sorted(desap)[:15]) + ("..." if len(desap) > 15 else ""))
                    _ultima_rodada["sem_fila"] = len(desap)
                elif nums_sem:
                    log(f"[api] ✅ {len(nums_sem)} sem etiqueta — todos nas filas ativas")
            except Exception as e_d:
                log(f"[api] ⚠️ Cruzamento sem-etiqueta falhou: {e_d}")

        except Exception as e:
            log(f"[api] ❌ Erro ao consultar API: {e}")
            return False
        finally:
            await context.close()

    if not todos_raw:
        log("[api] ℹ️ Nenhum pedido pendente no momento")
        # Salva JSON vazio para PCP não ficar com dados antigos
        arq = PASTA_DADOS / f"lista_{data_arquivo}_{token[:8]}.json"
        arq.write_text("[]", encoding="utf-8")
        return True

    # Mapeamento plataforma API → nome exibição
    PLAT = {
        "shopee":  "Shopee",
        "shein":   "Shein",
        "mercado": "Mercado Livre",
        "tiktok":  "TikTok",
        "kwai":    "Kwai",
    }

    pedidos_dict: dict = {}
    itens_list:   list = []

    for order in todos_raw:
        order_num = (order.get("orderNumber") or "").strip()
        if not order_num:
            continue

        plataforma = PLAT.get((order.get("platform") or "").lower(), order.get("platform") or "")
        try:
            valor = float(order.get("orderAmount") or 0) or None
        except Exception:
            valor = None

        items = order.get("orderItemList") or []

        for item in items:
            sku  = (item.get("variationSku") or item.get("productSku") or "").strip()
            nome = (item.get("productName") or "").strip()
            img  = (item.get("productImg")  or "").strip()
            qtd  = int(item.get("productCount") or 1)
            itens_list.append({
                "order_number": order_num,
                "sku":          sku,
                "product_name": nome,
                "image_url":    img,
                "quantidade":   qtd,
                "cliente":      token,
                "data":         data_arquivo,
            })

        first     = items[0] if items else {}
        qtd_total = sum(int(i.get("productCount") or 1) for i in items) or 1

        # Tenta capturar label_url da API (Shopee/TikTok/Shein retornam URL própria)
        api_label_url = None
        for campo in ("labelUrl", "printUrl", "waybillUrl", "expressLabel", "packageLabel"):
            v = (order.get(campo) or "").strip()
            if v and v.startswith("http"):
                api_label_url = v
                break

        # Data real do pedido: usa payTime > createTime > data_arquivo
        data_pedido = data_arquivo
        for campo_dt in ("payTime", "createTime", "orderTime"):
            v = (order.get(campo_dt) or "")[:10]
            if len(v) == 10 and v[4] == "-":
                data_pedido = v
                break

        if order_num not in pedidos_dict:
            nome_cliente = (
                order.get("receiverName") or order.get("buyerName") or
                order.get("recipientName") or order.get("receiveName") or
                order.get("receiver", {}).get("name") or ""
            ).strip() or None
            if nome_cliente is None and len(pedidos_dict) == 0:
                log(f"[api] 🔍 campos do 1º pedido: {sorted(order.keys())}")
            pedidos_dict[order_num] = {
                "order_number":      order_num,
                "numero_plataforma": str(order.get("orderId") or order.get("extendedId") or "") or None,
                "sku":               (first.get("variationSku") or first.get("productSku") or "").strip(),
                "product_name":      (first.get("productName") or "").strip(),
                "image_url":         (first.get("productImg")  or "").strip(),
                "data":              data_pedido,
                "quantidade":        qtd_total,
                "plataforma":        plataforma,
                "valor":             valor,
                "label_url":         api_label_url,
                "nome_cliente":      nome_cliente,
                "status_upseller":   _API_STATE_LABEL.get(order.get("_state", ""), None),
            }
        else:
            pedidos_dict[order_num]["quantidade"] += qtd_total
            if api_label_url and not pedidos_dict[order_num].get("label_url"):
                pedidos_dict[order_num]["label_url"] = api_label_url

    pedidos = list(pedidos_dict.values())
    multi   = sum(1 for p in pedidos if p.get("quantidade", 1) > 1)
    log(f"Pedidos: {len(pedidos)} ({multi} com 2+ itens, {len(itens_list)} itens)")

    # Recupera label_urls salvas no Supabase de execuções anteriores
    try:
        supa_lbl = create_client(*_supa())
        nums = list(pedidos_dict.keys())
        for i in range(0, len(nums), 100):
            lote = nums[i:i + 100]
            r = supa_lbl.table("pedidos").select("order_number,label_url").in_("order_number", lote).execute()
            for row in (r.data or []):
                on  = row["order_number"]
                url = row.get("label_url")
                if url and on in pedidos_dict:
                    pedidos_dict[on]["label_url"] = url
    except Exception as e:
        log(f"[api] ⚠️ Não foi possível recuperar label_urls do Supabase: {e}")

    # Baixa PDFs com URL original disponível (não a URL do robô)
    com_url = [
        (p["order_number"], p["label_url"]) for p in pedidos
        if p.get("label_url")
        and "127.0.0.1:5001"  not in p["label_url"]
        and "localhost:5001" not in p["label_url"]
    ]
    if com_url:
        log(f"📥 Baixando {len(com_url)} etiqueta(s)...")
        baixados = 0
        for order_num, url in com_url:
            pdf_path = PASTA_ETIQUETAS / f"etiqueta_{order_num}.pdf"
            if not pdf_path.exists():
                try:
                    _urllib.urlretrieve(url, str(pdf_path))
                    _corr = _corrigir_mediabox(pdf_path)
                    if _corr != pdf_path and _corr.exists():
                        _corr.replace(pdf_path)
                    baixados += 1
                except Exception:
                    pass
        if baixados:
            log(f"✅ {baixados} etiqueta(s) salvas em cache")

    # JSON para PCP: usa URL do robô local onde PDF já existe
    pedidos_json = _copy.deepcopy(pedidos)
    for p in pedidos_json:
        on   = p["order_number"]
        lurl = p.get("label_url")
        if lurl and "127.0.0.1:5001" not in lurl and (PASTA_ETIQUETAS / f"etiqueta_{on}.pdf").exists():
            p["label_url"] = f"http://127.0.0.1:5001/etiqueta/{on}"

    arq = PASTA_DADOS / f"lista_{data_arquivo}_{token[:8]}.json"
    arq.write_text(json.dumps(pedidos_json, ensure_ascii=False, indent=2), encoding="utf-8")
    log(f"JSON salvo: {arq.name}")

    # Envia para Supabase (mantém label_url original — fallback em /etiqueta/)
    log("Enviando para Supabase...")
    try:
        supa = create_client(*_supa())
        linhas = [{
            "order_number":      p["order_number"],
            "numero_plataforma": p.get("numero_plataforma"),
            "sku":               p["sku"],
            "product_name":      p["product_name"],
            "image_url":         p["image_url"],
            "data":              p.get("data") or data_arquivo,
            "cliente":           token,
            "quantidade":        p.get("quantidade", 1),
            "plataforma":        p.get("plataforma"),
            "valor":             p.get("valor"),
            "label_url":         p.get("label_url"),
            "nome_cliente":      p.get("nome_cliente"),
            "status_upseller":   p.get("status_upseller"),
        } for p in pedidos]

        if linhas:
            # Preserva data de pedidos já existentes — nunca sobrescreve
            ons2 = [l["order_number"] for l in linhas]
            existentes2: set = set()
            for i in range(0, len(ons2), 50):
                ex2 = supa.table("pedidos").select("order_number").in_("order_number", ons2[i:i+50]).execute()
                existentes2.update(r["order_number"] for r in (ex2.data or []))
            novos2   = [l for l in linhas if l["order_number"] not in existentes2]
            updates2 = [{k: v for k, v in l.items() if k != "data"}
                        for l in linhas if l["order_number"] in existentes2]

            colunas_opcionais = {"plataforma", "valor", "label_url", "numero_plataforma", "nome_cliente"}

            def _upsert_fb2(lote_orig, label):
                excluir: set = set()
                tentativa = lote_orig
                while True:
                    try:
                        supa.table("pedidos").upsert(tentativa, on_conflict="order_number").execute()
                        log(f"✅ {len(tentativa)} pedidos {label} ao Supabase!")
                        return
                    except Exception as e_col:
                        col_faltando = next((c for c in colunas_opcionais - excluir if c in str(e_col)), None)
                        if col_faltando:
                            excluir.add(col_faltando)
                            tentativa = [{k: v for k, v in l.items() if k not in excluir} for l in lote_orig]
                            log(f"⚠️ Coluna '{col_faltando}' ausente — tentando sem ela...")
                        else:
                            raise

            def _update_existentes2(lote):
                for linha in lote:
                    on = linha.get("order_number")
                    if not on:
                        continue
                    campos = {k: v for k, v in linha.items() if k != "order_number"}
                    try:
                        supa.table("pedidos").update(campos).eq("order_number", on).execute()
                    except Exception as e_upd:
                        log(f"⚠️ Erro ao atualizar {on}: {e_upd}")
                log(f"✅ {len(lote)} pedidos atualizados (data preservada)")

            if novos2:
                _upsert_fb2(novos2, "inseridos")
            if updates2:
                _update_existentes2(updates2)
        else:
            log("Nenhum pedido para enviar")
    except Exception as e:
        log(f"Erro Supabase (pedidos): {e}")

    try:
        if itens_list:
            order_numbers = list(pedidos_dict.keys())
            supa.table("pedido_itens").delete().in_("order_number", order_numbers).execute()
            batch = 200
            for i in range(0, len(itens_list), batch):
                supa.table("pedido_itens").insert(itens_list[i:i + batch]).execute()
            log(f"✅ {len(itens_list)} itens enviados a pedido_itens!")
    except Exception as e:
        log(f"Erro Supabase (pedido_itens): {e}")

    atualizar_ultima_execucao(token)
    log(f"✅ Concluído! {len(pedidos)} pedidos | {data_arquivo}")
    return True


# ── Agendador ─────────────────────────────────────────────────────────────────

def _gerar_horarios(hora_inicio: str, hora_fim: str, intervalo_horas: int) -> list:
    h, m   = map(int, hora_inicio.split(":"))
    hf, mf = map(int, hora_fim.split(":"))
    fim_min = hf * 60 + mf
    result  = []
    while (h * 60 + m) <= fim_min:
        result.append(f"{h:02d}:{m:02d}")
        h += intervalo_horas
    return result


_picklist_hoje: set = set()

def _loop_agendador():
    global _execucoes_hoje, _picklist_hoje
    while True:
        time.sleep(30)
        if not CONFIG_FILE.exists():
            continue
        try:
            config   = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
            _DIA_KEYS_WD = ["seg", "ter", "qua", "qui", "sex", "sab", "dom"]
            _dia_wd = _DIA_KEYS_WD[datetime.now().weekday()]
            _hs_wd  = config.get("horarios_semanais", {})
            if _hs_wd:
                horarios = _hs_wd.get(_dia_wd) or []
            else:
                horarios = config.get("horarios", [])
                if not horarios and config.get("hora_inicio"):
                    horarios = _gerar_horarios(
                        config.get("hora_inicio", "07:00"),
                        config.get("hora_fim", "18:00"),
                        int(config.get("intervalo_horas", 1)),
                    )

            if not horarios:
                continue

            agora = datetime.now().strftime("%H:%M")
            hoje  = str(date.today())
            chave = f"{hoje}_{agora}"

            _execucoes_hoje = {k for k in _execucoes_hoje if k.startswith(hoje)}
            _picklist_hoje  = {k for k in _picklist_hoje  if k.startswith(hoje)}

            if agora in horarios and chave not in _execucoes_hoje and not rodando:
                _execucoes_hoje.add(chave)
                _salvar_execucoes(_execucoes_hoje)

                agora_dt     = datetime.now().replace(second=0, microsecond=0)
                hoje_inicio  = agora_dt.replace(hour=0, minute=0)
                ontem        = date.today() - timedelta(days=1)
                ontem_inicio = datetime.combine(ontem, datetime.min.time())
                ontem_fim    = datetime.combine(ontem, datetime.max.time()).replace(microsecond=0)

                # Primeiro horário do dia → carga dupla (ontem + hoje)
                if agora == sorted(horarios)[0]:
                    log(f"⏰ Carga inicial ({agora}): ontem completo + hoje até agora")
                    threading.Thread(
                        target=_rodar_em_thread_duplo,
                        args=(config, ontem_inicio, ontem_fim, hoje_inicio, agora_dt),
                        daemon=True
                    ).start()
                else:
                    log(f"⏰ Atualização ({agora}): hoje até agora")
                    threading.Thread(
                        target=_rodar_em_thread,
                        args=(config, hoje_inicio, agora_dt),
                        daemon=True
                    ).start()
        except Exception:
            pass


# ── Inicia o servidor ─────────────────────────────────────────────────────────
def _limpar_etiquetas_antigas(dias: int = 7):
    """Apaga PDFs de etiquetas com mais de X dias ao iniciar o robô."""
    corte = date.today() - timedelta(days=dias)
    removidos = 0
    for pdf in PASTA_ETIQUETAS.glob("etiqueta_*.pdf"):
        try:
            if date.fromtimestamp(pdf.stat().st_mtime) <= corte:
                pdf.unlink()
                removidos += 1
        except Exception:
            pass
    if removidos:
        log(f"🗑️ {removidos} etiqueta(s) com mais de {dias} dias removidas")


def _loop_captura_etiquetas():
    """Captura etiquetas do UpSeller a cada 60s, com retroativo a cada 5 min."""
    INTERVALO = 60  # 1 minuto
    time.sleep(60)  # aguarda 1 min para o robô inicializar antes da primeira captura
    while True:
        try:
            if CONFIG_FILE.exists() and not rodando and not _pos_import_ativo:
                config = json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
                if config.get("token"):
                    log("━━ Captura automática de etiquetas (1min) ━━")
                    _executar_captura(config, aguardar=False, background=True)
        except Exception as e:
            log(f"[captura-auto] ⚠️ {e}")
        time.sleep(INTERVALO)


async def _verificar_cancelados_playwright(config: dict) -> None:
    """
    Cruza pedidos sem label_url no Supabase com o estado atual no UpSeller.
    - Voided (isVoided=1) → marca status='cancelado' no Supabase
    - Não encontrado em nenhuma fila ativa nem voided → reporta para investigação
    """
    try:
        from playwright.async_api import async_playwright
    except ImportError:
        log("[cancelados] ❌ Playwright não instalado")
        return

    supa = create_client(*_supa())

    try:
        res = supa.table("pedidos").select("order_number, status_upseller") \
            .eq("status", "ativo").is_("label_url", "null").execute()
        pedidos = res.data or []
    except Exception as e:
        log(f"[cancelados] ❌ Erro ao consultar Supabase: {e}")
        return

    if not pedidos:
        log("[cancelados] ℹ️ Nenhum pedido sem etiqueta no Supabase")
        return

    nums_sem_etiqueta: set[str] = {p["order_number"] for p in pedidos if p.get("order_number")}
    log(f"[cancelados] 🔍 {len(nums_sem_etiqueta)} pedido(s) sem etiqueta — verificando no UpSeller...")

    _limpar_crash_chrome()

    async with async_playwright() as pw:
        context = await pw.chromium.launch_persistent_context(
            user_data_dir=str(PASTA_SESSAO),
            headless=True,
            args=["--window-size=1280,900"],
        )
        try:
            page = context.pages[0] if context.pages else await context.new_page()
            await page.goto(
                "https://app.upseller.com/pt/order/in-process",
                wait_until="domcontentloaded",
                timeout=60_000,
            )
            await page.wait_for_timeout(2000)
            if "/login" in page.url:
                log("[cancelados] ⚠️ Sessão expirada — faça login manualmente")
                return

            # 1. Busca TODOS os voided (sem filtro de data)
            log("[cancelados] 📡 Consultando pedidos voided no UpSeller...")
            todos_voided: set[str] = set()
            pg_v = 1
            while True:
                r_v = await page.evaluate(f"""async () => {{
                    const r = await fetch('/api/order/index', {{
                        method:'POST',
                        headers:{{'Content-Type':'application/x-www-form-urlencoded'}},
                        body:'timeType=0&isVoided=1&searchType=0&sortName=1&sortValue=1&pageNum={pg_v}&pageSize=50'
                    }});
                    return await r.json();
                }}""")
                lista_v = (r_v.get("data") or {}).get("list") or []
                total_v = int((r_v.get("data") or {}).get("total") or 0)
                todos_voided.update((o.get("orderNumber") or "").strip() for o in lista_v if o.get("orderNumber"))
                if pg_v * 50 >= total_v or not lista_v:
                    break
                pg_v += 1
            log(f"[cancelados] 📊 {len(todos_voided)} pedido(s) voided no UpSeller")

            # 2. Busca TODOS os pedidos ativos de todas as filas
            log("[cancelados] 📡 Consultando filas ativas no UpSeller...")
            todos_ativos: set[str] = set()
            FILAS_ATIVAS = [
                ("allocate",        "allocateStatus=pending_review"),
                ("in_process",      "labelStatus=success&warehouseType=0"),
                ("invoice_pending", "invoiceStatus=to_issue&isVoided=0"),
                ("to_ship",         ""),
                ("to_pickup",       ""),
            ]
            for state, extra in FILAS_ATIVAS:
                pg_a = 1
                while True:
                    body_a = (
                        f"timeType=0&orderState={state}&isVoided=0&searchType=0"
                        f"&sortName=1&sortValue=1&pageNum={pg_a}&pageSize=50"
                        + (f"&{extra}" if extra else "")
                    )
                    r_a = await page.evaluate(f"""async () => {{
                        const r = await fetch('/api/order/index', {{
                            method:'POST',
                            headers:{{'Content-Type':'application/x-www-form-urlencoded'}},
                            body:'{body_a}'
                        }});
                        return await r.json();
                    }}""")
                    lista_a = (r_a.get("data") or {}).get("list") or []
                    total_a = int((r_a.get("data") or {}).get("total") or 0)
                    todos_ativos.update((o.get("orderNumber") or "").strip() for o in lista_a if o.get("orderNumber"))
                    if pg_a * 50 >= total_a or not lista_a:
                        break
                    pg_a += 1
            log(f"[cancelados] 📊 {len(todos_ativos)} pedido(s) ativos no UpSeller")

        finally:
            await context.close()

    # 3. Cruza resultados
    voided_match    = nums_sem_etiqueta & todos_voided
    ativos_match    = nums_sem_etiqueta & todos_ativos
    desaparecidos   = nums_sem_etiqueta - todos_voided - todos_ativos

    # 4. Atualiza Supabase para voided
    if voided_match:
        try:
            lote = list(voided_match)
            for i in range(0, len(lote), 50):
                supa.table("pedidos") \
                    .update({"status": "cancelado", "status_upseller": "Cancelado"}) \
                    .in_("order_number", lote[i:i+50]) \
                    .eq("status", "ativo").execute()
            log(f"[cancelados] ✅ {len(voided_match)} pedido(s) marcados como cancelado no Supabase")
            for n in sorted(voided_match)[:20]:
                log(f"[cancelados]   ⛔ {n}")
            if len(voided_match) > 20:
                log(f"[cancelados]   ... e mais {len(voided_match)-20}")
        except Exception as e:
            log(f"[cancelados] ❌ Erro ao atualizar Supabase (voided): {e}")
    else:
        log("[cancelados] ✅ Nenhum pedido sem etiqueta está voided no UpSeller")

    # 5. Reporta desaparecidos (não em nenhuma fila, não voided)
    if desaparecidos:
        log(f"[cancelados] ⚠️ {len(desaparecidos)} pedido(s) não encontrados em nenhuma fila do UpSeller (verificar manualmente):")
        for n in sorted(desaparecidos)[:30]:
            log(f"[cancelados]   ? {n}")
        if len(desaparecidos) > 30:
            log(f"[cancelados]   ... e mais {len(desaparecidos)-30}")
    else:
        log("[cancelados] ✅ Todos os pedidos sem etiqueta foram encontrados nas filas ativas")

    log(f"[cancelados] ✅ Concluído — voided: {len(voided_match)}, ativos: {len(ativos_match)}, sem fila: {len(desaparecidos)}")


if __name__ == "__main__":
    import socket as _socket

    # Verifica se já existe uma instância rodando na porta 5001
    _porta_livre = True
    try:
        _s = _socket.socket(_socket.AF_INET, _socket.SOCK_STREAM)
        _s.settimeout(0.5)
        _s.connect(("127.0.0.1", 5001))
        _s.close()
        _porta_livre = False  # conseguiu conectar → já tem instância rodando
    except Exception:
        _porta_livre = True   # porta livre → somos a única instância

    if not _porta_livre:
        # Já há uma instância do robô rodando — abre o browser e encerra
        print("Robô já está rodando — abrindo interface...")
        webbrowser.open("http://127.0.0.1:5001")
        sys.exit(0)

    PASTA_ETIQUETAS.mkdir(exist_ok=True)
    PASTA_PICKLISTS.mkdir(exist_ok=True)
    _limpar_etiquetas_antigas()
    # No Mac, mantém o sistema acordado enquanto o app estiver rodando
    if platform.system() == "Darwin":
        import os
        subprocess.Popen(
            ["caffeinate", "-i", "-w", str(os.getpid())],
            stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL
        )

    def _migrar_colunas():
        try:
            sb = create_client(*_supa())
            sb.rpc("exec_sql", {"sql": "ALTER TABLE pedidos ADD COLUMN IF NOT EXISTS nome_cliente TEXT;"}).execute()
            log("[startup] ✅ Coluna nome_cliente verificada/criada")
        except Exception:
            pass  # exec_sql pode não existir; coluna pode já existir — tudo ok

    threading.Thread(target=_migrar_colunas, daemon=True).start()
    threading.Thread(target=_loop_agendador, daemon=True).start()
    threading.Thread(target=_loop_captura_etiquetas, daemon=True).start()
    # Pré-baixa SumatraPDF no Windows para a primeira impressão ser imediata
    if platform.system() == "Windows":
        threading.Thread(target=_obter_sumatra, daemon=True).start()
    # Abre browser sempre (auto-start via pythonw não tem terminal)
    threading.Timer(1.5, lambda: webbrowser.open("http://127.0.0.1:5001")).start()
    print("=" * 48)
    print("  AdnSys Robo UpSeller")
    print("  Acesse: http://127.0.0.1:5001")
    print("  Ctrl+C para encerrar")
    print("=" * 48)
    app.run(debug=False, port=5001)
